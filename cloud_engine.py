import os as os_native

import sqlite3

import pandas as pd

import threading

import time

import json

import sys

from datetime import datetime, timedelta

import uuid

import shutil

import re

import unicodedata

import ftplib

import tarfile

import bz2

import numpy as np

from concurrent.futures import ThreadPoolExecutor

try:

    import clickhouse_connect

except ImportError:

    pass



from dotenv import load_dotenv

load_dotenv()



def remove_accents(input_str):

    if not input_str: return ""

    nfkd_form = unicodedata.normalize('NFKD', input_str)

    return "".join([c for c in nfkd_form if not unicodedata.combining(c)])



def normalize_name(name):

    """Remove common suffixes to improve matching (JUNIOR, FILHO, etc.)"""

    if not name: return ""

    name = remove_accents(str(name).upper().strip())

    suffixes = [' JUNIOR', ' FILHO', ' NETO', ' SOBRINHO', ' JR', ' SEGUNDO', ' TERCEIRO']

    for sfx in suffixes:

        if name.endswith(sfx):

            name = name[:-len(sfx)].strip()

            break

    return name



class CloudEngine:

    def __init__(self, **kwargs):

        import platform

        self.is_linux = (platform.system() == 'Linux')

        if self.is_linux:

            self.db_carrier = "/var/www/hemn_cloud/hemn_carrier.db"

            self.db_path = "/var/www/hemn_cloud/hemn_cloud.db"

        else:

            self.db_carrier = kwargs.get('db_carrier_path')

            self.db_path = kwargs.get('db_path') or "hemn_cloud.db"

        

        self._init_db()

        self._load_carrier_assets()

        

        # Monitor Cache (Otimização de CPU)

        self._monitor_cache = None

        self._monitor_cache_time = 0

        # V8: Semaforo global de concorrencia pra Enriquecimento PJ.

        # Limite de 2 enrichs simultaneos protege a maquina de saturacao em multi-cliente.

        self._enrich_semaphore = threading.Semaphore(2)

        # Disk breakdown — cache mais longo (60s)

        self._disk_cache = None

        self._disk_cache_time = 0



        # DETALHADO Cache (overview / map / operadoras) — chave -> {ts, data}

        self._detalhado_cache = {}

        # Lock pra serializar queries do DETALHADO (ClickHouse limita 2 simultâneas no user default)

        self._detalhado_query_lock = threading.Lock()



        # Mapa de DDDs por Estado (Brasil)

        self.UF_DDD_MAP = {

            'AC': ['68'], 'AL': ['82'], 'AM': ['92', '97'], 'AP': ['96'],

            'BA': ['71', '73', '74', '75', '77'], 'CE': ['85', '88'],

            'DF': ['61'], 'ES': ['27', '28'], 'GO': ['62', '64'],

            'MA': ['98', '99'], 'MG': ['31', '32', '33', '34', '35', '37', '38'],

            'MS': ['67'], 'MT': ['65', '66'], 'PA': ['91', '93', '94'],

            'PB': ['83'], 'PE': ['81', '87'], 'PI': ['86', '89'],

            'PR': ['41', '42', '43', '44', '45', '46'], 'RJ': ['21', '22', '24'],

            'RN': ['84'], 'RO': ['69'], 'RR': ['95'], 'RS': ['51', '53', '54', '55'],

            'SC': ['47', '48', '49'], 'SE': ['79'], 'SP': ['11', '12', '13', '14', '15', '16', '17', '18', '19'],

            'TO': ['63']

        }



    def _get_ch_client(self):

        """Retorna uma nova instância de cliente ClickHouse (Thread-Safe)"""

        if not self.is_linux: return None

        try:

            return clickhouse_connect.get_client(host='localhost', port=8123, username='default', password='')

        except Exception as e:

            print(f"[ERROR] Falha ao conectar no ClickHouse: {e}")

            return None



    def search_leads(self, search_type, search_term, scope, uf=None, regiao=None):

        search_type = str(search_type).lower().strip()

        client = self._get_ch_client()

        if not client:

            return {"error": "Conexão com ClickHouse indisponível (Ambiente Windows)"}

        

        where_clauses = []

        params = {}



        # Tipo de Busca

        if search_type == 'cpf':

            cpf_clean = ''.join(filter(str.isdigit, search_term))

            where_clauses.append("cpf = {cpf:String}")

            params['cpf'] = cpf_clean

        elif search_type == 'nome':

            where_clauses.append("nome LIKE {nome:String}")

            params['nome'] = f"%{remove_accents(search_term).upper().strip()}%"

        elif search_type == 'telefone':

            tel_clean = ''.join(filter(str.isdigit, search_term))

            where_clauses.append("(tel_fixo1 IN ({tel:String}, {tel_dot:String}) OR celular1 IN ({tel:String}, {tel_dot:String}))")

            params['tel'] = tel_clean

            params['tel_dot'] = tel_clean + '.0'

        

        # Filtros de Escopo

        if scope == 'ESTADO' and uf:

            where_clauses.append("uf = {uf:String}")

            params['uf'] = uf.upper()

        elif scope == 'REGIAO' and regiao:

            where_clauses.append("regiao = {regiao:String}")

            params['regiao'] = regiao.upper()

        

        if not where_clauses:

            return {"leads": [], "count": 0}



        where_str = " AND ".join(where_clauses)

        query = f"SELECT DISTINCT cpf, nome, dt_nascimento, uf, regiao FROM hemn.leads WHERE {where_str} LIMIT 100"

        

        try:

            # Otimização: Forçar uso de apenas 1 núcleo para buscas básicas/unitárias

            result = client.query(query, parameters=params, settings={'max_threads': 1})

            columns = ['cpf', 'nome', 'dt_nascimento', 'uf', 'regiao']

            leads = []

            today = datetime.now()

            for row in result.result_rows:

                lead = dict(zip(columns, row))

                # Calcular idade

                idade = "N/A"

                dt_nasc = lead.get('dt_nascimento', '')

                if dt_nasc and '/' in dt_nasc:

                    try:

                        parts = dt_nasc.split('/')

                        if len(parts) == 3:

                            day, month, year = int(parts[0]), int(parts[1]), int(parts[2])

                            birth_date = datetime(year, month, day)

                            age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))

                            idade = str(age)

                    except Exception:

                        pass

                lead['idade'] = idade

                leads.append(lead)

            

            return {"leads": leads, "count": len(leads)}

        except Exception as e:

            print(f"[ERROR] ClickHouse query failed: {e}")

            return {"error": str(e)}

        finally:

            client.close()



    def get_db_version(self):

        """Busca a versão atual do banco de dados na tabela hemn._metadata"""

        if not self.is_linux: return "Ambiente Windows (Local)"

        client = self._get_ch_client()

        if not client: return "Erro de Conexão (ClickHouse)"

        try:

            res = client.query("SELECT value FROM hemn._metadata WHERE key = 'db_version' LIMIT 1")

            if res.result_rows:

                return str(res.result_rows[0][0])

            return "Versão Desconhecida"

        except:

            return "Tabela _metadata não encontrada"

        finally:

            client.close()



    def _load_carrier_assets(self):

        """Loads prefix tree and operator dictionary from data_assets folder"""

        self.prefix_tree = {} # {prefix: operator_code}

        # Initial hardcoded fallback for most common

        self.anatel_dict = {

            "55320": "VIVO", "55321": "CLARO", "55341": "TIM", "55331": "OI",

            "55312": "ALGAR", "55343": "SERCOMTEL", "55306": "SURF", "55301": "ARQIA",

            "55315": "TELECALL", "55322": "BRISANET"

        }

        

        if self.is_linux:

            base_dir = "/var/www/hemn_cloud/data_assets"

        else:

            base_dir = os_native.path.join(os_native.path.dirname(os_native.path.abspath(__file__)), "data_assets")

            

        prefix_path = os_native.path.join(base_dir, "prefix_anatel.csv")

        dict_path = os_native.path.join(base_dir, "cod_operadora.csv")

        

        # Load Operadora Dictionary from CSV if available

        if os_native.path.exists(dict_path):

            try:

                import csv

                with open(dict_path, mode='r', encoding='latin1') as f:

                    reader = csv.reader(f)

                    for row in reader:

                        if len(row) >= 2: self.anatel_dict[row[0].strip()] = row[1].strip()

            except: pass

            

        # Load Prefix Tree (Standard ANATEL Base)

        if os_native.path.exists(prefix_path):

            try:

                df = pd.read_csv(prefix_path, sep=';', dtype=str)

                if 'number' in df.columns and 'company' in df.columns:

                    for _, row in df.iterrows():

                        self.prefix_tree[row['number'].strip()] = row['company'].strip()

            except: pass



    def get_op_name(self, code):

        """Helper to get operator name from RN1/Anatel code"""

        code = str(code).strip()

        name = self.anatel_dict.get(code, f"OUTRA ({code})" if code else "OUTRA")

        # NORMALIZAÇÃO HEMN (Híbrida para clareza e filtros)

        nu = name.upper().strip()

        if "TELEFONICA" in nu or "VIVO" in nu: return "VIVO / TELEFONICA"

        if "CLARO" in nu: return "CLARO"

        if "TIM" in nu: return "TIM"

        if nu == "OI" or nu.startswith("OI ") or "OI S.A" in nu or "OI MOVEL" in nu or "TELEMAR" in nu: 

            return "OI"

        if "ALGAR" in nu: return "ALGAR"

        if "BRISANET" in nu: return "BRISANET"

        if "TELECALL" in nu: return "TELECALL"

        if "DESKTOP" in nu: return "DESKTOP"

        if "SERCOMTEL" in nu: return "SERCOMTEL"

        return name



    def _init_db(self):

        conn = sqlite3.connect(self.db_path, timeout=30)

        conn.execute("PRAGMA journal_mode=WAL")

        conn.execute("""

            CREATE TABLE IF NOT EXISTS background_tasks (

                id TEXT PRIMARY KEY,

                username TEXT,

                module TEXT,

                status TEXT,

                progress REAL,

                message TEXT,

                result_file TEXT,

                record_count INTEGER,

                filters TEXT,

                hidden INTEGER DEFAULT 0,

                created_at TEXT

            )

        """)

        

        # Migração: Adiciona coluna 'filters' se não existir (para bancos legados)

        try:

            cursor = conn.cursor()

            cursor.execute("PRAGMA table_info(background_tasks)")

            columns = [info[1] for info in cursor.fetchall()]

            if 'filters' not in columns:

                print("[MIGRATION] Adicionando coluna 'filters' em background_tasks")

                conn.execute("ALTER TABLE background_tasks ADD COLUMN filters TEXT")

            if 'hidden' not in columns:

                print("[MIGRATION] Adicionando coluna 'hidden' em background_tasks")

                conn.execute("ALTER TABLE background_tasks ADD COLUMN hidden INTEGER DEFAULT 0")

        except Exception as e:

            print(f"[MIGRATION ERROR] Erro ao migrar background_tasks: {e}")



        conn.execute("""

            CREATE TABLE IF NOT EXISTS asaas_payments (

                id TEXT PRIMARY KEY,

                username TEXT,

                amount REAL,

                credits REAL,

                status TEXT,

                pix_payload TEXT,

                pix_image_base64 TEXT,

                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

                confirmed_at TIMESTAMP

            )

        """)

        conn.execute("""

            CREATE TABLE IF NOT EXISTS system_metadata (

                key TEXT PRIMARY KEY,

                value TEXT,

                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP

            )

        """)

        conn.execute("""

            CREATE TABLE IF NOT EXISTS carrier_update_history (

                id INTEGER PRIMARY KEY AUTOINCREMENT,

                timestamp_utc TEXT NOT NULL,

                started_by TEXT,

                duration_seconds INTEGER,

                total_records INTEGER,

                prev_records INTEGER,

                novos INTEGER,

                removidos INTEGER,

                trocaram_operadora INTEGER,

                status TEXT NOT NULL,

                error_msg TEXT

            )

        """)

        # Migration v2: per-operator breakdown columns (added incrementally — NULL for legacy rows)
        for _col in (
            "novos_vivo", "novos_claro", "novos_tim", "novos_oi", "novos_outras",
            "trocaram_vivo", "trocaram_claro", "trocaram_tim", "trocaram_oi", "trocaram_outras",
        ):
            try:
                conn.execute(f"ALTER TABLE carrier_update_history ADD COLUMN {_col} INTEGER")
            except sqlite3.OperationalError:
                pass

        # DETALHADO snapshot (pré-calculado 1× ao dia, lido instantâneo)

        conn.execute("""

            CREATE TABLE IF NOT EXISTS detalhado_snapshot (

                uf TEXT PRIMARY KEY,

                payload TEXT NOT NULL,

                updated_at TEXT NOT NULL

            )

        """)

        # Solicitações de contato vindas do site institucional
        conn.execute("""
            CREATE TABLE IF NOT EXISTS contact_leads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                full_name TEXT NOT NULL,
                empresa TEXT,
                email TEXT NOT NULL,
                phone TEXT,
                message TEXT NOT NULL,
                origin TEXT DEFAULT 'site',
                status TEXT DEFAULT 'PENDING',
                notes TEXT,
                created_at TEXT NOT NULL,
                contacted_at TEXT,
                contacted_by TEXT,
                ip_origin TEXT,
                user_agent TEXT
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_leads_status_created ON contact_leads(status, created_at DESC)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_leads_created ON contact_leads(created_at DESC)")
        # Rate limit anti-spam pro endpoint público de contato
        conn.execute("""
            CREATE TABLE IF NOT EXISTS contact_lead_rate (
                ip TEXT NOT NULL,
                ts TEXT NOT NULL
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_lead_rate_ip_ts ON contact_lead_rate(ip, ts DESC)")

        conn.commit()

        conn.close()

        # === V8.1: Inicializa carrier_diff_history e faz seed se necessário ===
        try:
            from datetime import datetime as _dt2, timezone as _tz2, timedelta as _td2
            cc = sqlite3.connect(self.db_carrier, timeout=30)
            cc.execute("""
                CREATE TABLE IF NOT EXISTS carrier_diff_history (
                  kind TEXT NOT NULL,
                  telefone TEXT NOT NULL,
                  operadora_id INT,
                  created_at TEXT NOT NULL
                )
            """)
            cc.execute("CREATE INDEX IF NOT EXISTS idx_cdh_created ON carrier_diff_history(created_at)")
            cc.execute("CREATE INDEX IF NOT EXISTS idx_cdh_kind_created ON carrier_diff_history(kind, created_at)")
            # Seed: se history está vazio mas carrier_diff_{novos,portados} existem com dados,
            # popula com timestamp = NOW (assim "hoje" e "semana" já têm dados imediatamente)
            n_hist = cc.execute("SELECT count(*) FROM carrier_diff_history").fetchone()[0]
            if n_hist == 0:
                _sp_tz2 = _tz2(_td2(hours=-3))
                _now_str = _dt2.now(_sp_tz2).strftime('%Y-%m-%d %H:%M:%S')
                for k in ("novos", "portados"):
                    table = f"carrier_diff_{k}"
                    has = cc.execute(
                        "SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,)
                    ).fetchone()
                    if has:
                        cc.execute(
                            f"INSERT INTO carrier_diff_history (kind, telefone, operadora_id, created_at) "
                            f"SELECT ?, telefone, operadora_id, ? FROM {table}",
                            (k, _now_str)
                        )
                cc.commit()
                seeded = cc.execute("SELECT count(*) FROM carrier_diff_history").fetchone()[0]
                if seeded > 0:
                    print(f"[INIT] carrier_diff_history seed inicial: {seeded:,} linhas (ts={_now_str})")
            cc.close()
        except Exception as ex:
            print(f"[INIT] carrier_diff_history init falhou (nao bloqueante): {ex}")



    def _create_task(self, module="ENRICH", username=None, filters=None):

        tid = str(uuid.uuid4())

        created_at = datetime.now().isoformat()

        conn = sqlite3.connect(self.db_path, timeout=30)

        conn.execute("PRAGMA journal_mode=WAL")

        conn.execute(

            "INSERT INTO background_tasks (id, username, module, status, progress, message, filters, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",

            (tid, username, module, "QUEUED", 0, "Aguardando início...", filters, created_at)

        )

        conn.commit()

        conn.close()

        return tid



    def cancel_task(self, tid):

        """

        Cancela uma tarefa de forma IMEDIATA, inclusive no banco de dados.

        """

        # 1. Marcar no SQLite (para o loop do Python parar)

        conn = sqlite3.connect(self.db_path, timeout=30)

        conn.execute("PRAGMA journal_mode=WAL")

        conn.execute("UPDATE background_tasks SET status = 'CANCELLED', message = 'Processo cancelado pelo usuário.' WHERE id = ?", (tid,))

        conn.commit()

        conn.close()

        

        # 2. Matar consulta no ClickHouse (se houver uma rodando com esse ID)

        if self.is_linux:

            try:

                ch = self._get_ch_client()

                if ch:

                    # KILL QUERY async para não travar a resposta da API

                    # Matar tanto a query principal quanto possíveis sub-lotes

                    ch.command(f"KILL QUERY WHERE query_id = '{tid}' OR query_id LIKE '{tid}_%' ASYNC")

                    ch.close()

            except Exception as e:

                print(f"[CANCEL] Erro ao matar query no ClickHouse: {e}")

                

        return True



    def cleanup_all_tasks(self):

        """Cancela todas as tarefas pendentes ou em processamento (Limpeza Global)."""

        try:

            conn = sqlite3.connect(self.db_path, timeout=30)

            conn.execute("PRAGMA journal_mode=WAL")

            # Cancela tarefas que não terminaram

            conn.execute("UPDATE background_tasks SET status = 'CANCELLED', message = 'Cancelado via Limpeza Global pelo Administrador.' WHERE status IN ('QUEUED', 'PROCESSING')")

            conn.commit()

            conn.close()

            

            # Tentar limpar ClickHouse também (opcional mas recomendado)

            if self.is_linux:

                try:

                    ch = self._get_ch_client()

                    if ch:

                        ch.command("KILL QUERY WHERE query_id != '' ASYNC")

                        ch.close()

                except: pass

            return True

        except Exception as e:

            print(f"[CLEANUP] Erro na limpeza global: {e}")

            return False



    def hide_task(self, tid):

        conn = sqlite3.connect(self.db_path, timeout=30)

        conn.execute("PRAGMA journal_mode=WAL")

        conn.execute("UPDATE background_tasks SET hidden = 1 WHERE id = ?", (tid,))

        conn.commit()

        conn.close()

        return True



    def get_user_tasks(self, username):

        conn = sqlite3.connect(self.db_path, timeout=30)

        conn.execute("PRAGMA journal_mode=WAL")

        conn.row_factory = sqlite3.Row

        # Include COMPLETED and FAILED tasks from the last 24 hours for UI persistence

        # Only show tasks that are NOT hidden

        rows = conn.execute(

            "SELECT * FROM background_tasks WHERE username = ? COLLATE NOCASE AND hidden = 0 AND (status IN ('QUEUED', 'PROCESSING') OR (status = 'COMPLETED' AND created_at > datetime('now','-24 hours')) OR (status = 'FAILED' AND created_at > datetime('now','-24 hours'))) ORDER BY created_at DESC", 

            (username,)

        ).fetchall()

        conn.close()

        return [dict(r) for r in rows]



    def _update_task(self, tid, **kwargs):

        if not kwargs: return

        cols = ", ".join([f"{k} = ?" for k in kwargs.keys()])

        conn = sqlite3.connect(self.db_path, timeout=30)

        conn.execute("PRAGMA journal_mode=WAL")

        conn.execute(f"UPDATE background_tasks SET {cols} WHERE id = ?", list(kwargs.values()) + [tid])

        conn.commit()

        conn.close()



    def get_task_status(self, tid):

        conn = sqlite3.connect(self.db_path, timeout=30)

        conn.execute("PRAGMA journal_mode=WAL")

        conn.row_factory = sqlite3.Row

        row = conn.execute("SELECT * FROM background_tasks WHERE id = ?", (tid,)).fetchone()

        conn.close()

        if not row: return {"status": "NOT_FOUND"}

        return dict(row)



    def get_internal_stats(self):

        try:

            conn = sqlite3.connect(self.db_path, timeout=30)

            conn.execute("PRAGMA journal_mode=WAL")

            # Return data in the format expected by the frontend (index_vps.html)

            stats = {

                "tasks": {

                    "active": conn.execute("SELECT COUNT(*) FROM background_tasks WHERE status = 'PROCESSING'").fetchone()[0],

                    "queued": conn.execute("SELECT COUNT(*) FROM background_tasks WHERE status = 'QUEUED'").fetchone()[0],

                    "completed": conn.execute("SELECT COUNT(*) FROM background_tasks WHERE status = 'COMPLETED' AND (date(created_at) = date('now') OR created_at > datetime('now','-24 hours'))").fetchone()[0],

                },

                "enrich_slots_available": 2, # Fallback value

                "recent_activities": []

            }

            

            # Fetch recent activities

            conn.row_factory = sqlite3.Row

            recent = conn.execute("SELECT id, module, status, progress, message, created_at FROM background_tasks ORDER BY created_at DESC LIMIT 10").fetchall()

            stats["recent_activities"] = [dict(r) for r in recent]

            conn.close()



            # Add potential external ingestion progress (DB Update)

            ingest_task = self._get_ingestion_progress()

            if ingest_task:

                stats["recent_activities"].insert(0, ingest_task)



            return stats

        except Exception as e:

            return {

                "tasks": {"active": 0, "queued": 0, "completed": 0},

                "enrich_slots_available": 2,

                "recent_activities": [],

                "error": str(e)

            }



    def _get_ingestion_progress(self):

        """Checks for external ingestion logs and returns a virtual task if active."""

        log_path = "/var/www/hemn_cloud/ingest_march_2026.log"

        if not os_native.path.exists(log_path): return None

        

        try:

            # Check if log was updated recently (last 30 minutes)

            mtime = os_native.path.getmtime(log_path)

            if time.time() - mtime > 1800: return None

            

            with open(log_path, "r", encoding="utf-8", errors="ignore") as f:

                lines = f.readlines()

                if not lines: return None

                

                # Simple progress heuristic: find "Starting <file>..." and "Finished <file>..."

                current_file = "Processando..."

                completed_count = 0

                for line in reversed(lines):

                    if "Starting" in line and "..." in line:

                        current_file = line.split("Starting")[-1].strip().replace("...", "")

                        break

                    if "Finished" in line:

                        completed_count += 1 # This is not precise but gives a hint



                # Calculate progress based on total FILES (40 files in the script)

                total_files = 40

                # Scan full log once to count "Finished" (expensive but 1170 lines is small)

                all_finished = [l for l in lines if "Finished" in l]

                prog = min(99, int((len(all_finished) / total_files) * 100))

                

                return {

                    "module": "DATABASE_UPDATE",

                    "status": "PROCESSING",

                    "progress": prog,

                    "message": f"Atualizando Base Mar/2026: {current_file}",

                    "created_at": datetime.fromtimestamp(mtime).isoformat()

                }

        except:

            return None



    def get_ch_metrics(self):

        client = self._get_ch_client()

        if not client:

            return {"status": "DISCONNECTED", "uptime": "0", "active_queries": 0}

        try:

            res = client.query("SELECT version() as v, uptime() as up, count(*) as q FROM system.processes LIMIT 1")

            row = res.result_rows[0]

            

            # Fetch memory usage

            mem_res = client.query("SELECT value FROM system.asynchronous_metrics WHERE metric = 'MemoryAmount' LIMIT 1")

            mem_total = mem_res.result_rows[0][0] if mem_res.result_rows else 0

            

            mem_used_res = client.query("SELECT value FROM system.metrics WHERE metric = 'MemoryTracking' LIMIT 1")

            mem_used = mem_used_res.result_rows[0][0] if mem_used_res.result_rows else 0



            return {

                "status": "ONLINE",

                "version": str(row[0]),

                "uptime_seconds": int(row[1]),

                "active_queries": int(row[2]),

                "memory_usage_bytes": int(mem_used),

                "memory_total_bytes": int(mem_total),

                "ram_usage_mb": round(mem_used / (1024*1024), 1)

            }

        except Exception as e:

            return {"status": "ERROR", "message": str(e), "uptime_seconds": 0, "active_queries": 0}

        finally:

            client.close()



    def get_cities_by_uf(self, uf):

        """Retorna lista de cidades distintas para uma UF (com cache em memória)."""

        uf = (uf or "").upper().strip()

        if not uf or len(uf) != 2: return []

        if not hasattr(self, "_cities_cache"): self._cities_cache = {}

        if uf in self._cities_cache: return self._cities_cache[uf]

        try:

            ch = self._get_ch_client()

            if not ch: return []

            res = ch.query(

                "SELECT DISTINCT municipio_nome FROM hemn.comercial_pj WHERE uf=%(uf)s AND municipio_nome != '' ORDER BY municipio_nome",

                {"uf": uf}

            )

            cities = [r[0] for r in res.result_rows if r and r[0]]

            self._cities_cache[uf] = cities

            return cities

        except Exception as e:

            print(f"[ERROR] get_cities_by_uf {uf}: {e}")

            return []



    def _path_size_bytes(self, path):

        """Soma rápida do tamanho de uma pasta (sem seguir symlinks). Retorna 0 em erro."""

        import os

        total = 0

        try:

            for dirpath, dirnames, filenames in os.walk(path, followlinks=False):

                for fn in filenames:

                    fp = os.path.join(dirpath, fn)

                    try: total += os.path.getsize(fp)

                    except OSError: pass

        except Exception:

            pass

        return total



    def _get_disk_breakdown(self):

        """Retorna breakdown do uso de disco com cache de 60s.

        Lista de dicts: [{label, gb, pct, color}], + total_gb, used_gb, free_gb."""

        import os, glob, shutil, time

        now = time.time()

        if self._disk_cache and (now - self._disk_cache_time < 60):

            return self._disk_cache



        result = {

            "total_gb": 0, "used_gb": 0, "free_gb": 0, "used_pct": 0,

            "items": []

        }

        try:

            usage = shutil.disk_usage("/")

            total_b = usage.total

            used_b  = usage.used

            free_b  = usage.free

            GB = 1024 ** 3



            # ClickHouse via SQL (mais rápido e confiável que du)

            ch_b = 0

            try:

                ch = self._get_ch_client()

                if ch:

                    res = ch.query("SELECT sum(bytes_on_disk) FROM system.parts WHERE active")

                    if res.result_rows and res.result_rows[0] and res.result_rows[0][0]:

                        ch_b = int(res.result_rows[0][0])

            except Exception:

                ch_b = 0

            # fallback: tamanho da pasta no disco se SQL falhou

            if ch_b == 0 and os.path.exists("/var/lib/clickhouse"):

                ch_b = self._path_size_bytes("/var/lib/clickhouse")



            # SQLite operacionais

            sqlite_b = 0

            for db in glob.glob("/var/www/hemn_cloud/*.db"):

                try: sqlite_b += os.path.getsize(db)

                except OSError: pass



            # Storage (uploads/downloads)

            storage_b = 0

            for sub in ("storage", "downloads", "data_assets", "data_analysis"):

                p = f"/var/www/hemn_cloud/{sub}"

                if os.path.exists(p): storage_b += self._path_size_bytes(p)



            # Logs

            logs_b = 0

            for p in ("/var/log", "/var/www/hemn_cloud/logs"):

                if os.path.exists(p): logs_b += self._path_size_bytes(p)



            categorized = ch_b + sqlite_b + storage_b + logs_b

            others_b = max(0, used_b - categorized)



            def gb(b): return round(b / GB, 1)

            def pct(b): return round(100 * b / total_b, 1) if total_b else 0



            items = [

                {"label": "ClickHouse (Inteligência de Dados)", "gb": gb(ch_b),      "pct": pct(ch_b),      "color": "#3b82f6"},

                {"label": "Banco SQLite (Operacional)",         "gb": gb(sqlite_b),  "pct": pct(sqlite_b),  "color": "#10b981"},

                {"label": "Storage (Uploads/Downloads)",        "gb": gb(storage_b), "pct": pct(storage_b), "color": "#f59e0b"},

                {"label": "Logs",                               "gb": gb(logs_b),    "pct": pct(logs_b),    "color": "#8b5cf6"},

                {"label": "Outros (Sistema/App)",               "gb": gb(others_b),  "pct": pct(others_b),  "color": "#94a3b8"},

                {"label": "Livre",                              "gb": gb(free_b),    "pct": pct(free_b),    "color": "#e5e7eb"},

            ]



            result = {

                "total_gb": gb(total_b),

                "used_gb":  gb(used_b),

                "free_gb":  gb(free_b),

                "used_pct": round(100 * used_b / total_b, 1) if total_b else 0,

                "items": items,

            }

        except Exception:

            pass



        self._disk_cache = result

        self._disk_cache_time = now

        return result



    def _read_cpu_stat(self):

        """Lê snapshot de /proc/stat (linha agregada 'cpu'). Retorna (total, idle, user, system, iowait)."""

        with open("/proc/stat") as f:

            line = f.readline()

        parts = line.split()

        vals = [int(x) for x in parts[1:]]

        user    = vals[0] if len(vals) > 0 else 0

        nice    = vals[1] if len(vals) > 1 else 0

        system  = vals[2] if len(vals) > 2 else 0

        idle    = vals[3] if len(vals) > 3 else 0

        iowait  = vals[4] if len(vals) > 4 else 0

        irq     = vals[5] if len(vals) > 5 else 0

        softirq = vals[6] if len(vals) > 6 else 0

        steal   = vals[7] if len(vals) > 7 else 0

        idle_total = idle + iowait

        non_idle   = user + nice + system + irq + softirq + steal

        total      = idle_total + non_idle

        return total, idle_total, user + nice, system + irq + softirq, iowait



    def _measure_cpu(self, sample_ms=100):

        """Mede CPU agregada em janela de sample_ms ms. Retorna (cpu_pct, user_pct, system_pct, iowait_pct)."""

        import time

        try:

            a_total, a_idle, a_user, a_sys, a_io = self._read_cpu_stat()

            time.sleep(max(sample_ms, 50) / 1000.0)

            b_total, b_idle, b_user, b_sys, b_io = self._read_cpu_stat()

            d_total = b_total - a_total

            if d_total <= 0: return 0.0, 0.0, 0.0, 0.0

            cpu_pct = round(100 * (d_total - (b_idle - a_idle)) / d_total, 1)

            user_pct = round(100 * (b_user - a_user) / d_total, 1)

            sys_pct  = round(100 * (b_sys  - a_sys)  / d_total, 1)

            io_pct   = round(100 * (b_io   - a_io)   / d_total, 1)

            return cpu_pct, user_pct, sys_pct, io_pct

        except Exception:

            return 0.0, 0.0, 0.0, 0.0



    def get_monitor_stats(self):

        """Nova função agregadora de monitoramento para a VPS"""

        import os, shutil, time



        # Otimização: Cache de 5 segundos

        now = time.time()

        if self._monitor_cache and (now - self._monitor_cache_time < 5):

            return self._monitor_cache



        # 1. System Stats (Linux /proc fallback para evitar dependência de psutil)

        sys_stats = {

            "cpu": 0, "ram": 0, "disk": 0,

            "load1": 0.0, "load5": 0.0, "load15": 0.0,

            "cores": 1,

            "cpu_user": 0.0, "cpu_system": 0.0, "cpu_iowait": 0.0,

        }

        try:

            # RAM

            if os.path.exists("/proc/meminfo"):

                mem = {}

                with open("/proc/meminfo", "r") as f:

                    for line in f:

                        k, _, v = line.partition(":")

                        if v: mem[k.strip()] = int(v.strip().split()[0])

                total = mem.get("MemTotal", 1)

                available = mem.get("MemAvailable", mem.get("MemFree", 0))

                sys_stats["ram"] = round(100 - (available / total * 100), 1)



            # Cores

            sys_stats["cores"] = os.cpu_count() or 1



            # Load average

            if os.path.exists("/proc/loadavg"):

                with open("/proc/loadavg", "r") as f:

                    parts = f.read().split()

                    sys_stats["load1"]  = float(parts[0])

                    sys_stats["load5"]  = float(parts[1])

                    sys_stats["load15"] = float(parts[2])



            # CPU real (delta de /proc/stat em janela curta)

            cpu_pct, user_pct, sys_pct, io_pct = self._measure_cpu(sample_ms=100)

            sys_stats["cpu"]        = cpu_pct

            sys_stats["cpu_user"]   = user_pct

            sys_stats["cpu_system"] = sys_pct

            sys_stats["cpu_iowait"] = io_pct



            # Disk

            usage = shutil.disk_usage("/")

            sys_stats["disk"] = round((usage.used / usage.total) * 100, 1)

        except:

            pass



        eng_stats = self.get_internal_stats()

        result = {

            "system": sys_stats,

            "engine": eng_stats,

            "recent_activities": eng_stats.get("recent_activities", []),

            "clickhouse": self.get_ch_metrics(),

            "disk_breakdown": self._get_disk_breakdown(),

            "uptime": 99.9 # Valor informativo de SLA

        }

        

        # Atualizar cache

        self._monitor_cache = result

        self._monitor_cache_time = now

        

        return result



    def count_active_tasks(self, username):

        """Retorna o número de tarefas QUEUED ou PROCESSING de um usuário."""

        if not username: return 0

        try:

            conn = sqlite3.connect(self.db_path, timeout=30)

            conn.execute("PRAGMA journal_mode=WAL")

            count = conn.execute(

                "SELECT COUNT(*) FROM background_tasks WHERE username = ? COLLATE NOCASE AND status IN ('QUEUED', 'PROCESSING')",

                (username,)

            ).fetchone()[0]

            conn.close()

            return count

        except:

            return 0



    def _batch_query(self, sql_template, key_param, values, batch_size=3000, tid=None, base_prog=0, max_prog=0, msg_prefix="", extra_params=None, extra_settings=None):

        """Execute a query with a large IN() list in safe-sized batches with progress tracking."""

        ch_local = self._get_ch_client()

        if not ch_local: return [], []



        all_rows = []

        col_names = []

        total = len(values)

        if total == 0: return [], []



        for i in range(0, total, batch_size):

            # Check for cancellation

            if tid:

                status = self.get_task_status(tid)

                if status.get("status") == "CANCELLED": return [], []



            chunk = values[i:i + batch_size]

            params = {key_param: chunk}

            if extra_params: params.update(extra_params)

            # Otimização v1.8.12: Limitar threads para permitir concorrência de usuários

            _settings = {'query_id': tid, 'max_threads': 1}

            if extra_settings:

                _settings.update(extra_settings)

            res = ch_local.query(sql_template, params, settings=_settings)

            all_rows.extend(res.result_rows)

            # Log names once

            if not col_names:

                print(f"[DEBUG] _batch_query: col_names from res: {getattr(res, 'column_names', 'ATTR_NOT_FOUND')}")

            

            if not col_names and getattr(res, 'column_names', None):

                col_names = list(res.column_names)

            

            if tid and max_prog > base_prog:

                prog = base_prog + int((i / total) * (max_prog - base_prog))

                self._update_task(tid, progress=prog, message=f"{msg_prefix} ({i:,}/{total:,})...")

                # Yield to OS for concurrency in multi-user environments (High Priority v1.8.12)

                time.sleep(0.1) 



        return all_rows, col_names



    def _parse_address_columns(self, row):

        # Simplificado para o exemplo, manter original se possível

        tipo = row.get('tipo_logradouro', '')

        logra = row.get('logradouro', '')

        num = row.get('numero', '')

        comp = row.get('complemento', '')

        bairro = row.get('bairro', '')

        muni = row.get('municipio_nome', '')

        uf = row.get('uf', '')

        cep = row.get('cep', '')

        full_logra = f"{tipo} {logra}".strip()

        return [full_logra, num, comp, bairro, muni, uf, cep]



    def _parse_contact_columns(self, row):

        ddd1 = str(row.get('ddd1', '')).strip()

        tel1 = str(row.get('telefone1', '')).strip()

        ddd2 = str(row.get('ddd2', '')).strip()

        tel2 = str(row.get('telefone2', '')).strip()

        email = str(row.get('correio_eletronico', '')).strip()

        

        def format_tel(d, t):

            t = ''.join(filter(str.isdigit, t))

            d = ''.join(filter(str.isdigit, d))

            if not t: return ""

            # Adicionar 9º dígito se for celular (8 dígitos começando com 6-9)

            if len(t) == 8 and t[0] in '6789':

                t = '9' + t

            return f"{d}{t}"

        

        f1 = format_tel(ddd1, tel1)

        if f1: return [f1, "FIXO/CEL", email]

        f2 = format_tel(ddd2, tel2)

        if f2: return [f2, "FIXO/CEL", email]

        return ["", "", email]



    def _format_phone(self, ddd, tel):

        """Formata telefone com 9o digito quando movel (8 digitos comecando com 6-9).

        Concatena DDD + numero. Retorna string vazia se telefone invalido."""

        d = ''.join(filter(str.isdigit, str(ddd or '')))

        t = ''.join(filter(str.isdigit, str(tel or '')))

        if not t:

            return ""

        # Mobile sem o 9 -> adiciona

        if len(t) == 8 and t[0] in '6789':

            t = '9' + t

        return f"{d}{t}"



    def start_enrich(self, input_file, output_dir, name_col, cpf_col, username=None, perfil="TODOS"):

        print(f"[DEBUG] start_enrich called: input={input_file}, name_col={name_col}, cpf_col={cpf_col}, user={username}, perfil={perfil}")

        fname = os_native.path.basename(input_file)

        f_summary = f"[v2.2.0-PREMIUM] Enriquecer: {fname} (Perfil: {perfil})"

        tid = self._create_task(module="ENRICH", username=username, filters=f_summary)

        threading.Thread(target=self._run_enrich, args=(tid, input_file, output_dir, name_col, cpf_col, perfil), daemon=True).start()

        return tid



    def start_carrier_update(self, username="admin"):

        """Inicia atualização da base de operadoras via FTP"""

        tid = self._create_task(module="CARRIER_UPDATE", username=username, filters="[ADMIN] Atualizar Base Operadoras")

        threading.Thread(target=self._run_carrier_update, args=(tid, username), daemon=True).start()

        return tid



    def get_carrier_status(self):

        """Retorna informações sobre o status da base de operadoras e atualizações disponíveis"""

        # 1. Obter timestamps do banco

        info = {

            "last_check": None,

            "last_ftp": None,

            "last_update": None,

            "update_available": False

        }

        

        try:

            conn = sqlite3.connect(self.db_path, timeout=5)

            conn.row_factory = sqlite3.Row

            # 1. Buscar metadados manuais

            rows = conn.execute("SELECT key, value FROM system_metadata WHERE key LIKE 'last_carrier_%'").fetchall()

            for r in rows:

                if r['key'] == 'last_carrier_check_timestamp': info["last_check"] = r['value']

                if r['key'] == 'last_carrier_ftp_timestamp': info["last_ftp"] = r['value']

                if r['key'] == 'last_carrier_vps_timestamp': info["last_update"] = r['value']

            

            # 2. Fallback: Se não houver timestamp VPS ou ele for antigo, buscar no histórico de tarefas

            # Isso resolve casos onde a atualização terminou mas o metadado falhou por trava de base ou permissão.

            task_row = conn.execute("""

                SELECT created_at 

                FROM background_tasks 

                WHERE module = 'CARRIER_UPDATE' AND (status = 'COMPLETED' OR progress = 100)

                ORDER BY created_at DESC LIMIT 1

            """).fetchone()

            

            if task_row:

                import dateutil.parser

                # Use a data de criação ou atualização da tarefa como fallback (Normalizamos para ISO UTC)

                task_ts = task_row['created_at']

                

                # Se o timestamp da tarefa for mais recente que o metadado manual, usamos ele

                if not info["last_update"] or task_ts > info["last_update"]:

                    info["last_update"] = task_ts

                    print(f"[CARRIER] Usando fallback de tarefa para status: {task_ts}")



            conn.close()

        except Exception as e:

            print(f"[ERROR] Fail to get carrier status: {e}")



        # 2. Se a última verificação foi há mais de 12 horas, verificar agora (Usando UTC para consistência com FTP)

        now = datetime.utcnow()

        should_check = True

        if info["last_check"]:

            try:

                # Robustez: aceita múltiplos formatos de ISO

                import dateutil.parser

                last_dt = dateutil.parser.parse(info["last_check"])

                if (now - last_dt).total_seconds() < 12 * 3600:

                    should_check = False

            except: pass

            

        if should_check:

            print("[CARRIER] Iniciando verificação programada de 12h no FTP...")

            remote_ts = self._check_ftp_carrier_timestamp()

            if remote_ts:

                info["last_check"] = now.isoformat() + "Z"

                info["last_ftp"] = remote_ts

                # Salvar no banco

                try:

                    conn = sqlite3.connect(self.db_path, timeout=5)

                    conn.execute("INSERT OR REPLACE INTO system_metadata (key, value) VALUES ('last_carrier_check_timestamp', ?)", (info["last_check"],))

                    conn.execute("INSERT OR REPLACE INTO system_metadata (key, value) VALUES ('last_carrier_ftp_timestamp', ?)", (info["last_ftp"],))

                    conn.commit()

                    conn.close()

                except: pass



        # 3. Comparar FTP vs VPS (Usando string robusta ou objeto datetime)

        if info["last_ftp"] and info["last_update"]:

            try:

                # Adicionamos uma pequena folga de 60 segundos para evitar problemas de arredondamento.

                # Como trabalhamos com strings ISO ordenáveis, a comparação direta funciona.

                if info["last_ftp"] > info["last_update"]:

                    # Verificação Final de Robustez: Se a diferença for menor que 24 horas, ignoramos o alerta.

                    # Isso evita falsos positivos constantes se o FTP for atualizado logo antes do VPS terminar.

                    import dateutil.parser

                    ftp_dt = dateutil.parser.parse(info["last_ftp"])

                    vps_dt = dateutil.parser.parse(info["last_update"])

                    

                    if ftp_dt.tzinfo: ftp_dt = ftp_dt.replace(tzinfo=None)

                    if vps_dt.tzinfo: vps_dt = vps_dt.replace(tzinfo=None)

                    

                    if (ftp_dt - vps_dt).total_seconds() > 3600: # Mais de 1h de diferença real

                        info["update_available"] = True

            except: pass

        elif info["last_ftp"] and not info["last_update"]:

            # Se nunca atualizamos mas tem no FTP, sinaliza disponível

            info["update_available"] = True



        return info



    def get_carrier_history(self, limit=20):

        """Retorna histórico de atualizações da base de operadoras (mais recentes primeiro)"""

        try:

            conn = sqlite3.connect(self.db_path, timeout=5)

            conn.row_factory = sqlite3.Row

            rows = conn.execute(

                "SELECT id, timestamp_utc, started_by, duration_seconds, total_records, prev_records, novos, removidos, trocaram_operadora, status, error_msg FROM carrier_update_history ORDER BY id DESC LIMIT ?",

                (int(limit),)

            ).fetchall()

            conn.close()

            return [dict(r) for r in rows]

        except Exception as e:

            print(f"[ERROR] get_carrier_history: {e}")

            return []



    def export_carrier_diff(self, kind="novos", output_dir=None, period="hoje"):

        """Exporta XLSX com a lista de telefones do diff.

        kind='novos'    -> telefones que apareceram no diff (carrier_diff_history kind='novos')

        kind='portados' -> telefones que trocaram de operadora (carrier_diff_history kind='portados')

        period='hoje'   -> filtra apenas registros do dia atual (fuso SP)

        period='semana' -> filtra registros da semana corrente (segunda-domingo, fuso SP)

        Fallback: se carrier_diff_history nao tem dados pra janela, usa carrier_diff_{kind} direto.

        Retorna path do arquivo XLSX gerado, ou None se nao houver dados.

        """

        from datetime import datetime as _dt, timezone as _tz, timedelta as _td

        if kind not in ("novos", "portados"):

            raise ValueError("kind deve ser 'novos' ou 'portados'")

        if period not in ("hoje", "semana"):

            period = "hoje"

        # Calcula janela em fuso SP

        _sp_tz = _tz(_td(hours=-3))

        _now_sp = _dt.now(_sp_tz)

        if period == "hoje":

            _start = _now_sp.replace(hour=0, minute=0, second=0, microsecond=0)

        else:  # semana = segunda-feira da semana corrente até agora

            _monday = _now_sp - _td(days=_now_sp.weekday())

            _start = _monday.replace(hour=0, minute=0, second=0, microsecond=0)

        _start_str = _start.strftime('%Y-%m-%d %H:%M:%S')

        rows = []

        try:

            conn = sqlite3.connect(self.db_carrier, timeout=30)

            # Tenta primeiro o history (com janela)

            has_history = conn.execute(

                "SELECT name FROM sqlite_master WHERE type='table' AND name='carrier_diff_history'"

            ).fetchone()

            if has_history:

                rows = conn.execute(

                    "SELECT telefone, operadora_id FROM carrier_diff_history "

                    "WHERE kind = ? AND created_at >= ? ORDER BY operadora_id",

                    (kind, _start_str)

                ).fetchall()

                # Dedupe por telefone (mesmo numero pode aparecer em multiplos updates da semana)

                seen = set()

                deduped = []

                for tel, opid in rows:

                    if tel not in seen:

                        seen.add(tel)

                        deduped.append((tel, opid))

                rows = deduped

            # Fallback: se history vazio (primeiro uso), usa tabela atual

            if not rows:

                table = "carrier_diff_novos" if kind == "novos" else "carrier_diff_portados"

                has_table = conn.execute(

                    "SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,)

                ).fetchone()

                if has_table:

                    rows = conn.execute(

                        f"SELECT telefone, operadora_id FROM {table} ORDER BY operadora_id"

                    ).fetchall()

            conn.close()

        except Exception as e:

            print(f"[ERROR] export_carrier_diff: {e}")

            return None

        if not rows:

            return None

        # Resolve nome da operadora

        df_rows = []

        for tel, opid in rows:

            df_rows.append({

                "Telefone": tel,

                "Operadora": self.get_op_name(opid) or "DESCONHECIDA",

                "Operadora_ID": opid,

            })

        df = pd.DataFrame(df_rows)

        out_dir = output_dir or "/var/www/hemn_cloud/uploads"

        os_native.makedirs(out_dir, exist_ok=True)

        stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

        label = "Novos_Numeros" if kind == "novos" else "Numeros_Portados"

        period_label = "Hoje" if period == "hoje" else "Semana"

        out_file = os_native.path.join(out_dir, f"{label}_{period_label}_{stamp}.xlsx")

        df.to_excel(out_file, index=False, engine="openpyxl")

        return out_file



    def get_carrier_aggregates(self, period="hoje"):

        """Agrega carrier_update_history por janela em fuso America/Sao_Paulo.

        period='hoje'   → SP date == hoje (SP)

        period='semana' → SP date entre segunda e domingo da semana corrente (SP)

        Retorna dict com totais e breakdown por operadora.

        """

        period = (period or "hoje").lower()

        if period not in ("hoje", "semana"):

            period = "hoje"

        # Janela em SP (UTC-3, sem DST)

        sp_now = datetime.utcnow() - timedelta(hours=3)

        sp_today = sp_now.date()

        if period == "hoje":

            d_from = d_to = sp_today

            label = "hoje"

        else:

            # Semana ISO: segunda → domingo

            weekday = sp_today.weekday()  # 0=segunda, 6=domingo

            d_from = sp_today - timedelta(days=weekday)

            d_to   = d_from + timedelta(days=6)

            label = "esta semana"

        empty = {

            "period": period,

            "label": label,

            "from": d_from.isoformat(),

            "to": d_to.isoformat(),

            "portados": {"total": 0, "breakdown": {"VIVO": 0, "CLARO": 0, "TIM": 0, "OI": 0, "OUTRAS": 0}, "has_breakdown": False},

            "novos":    {"total": 0, "breakdown": {"VIVO": 0, "CLARO": 0, "TIM": 0, "OI": 0, "OUTRAS": 0}, "has_breakdown": False},

        }

        # === Fonte primária: carrier_update_history (timestamps REAIS de cada refresh) ===

        try:

            conn = sqlite3.connect(self.db_path, timeout=5)

            row = conn.execute(

                "SELECT "

                "COALESCE(SUM(novos),0)             AS sum_novos, "

                "COALESCE(SUM(trocaram_operadora),0) AS sum_trocou, "

                "SUM(novos_vivo)    AS nv, SUM(novos_claro) AS nc, SUM(novos_tim) AS nt, SUM(novos_oi) AS noi, SUM(novos_outras) AS nou, "

                "SUM(trocaram_vivo) AS tv, SUM(trocaram_claro) AS tc, SUM(trocaram_tim) AS tt, SUM(trocaram_oi) AS toi, SUM(trocaram_outras) AS tou, "

                "SUM(CASE WHEN novos_vivo IS NULL THEN 0 ELSE 1 END)    AS bd_n_rows, "

                "SUM(CASE WHEN trocaram_vivo IS NULL THEN 0 ELSE 1 END) AS bd_t_rows "

                "FROM carrier_update_history "

                "WHERE status = 'SUCCESS' "

                "AND date(timestamp_utc, '-3 hours') BETWEEN ? AND ?",

                (d_from.isoformat(), d_to.isoformat())

            ).fetchone()

            conn.close()

        except Exception as e:

            print(f"[ERROR] get_carrier_aggregates: {e}")

            return empty

        if not row:

            return empty

        sum_novos, sum_trocou = int(row[0] or 0), int(row[1] or 0)

        nv, nc, nt, noi, nou = row[2], row[3], row[4], row[5], row[6]

        tv, tc, tt, toi, tou = row[7], row[8], row[9], row[10], row[11]

        bd_n_rows, bd_t_rows = int(row[12] or 0), int(row[13] or 0)

        # Fallback "hoje": se nao houve refresh hoje, mostra o ultimo refresh com label especial

        is_fallback = False

        fallback_date = None

        if period == "hoje" and sum_novos == 0 and sum_trocou == 0:

            try:

                conn = sqlite3.connect(self.db_path, timeout=5)

                last = conn.execute(

                    "SELECT date(timestamp_utc, '-3 hours') AS d FROM carrier_update_history "

                    "WHERE status = 'SUCCESS' ORDER BY id DESC LIMIT 1"

                ).fetchone()

                if last and last[0]:

                    fallback_date = last[0]

                    row = conn.execute(

                        "SELECT "

                        "COALESCE(SUM(novos),0), COALESCE(SUM(trocaram_operadora),0), "

                        "SUM(novos_vivo), SUM(novos_claro), SUM(novos_tim), SUM(novos_oi), SUM(novos_outras), "

                        "SUM(trocaram_vivo), SUM(trocaram_claro), SUM(trocaram_tim), SUM(trocaram_oi), SUM(trocaram_outras), "

                        "SUM(CASE WHEN novos_vivo IS NULL THEN 0 ELSE 1 END), "

                        "SUM(CASE WHEN trocaram_vivo IS NULL THEN 0 ELSE 1 END) "

                        "FROM carrier_update_history "

                        "WHERE status = 'SUCCESS' AND date(timestamp_utc, '-3 hours') = ?",

                        (fallback_date,)

                    ).fetchone()

                    if row:

                        sum_novos, sum_trocou = int(row[0] or 0), int(row[1] or 0)

                        nv, nc, nt, noi, nou = row[2], row[3], row[4], row[5], row[6]

                        tv, tc, tt, toi, tou = row[7], row[8], row[9], row[10], row[11]

                        bd_n_rows, bd_t_rows = int(row[12] or 0), int(row[13] or 0)

                        is_fallback = True

                        try:

                            from datetime import date as _date

                            fb_d = _date.fromisoformat(fallback_date)

                            label = f"última atualização · {fb_d.strftime('%d/%m')}"

                            d_from = d_to = fallback_date

                        except Exception:

                            label = f"última atualização · {fallback_date}"

                            d_from = d_to = fallback_date

                conn.close()

            except Exception as e:

                print(f"[WARN] aggregates fallback: {e}")

        novos_has = bd_n_rows > 0

        trocou_has = bd_t_rows > 0

        return {

            "period": period,

            "label": label,

            "from": str(d_from),

            "to": str(d_to),

            "is_fallback": is_fallback,

            "portados": {

                "total": sum_trocou,

                "breakdown": {

                    "VIVO": int(tv or 0), "CLARO": int(tc or 0), "TIM": int(tt or 0),

                    "OI": int(toi or 0), "OUTRAS": int(tou or 0),

                },

                "has_breakdown": trocou_has,

            },

            "novos": {

                "total": sum_novos,

                "breakdown": {

                    "VIVO": int(nv or 0), "CLARO": int(nc or 0), "TIM": int(nt or 0),

                    "OI": int(noi or 0), "OUTRAS": int(nou or 0),

                },

                "has_breakdown": novos_has,

            },

        }



    def _check_ftp_carrier_timestamp(self):

        """Acessa o FTP apenas para ler o timestamp de modificação"""

        host = os_native.getenv("FTP_HOST", "ftp.portabilidadecelular.com")

        port = int(os_native.getenv("FTP_PORT", 2157))

        user = os_native.getenv("FTP_USER", "MAYK")

        passwd = os_native.getenv("FTP_PASS", "Mayk@2025")

        filename = "portabilidade.tar.bz2"

        try:

            ftp = ftplib.FTP()

            ftp.connect(host, port, timeout=10)

            ftp.login(user, passwd)

            timestamp_raw = ftp.voidcmd(f"MDTM {filename}").split()[1]

            dt = datetime.strptime(timestamp_raw, "%Y%m%d%H%M%S")

            ftp.quit()

            return dt.isoformat() + "Z"

        except Exception as e:

            print(f"[ERROR] Fail to check FTP timestamp: {e}")

            try: ftp.quit()

            except: pass

            return None



    def _run_carrier_update(self, tid, started_by="admin"):

        import sys

        _carrier_started_at = time.time()

        print(f"[CRITICAL DEBUG] _run_carrier_update started for TID: {tid}")

        sys.stdout.flush()

        try:

            # Definir caminhos absolutos para evitar erros de CWD

            base_dir = "/var/www/hemn_cloud"

            local_zip = os_native.path.join(base_dir, "portabilidade.tar.bz2")

            print(f"[CRITICAL DEBUG] local_zip set to: {local_zip}")

            sys.stdout.flush()

            

            host = os_native.getenv("FTP_HOST", "ftp.portabilidadecelular.com")

            port = int(os_native.getenv("FTP_PORT", 2157))

            user = os_native.getenv("FTP_USER", "MAYK")

            passwd = os_native.getenv("FTP_PASS", "Mayk@2025")

            filename = "portabilidade.tar.bz2"

            

            print(f"[CRITICAL DEBUG] Attempting to update task status to 5%...")

            self._update_task(tid, progress=5, message="Baixando base de dados (CURL)...")

            print(f"[CRITICAL DEBUG] Task status updated to 5%.")

            

            # Tentar obter o tamanho do arquivo para progresso

            size = 467616948 # Default

            try:

                print(f"[CRITICAL DEBUG] Connecting to FTP for size check...")

                ftp_size = ftplib.FTP()

                ftp_size.connect(host, port, timeout=10)

                ftp_size.login(user, passwd)

                size = ftp_size.size(filename)

                ftp_size.quit()

                print(f"[CRITICAL DEBUG] FTP size obtained: {size}")

            except Exception as fe:

                print(f"[CRITICAL DEBUG] FTP size fail: {fe}")

            

            self._update_task(tid, progress=5, message="Baixando base de dados (CURL)...")

            

            # Usar curl para máxima estabilidade no download FTP

            import subprocess

            import os

            

            # Garantir que o diretório atual é o correto

            os_native.chdir("/var/www/hemn_cloud")

            

            # Download usando CURL (mais robusto)

            cmd = [

                "/usr/bin/curl", "-u", f"{user}:{passwd}",

                f"ftp://{host}:{port}/{filename}",

                "-o", local_zip,

                "--silent", "--show-error",

                "--retry", "3", "--retry-delay", "5"

            ]

            print(f"[CRITICAL DEBUG] CurL command prepared: {' '.join(cmd)}")

            sys.stdout.flush()

            process = subprocess.Popen(cmd, stderr=subprocess.PIPE, stdout=subprocess.PIPE)

            

            # Monitorar progresso (simples via tamanho do arquivo em disco)

            start_time = time.time()

            print(f"[CRITICAL DEBUG] Monitoring download of {local_zip}...")

            sys.stdout.flush()

            

            while process.poll() is None:

                if os_native.path.exists(local_zip):

                    current_size = os_native.path.getsize(local_zip)

                    p = int((current_size / size) * 45) + 5

                    if p > 50: p = 50

                    self._update_task(tid, progress=p, message=f"Baixando base... ({current_size//1024//1024}MB)")

                    if int(time.time() - start_time) % 10 == 0:

                        print(f"[CRITICAL DEBUG] Download Progress: {p}% ({current_size} bytes)")

                        sys.stdout.flush()

                time.sleep(1)

            

            print(f"[CRITICAL DEBUG] Curl finished. Return code: {process.returncode}")

            sys.stdout.flush()

            

            if process.returncode != 0:

                stdout, stderr = process.communicate()

                print(f"[CRITICAL ERROR] Download failed: {stderr.decode() if stderr else 'Unknown error'}")

                sys.stdout.flush()

                raise Exception(f"Download falhou: {stderr.decode() if stderr else 'Unknown error'}")

            

            if not os_native.path.exists(local_zip) or os_native.path.getsize(local_zip) < 10000000: # 10MB min

                raise Exception("Arquivo inexistente ou muito pequeno após download.")

            

            self._update_task(tid, progress=55, message="Extraindo dados (bzip2)...")

            print(f"[CRITICAL DEBUG] Extracting {local_zip}...")

            sys.stdout.flush()

            

            import tarfile

            extracted_file = None

            with tarfile.open(local_zip, "r:bz2") as tar:

                tar.extractall(path="/var/www/hemn_cloud")

                for member in tar.getmembers():

                    if member.name.endswith(".csv"):

                        extracted_file = os_native.path.join("/var/www/hemn_cloud", member.name)

                        break

            

            if not extracted_file or not os_native.path.exists(extracted_file):

                print(f"[CRITICAL ERROR] Extracted CSV NOT FOUND after bz2 extraction.")

                sys.stdout.flush()

                raise Exception("Arquivo extraído não encontrado!")

            

            print(f"[CRITICAL DEBUG] Extraction SUCCESS. File: {extracted_file}")

            sys.stdout.flush()



            self._update_task(tid, progress=60, message="Iniciando ingestão SQLite...")

            

            # Ingestão no SQLite hemn_carrier.db

            conn = sqlite3.connect(self.db_carrier)

            # === V8 FAST INGEST: portabilidade_new e descartavel ate o swap; sem journal/fsync ate la. ===

            # Se o VPS cair durante INSERT/INDEX, portabilidade_new e dropada e refeita na proxima execucao;

            # portabilidade (producao) so muda no RENAME final, com WAL ja restaurado.

            conn.execute("PRAGMA journal_mode=OFF")     # zero overhead de journal

            conn.execute("PRAGMA synchronous=OFF")      # zero fsync

            conn.execute("PRAGMA cache_size=-1048576")  # 1 GB cache (era 64MB)

            conn.execute("PRAGMA temp_store=MEMORY")    # temp tables em RAM

            try:

                conn.execute("PRAGMA mmap_size=8589934592")  # 8 GB mmap

            except Exception: pass

            conn.execute("DROP TABLE IF EXISTS portabilidade_new")

            conn.execute("CREATE TABLE portabilidade_new (telefone TEXT, operadora_id INTEGER)")

            

            batch = []

            count = 0

            ingested = 0

            print(f"[CRITICAL DEBUG] Starting ingestion from {extracted_file}...")

            sys.stdout.flush()

            

            with open(extracted_file, 'r') as f:

                for line in f:

                    parts = line.strip().split(';') # Formato: ID;TELEFONE;OP_ID;DATA

                    if len(parts) >= 3:

                        # parts[1] is phone, parts[2] is operator_id

                        batch.append((parts[1], parts[2]))

                        count += 1

                        ingested += 1

                    

                    if len(batch) >= 50000:

                        conn.executemany("INSERT INTO portabilidade_new VALUES (?,?)", batch)

                        batch = []

                        p = 60 + int((ingested / 50000000) * 35) # Estimativa de 50M records

                        if p > 95: p = 95

                        self._update_task(tid, progress=p, message=f"Ingerindo dados... ({ingested//1000000}M)")

                        if ingested % 1000000 == 0:

                            print(f"[CRITICAL DEBUG] Ingested: {ingested} records")

                            sys.stdout.flush()

            

            if batch:

                conn.executemany("INSERT INTO portabilidade_new VALUES (?,?)", batch)

            

            print(f"[CRITICAL DEBUG] Ingestion finished. Total: {ingested} records. Optimizing...")

            sys.stdout.flush()

            self._update_task(tid, progress=96, message="Finalizando e otimizando base...")

            

            conn.execute("CREATE INDEX IF NOT EXISTS idx_port_tel_new ON portabilidade_new (telefone)")

            # === VALIDATION GATES (antes de promover a nova tabela) ===

            new_count = conn.execute("SELECT count(*) FROM portabilidade_new").fetchone()[0]

            new_distinct_ops = conn.execute("SELECT count(DISTINCT operadora_id) FROM portabilidade_new").fetchone()[0]

            existing_count = 0

            existing_table = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='portabilidade'").fetchone()

            if existing_table:

                try:

                    existing_count = conn.execute("SELECT count(*) FROM portabilidade").fetchone()[0]

                except Exception:

                    existing_count = 0

            print(f"[CARRIER VALIDATION] new_count={new_count} existing_count={existing_count} new_distinct_ops={new_distinct_ops}")

            sys.stdout.flush()

            def _fail(reason):

                conn.execute("DROP TABLE IF EXISTS portabilidade_new")

                conn.commit()

                conn.close()

                raise RuntimeError(f"Validacao falhou: {reason}")

            if new_count == 0:

                _fail("portabilidade_new esta vazia")

            if new_distinct_ops < 5:

                _fail(f"apenas {new_distinct_ops} operadoras distintas (esperado >=5)")

            if existing_count > 0 and new_count < existing_count * 0.8:

                _fail(f"novo count {new_count} e <80% do existente {existing_count}")

            ic = conn.execute("PRAGMA quick_check").fetchone()

            if ic and ic[0] != "ok":

                _fail(f"quick_check retornou {ic[0]}")

            print("[CARRIER VALIDATION] Todos os gates passaram. Promovendo nova tabela.")

            sys.stdout.flush()

            # === V8: Restaura durabilidade ANTES do swap (RENAME na producao precisa de WAL+fsync). ===

            try:

                conn.commit()

                conn.execute("PRAGMA journal_mode=WAL")

                conn.execute("PRAGMA synchronous=NORMAL")

                print("[CARRIER] Modo durabilidade restaurado (WAL + synchronous=NORMAL).")

                sys.stdout.flush()

            except Exception as pe:

                print(f"[CARRIER] Falha ao restaurar PRAGMAs (nao critico): {pe}")

            # === ATOMIC SWAP COM BACKUP (preserva tabela anterior) ===

            stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

            backup_table = f"portabilidade_bak_{stamp}"

            if existing_table:

                conn.execute("DROP INDEX IF EXISTS idx_port_tel")

                conn.execute(f"ALTER TABLE portabilidade RENAME TO {backup_table}")

            conn.execute("ALTER TABLE portabilidade_new RENAME TO portabilidade")

            conn.execute("DROP INDEX IF EXISTS idx_port_tel_new")

            conn.execute("CREATE INDEX IF NOT EXISTS idx_port_tel ON portabilidade (telefone)")

            # === CLEANUP: manter apenas os 2 backups mais recentes ===

            old_backups = conn.execute(

                "SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'portabilidade_bak_%' ORDER BY name DESC"

            ).fetchall()

            for (bak_name,) in old_backups[2:]:

                conn.execute(f"DROP TABLE IF EXISTS {bak_name}")

                print(f"[CARRIER UPDATE] Backup antigo removido: {bak_name}")

                sys.stdout.flush()

            print(f"[CARRIER UPDATE] Swap concluido. Backup preservado: {backup_table if existing_table else '(primeira execucao)'}")

            sys.stdout.flush()



            # === V8: Commit do swap e fecha conexao principal ANTES da background. ===

            conn.commit()

            try:

                conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")

            except Exception: pass

            conn.close()



            # Limpeza dos arquivos temporarios (CSV/tar.bz2)

            if os_native.path.exists(local_zip): os_native.remove(local_zip)

            if extracted_file and os_native.path.exists(extracted_file): os_native.remove(extracted_file)



            # Atualiza metadado global imediatamente (frontend ja sabe que a base e nova)

            try:

                meta = sqlite3.connect(self.db_path, timeout=10)

                now_iso = datetime.utcnow().isoformat() + "Z"

                meta.execute("INSERT OR REPLACE INTO system_metadata (key, value) VALUES ('last_carrier_vps_timestamp', ?)", (now_iso,))

                meta.commit()

                meta.close()

                print(f"[CRITICAL DEBUG] Global metadata updated. last_carrier_vps_timestamp: {now_iso}")

            except Exception as me:

                print(f"[ERROR] Fail to update carrier metadata: {me}")



            # === MARCA TASK COMPLETED IMEDIATAMENTE: DIFF/breakdown/history vao pra background ===

            self._update_task(tid, progress=100, message="Base atualizada! Calculando diff em background...", status="COMPLETED")

            print(f"[CRITICAL DEBUG] CARRIER UPDATE swap done — TID {tid} marcada COMPLETED. DIFF rodando em background.")

            sys.stdout.flush()



            # === BACKGROUND: DIFF + breakdown + history INSERT + DETALHADO refresh ===

            _bg_db_carrier = self.db_carrier

            _bg_db_path = self.db_path

            _bg_existing_table = existing_table

            _bg_backup_table = backup_table

            _bg_new_count = new_count

            _bg_existing_count = existing_count

            _bg_started_by = started_by

            _bg_carrier_started_at = _carrier_started_at

            _bg_get_op_name = self.get_op_name

            _bg_refresh_detalhado = self.refresh_detalhado_snapshot



            def _carrier_post_completion_bg():

                try:

                    print(f"[CARRIER BG] Iniciando DIFF + breakdown em background para TID {tid}...")

                    sys.stdout.flush()

                    bg_conn = sqlite3.connect(_bg_db_carrier, timeout=120)

                    bg_conn.execute("PRAGMA journal_mode=WAL")

                    bg_conn.execute("PRAGMA synchronous=NORMAL")

                    bg_conn.execute("PRAGMA cache_size=-262144")  # 256 MB

                    bg_conn.execute("PRAGMA temp_store=MEMORY")

                    diff_novos = 0

                    diff_removidos = 0

                    diff_trocaram = 0

                    novos_bd = {"VIVO": 0, "CLARO": 0, "TIM": 0, "OI": 0, "OUTRAS": 0}

                    trocou_bd = {"VIVO": 0, "CLARO": 0, "TIM": 0, "OI": 0, "OUTRAS": 0}

                    breakdown_ok = False

                    if _bg_existing_table and _bg_backup_table:

                        try:

                            print("[CARRIER DIFF BG] Calculando diff (EXCEPT em telefone/operadora_id)...")

                            sys.stdout.flush()

                            _diff_t0 = time.time()

                            bg_conn.execute("DROP TABLE IF EXISTS _novos_tels")

                            bg_conn.execute(

                                f"CREATE TEMP TABLE _novos_tels AS SELECT telefone FROM portabilidade EXCEPT SELECT telefone FROM {_bg_backup_table}"

                            )

                            diff_novos = bg_conn.execute("SELECT count(*) FROM _novos_tels").fetchone()[0]

                            diff_removidos = bg_conn.execute(

                                f"SELECT count(*) FROM (SELECT telefone FROM {_bg_backup_table} EXCEPT SELECT telefone FROM portabilidade)"

                            ).fetchone()[0]

                            bg_conn.execute("DROP TABLE IF EXISTS _changed_pairs")

                            bg_conn.execute(

                                f"CREATE TEMP TABLE _changed_pairs AS SELECT telefone, operadora_id FROM portabilidade EXCEPT SELECT telefone, operadora_id FROM {_bg_backup_table}"

                            )

                            new_minus_old_tuples = bg_conn.execute("SELECT count(*) FROM _changed_pairs").fetchone()[0]

                            diff_trocaram = max(0, new_minus_old_tuples - diff_novos)

                            print(f"[CARRIER DIFF BG] +{diff_novos:,} novos | -{diff_removidos:,} removidos | {diff_trocaram:,} trocaram operadora (em {time.time()-_diff_t0:.1f}s)")

                            sys.stdout.flush()

                            try:

                                _bd_t0 = time.time()

                                bg_conn.execute("CREATE INDEX IF NOT EXISTS _idx_novos_tels ON _novos_tels(telefone)")

                                novos_by_op = dict(bg_conn.execute(

                                    "SELECT c.operadora_id, COUNT(*) FROM _changed_pairs c "

                                    "INNER JOIN _novos_tels n ON c.telefone = n.telefone "

                                    "GROUP BY c.operadora_id"

                                ).fetchall())

                                trocou_by_op = dict(bg_conn.execute(

                                    "SELECT c.operadora_id, COUNT(*) FROM _changed_pairs c "

                                    "LEFT JOIN _novos_tels n ON c.telefone = n.telefone "

                                    "WHERE n.telefone IS NULL "

                                    "GROUP BY c.operadora_id"

                                ).fetchall())

                                def _bucket_op(opid):

                                    norm = (_bg_get_op_name(opid) or "").upper()

                                    if "VIVO" in norm or "TELEFONICA" in norm: return "VIVO"

                                    if "CLARO" in norm or "EMBRATEL" in norm: return "CLARO"

                                    if "TIM" in norm: return "TIM"

                                    if norm == "OI" or norm.startswith("OI ") or "TELEMAR" in norm: return "OI"

                                    return "OUTRAS"

                                for opid, cnt in novos_by_op.items():

                                    novos_bd[_bucket_op(opid)] += int(cnt or 0)

                                for opid, cnt in trocou_by_op.items():

                                    trocou_bd[_bucket_op(opid)] += int(cnt or 0)

                                breakdown_ok = True

                                print(f"[CARRIER DIFF BG] Breakdown novos:   {novos_bd}")

                                print(f"[CARRIER DIFF BG] Breakdown trocou:  {trocou_bd}")

                                print(f"[CARRIER DIFF BG] Breakdown calculado em {time.time()-_bd_t0:.1f}s")

                                sys.stdout.flush()

                                # === V8: Persistir lista de telefones para download (rotação simples — substitui run anterior) ===

                                try:

                                    _exp_t0 = time.time()

                                    bg_conn.execute("DROP TABLE IF EXISTS carrier_diff_novos")

                                    bg_conn.execute(

                                        "CREATE TABLE carrier_diff_novos AS "

                                        "SELECT n.telefone, p.operadora_id "

                                        "FROM _novos_tels n "

                                        "INNER JOIN portabilidade p ON n.telefone = p.telefone"

                                    )

                                    bg_conn.execute("DROP TABLE IF EXISTS carrier_diff_portados")

                                    bg_conn.execute(

                                        "CREATE TABLE carrier_diff_portados AS "

                                        "SELECT c.telefone, c.operadora_id "

                                        "FROM _changed_pairs c "

                                        "LEFT JOIN _novos_tels n ON c.telefone = n.telefone "

                                        "WHERE n.telefone IS NULL"

                                    )

                                    n_novos = bg_conn.execute("SELECT count(*) FROM carrier_diff_novos").fetchone()[0]

                                    n_port = bg_conn.execute("SELECT count(*) FROM carrier_diff_portados").fetchone()[0]

                                    bg_conn.commit()

                                    print(f"[CARRIER DIFF BG] Persistidas tabelas para download: carrier_diff_novos={n_novos:,}, carrier_diff_portados={n_port:,} (em {time.time()-_exp_t0:.1f}s)")

                                    sys.stdout.flush()

                                    # === V8.1: Acumular em carrier_diff_history pra suportar period=semana ===

                                    try:

                                        from datetime import datetime as _dt, timezone as _tz, timedelta as _td

                                        _hist_t0 = time.time()

                                        bg_conn.execute("""

                                            CREATE TABLE IF NOT EXISTS carrier_diff_history (

                                              kind TEXT NOT NULL,

                                              telefone TEXT NOT NULL,

                                              operadora_id INT,

                                              created_at TEXT NOT NULL

                                            )

                                        """)

                                        bg_conn.execute("CREATE INDEX IF NOT EXISTS idx_cdh_created ON carrier_diff_history(created_at)")

                                        bg_conn.execute("CREATE INDEX IF NOT EXISTS idx_cdh_kind_created ON carrier_diff_history(kind, created_at)")

                                        # Timestamp em fuso SP (-3)

                                        _sp_tz = _tz(_td(hours=-3))

                                        _now_sp = _dt.now(_sp_tz).strftime('%Y-%m-%d %H:%M:%S')

                                        # INSERT (acumula com timestamp)

                                        bg_conn.execute(

                                            "INSERT INTO carrier_diff_history (kind, telefone, operadora_id, created_at) "

                                            "SELECT 'novos', telefone, operadora_id, ? FROM carrier_diff_novos",

                                            (_now_sp,)

                                        )

                                        bg_conn.execute(

                                            "INSERT INTO carrier_diff_history (kind, telefone, operadora_id, created_at) "

                                            "SELECT 'portados', telefone, operadora_id, ? FROM carrier_diff_portados",

                                            (_now_sp,)

                                        )

                                        # Limpa registros > 30 dias (evita inchar)

                                        _cutoff = (_dt.now(_sp_tz) - _td(days=30)).strftime('%Y-%m-%d %H:%M:%S')

                                        bg_conn.execute("DELETE FROM carrier_diff_history WHERE created_at < ?", (_cutoff,))

                                        bg_conn.commit()

                                        n_hist = bg_conn.execute("SELECT count(*) FROM carrier_diff_history").fetchone()[0]

                                        print(f"[CARRIER DIFF BG] History acumulado: total={n_hist:,} linhas (em {time.time()-_hist_t0:.1f}s, ts={_now_sp})")

                                        sys.stdout.flush()

                                    except Exception as he:

                                        print(f"[CARRIER DIFF BG] Falha ao gravar history (nao bloqueante): {he}")

                                        sys.stdout.flush()

                                except Exception as pe:

                                    print(f"[CARRIER DIFF BG] Falha ao persistir tabelas de download (nao bloqueante): {pe}")

                                    sys.stdout.flush()

                            except Exception as bde:

                                print(f"[CARRIER DIFF BG] Falha no breakdown (nao bloqueante): {bde}")

                                sys.stdout.flush()

                                breakdown_ok = False

                            finally:

                                try:

                                    bg_conn.execute("DROP TABLE IF EXISTS _novos_tels")

                                    bg_conn.execute("DROP TABLE IF EXISTS _changed_pairs")

                                except Exception: pass

                        except Exception as de:

                            print(f"[CARRIER DIFF BG] Erro ao calcular diff: {de}")

                            sys.stdout.flush()

                            diff_novos = max(0, _bg_new_count - _bg_existing_count)

                            diff_removidos = max(0, _bg_existing_count - _bg_new_count)

                            diff_trocaram = -1

                            breakdown_ok = False

                    bg_conn.commit()

                    try:

                        bg_conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")

                    except Exception: pass

                    bg_conn.close()



                    # Gravar history (BG)

                    try:

                        hc = sqlite3.connect(_bg_db_path, timeout=10)

                        _bd_n = novos_bd if breakdown_ok else {"VIVO": None, "CLARO": None, "TIM": None, "OI": None, "OUTRAS": None}

                        _bd_t = trocou_bd if breakdown_ok else {"VIVO": None, "CLARO": None, "TIM": None, "OI": None, "OUTRAS": None}

                        hc.execute(

                            "INSERT INTO carrier_update_history ("

                            "timestamp_utc, started_by, duration_seconds, total_records, prev_records, "

                            "novos, removidos, trocaram_operadora, status, error_msg, "

                            "novos_vivo, novos_claro, novos_tim, novos_oi, novos_outras, "

                            "trocaram_vivo, trocaram_claro, trocaram_tim, trocaram_oi, trocaram_outras"

                            ") VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",

                            (

                                datetime.utcnow().isoformat() + "Z", _bg_started_by, int(time.time() - _bg_carrier_started_at),

                                int(_bg_new_count), int(_bg_existing_count),

                                int(diff_novos), int(diff_removidos), int(diff_trocaram), "SUCCESS", None,

                                _bd_n["VIVO"], _bd_n["CLARO"], _bd_n["TIM"], _bd_n["OI"], _bd_n["OUTRAS"],

                                _bd_t["VIVO"], _bd_t["CLARO"], _bd_t["TIM"], _bd_t["OI"], _bd_t["OUTRAS"],

                            )

                        )

                        hc.commit()

                        hc.close()

                        print(f"[CARRIER HISTORY BG] Registro gravado (breakdown={'ok' if breakdown_ok else 'skip'}).")

                        sys.stdout.flush()

                    except Exception as he:

                        print(f"[CARRIER HISTORY BG] Falha ao gravar historico: {he}")

                        sys.stdout.flush()



                    # Refresh DETALHADO (BG)

                    try:

                        refresh_result = _bg_refresh_detalhado()

                        print(f"[DETALHADO hook BG] refresh: {refresh_result}")

                        sys.stdout.flush()

                    except Exception as de:

                        print(f"[DETALHADO hook BG] erro: {de}")

                        sys.stdout.flush()

                except Exception as bge:

                    print(f"[CARRIER BG] Erro fatal: {bge}")

                    sys.stdout.flush()



            threading.Thread(target=_carrier_post_completion_bg, daemon=True, name=f"carrier_bg_{tid[:8]}").start()



        except Exception as e:

            print(f"Error in carrier update: {e}")

            self._update_task(tid, status="FAILED", message=f"Erro: {str(e)}")

            # Gravar entrada de historico (FAILED)

            try:

                _conn_h = sqlite3.connect(self.db_path, timeout=10)

                _conn_h.execute(

                    "INSERT INTO carrier_update_history (timestamp_utc, started_by, duration_seconds, total_records, prev_records, novos, removidos, trocaram_operadora, status, error_msg) VALUES (?,?,?,?,?,?,?,?,?,?)",

                    (datetime.utcnow().isoformat() + "Z", started_by, int(time.time() - _carrier_started_at), 0, 0, 0, 0, 0, "FAILED", str(e)[:500])

                )

                _conn_h.commit()

                _conn_h.close()

            except Exception as he:

                print(f"[CARRIER HISTORY] Falha ao gravar historico FAILED: {he}")

    def deep_search(self, name=None, cpf=None, cnpj=None, phone=None, uf=None):

        """Busca rápida unitária no ClickHouse"""

        ch_local = self._get_ch_client()

        if not ch_local:

            return pd.DataFrame()

            

        basics = []

        

        if cpf:

            # Robust cleaning

            cpf_clean = ''.join(filter(str.isdigit, str(cpf)))

            if len(cpf_clean) >= 11:

                cpf_mask = f"***{cpf_clean[3:9]}**"

                res = ch_local.query("SELECT cnpj_basico FROM hemn.socios WHERE cnpj_cpf_socio IN (%(c1)s, %(c2)s) LIMIT 50", 

                                          {'c1': cpf_clean, 'c2': cpf_mask})

                basics.extend([r[0] for r in res.result_rows])

                # Add check for MEIs where Razão Social contains the CPF

                res = ch_local.query("SELECT cnpj_basico FROM hemn.empresas WHERE razao_social LIKE %(c)s LIMIT 50",

                                          {'c': f"%{cpf_clean}%"})

                basics.extend([r[0] for r in res.result_rows])

        

        if cnpj:

            cnpj_clean = ''.join(filter(str.isdigit, str(cnpj)))

            if len(cnpj_clean) >= 8:

                basics.append(cnpj_clean[:8])



        if phone:

            tel_clean = ''.join(filter(str.isdigit, str(phone)))

            if len(tel_clean) >= 10:

                ddd = tel_clean[:2]

                num_8 = tel_clean[-8:]

                nums = [num_8]

                if len(tel_clean) == 11:

                    # Inclui o número completo de 9 dígitos se fornecido

                    nums.append(tel_clean[2:])

                

                # Busca na gaveta phone_lookup (ORDER BY ddd, telefone) — substitui 2 full scans em estabelecimento

                res = ch_local.query("SELECT cnpj_basico FROM hemn.phone_lookup WHERE ddd = %(ddd)s AND telefone IN %(nums)s LIMIT 100",

                                          {'ddd': ddd, 'nums': nums})

                basics.extend([r[0] for r in res.result_rows])



        if name:

            name_clean = remove_accents(str(name).strip().upper())

            # Try both prefix for speed and contains for flexibility

            res = ch_local.query("SELECT cnpj_basico FROM hemn.socios WHERE nome_socio LIKE %(n)s OR nome_socio LIKE %(nc)s LIMIT 50", 

                                      {'n': f'{name_clean}%', 'nc': f'%{name_clean}%'})

            basics.extend([r[0] for r in res.result_rows])

            res = ch_local.query("SELECT cnpj_basico FROM hemn.empresas WHERE razao_social LIKE %(n)s OR razao_social LIKE %(nc)s LIMIT 50", 

                                      {'n': f'{name_clean}%', 'nc': f'%{name_clean}%'})

            basics.extend([r[0] for r in res.result_rows])

            

        if not basics:

            return pd.DataFrame()

            

        basics = list(set(basics))

        # Filtro de impressão digital (Fingerprint): Nome + CPF se fornecidos

        socio_filters = []

        empresa_filters = []

        

        params_socio = []

        params_empresa = []

        

        if cpf:

            cpf_clean = ''.join(filter(str.isdigit, str(cpf)))

            if len(cpf_clean) >= 11:

                cpf_mask = f"***{cpf_clean[3:9]}**"

                socio_filters.append("s.cnpj_cpf_socio IN (%s, %s)")

                params_socio.extend([cpf_clean, cpf_mask])

                

                # Permite que MEIs anonimizados pela RFB (que começam com o próprio CNPJ no nome) passem pelo filtro de CPF restrito, 

                # desde que eles passem pelo filtro do NOME EXATO que adicionamos abaixo.

                empresa_filters.append("(e.razao_social LIKE %s OR (e.natureza_juridica = '2135' AND startsWith(e.razao_social, concat(substring(e.cnpj_basico, 1, 2), '.', substring(e.cnpj_basico, 3, 3), '.', substring(e.cnpj_basico, 6, 3)))))")

                params_empresa.append(f"%{cpf_clean}%")

        

        if name:

            name_norm = normalize_name(name)

            socio_filters.append("(s.nome_socio LIKE %s OR s.nome_socio LIKE %s)")

            params_socio.extend([f"{name_norm}%", f"%{name_norm}%"])

            

            empresa_filters.append("(e.razao_social LIKE %s OR e.razao_social LIKE %s)")

            params_empresa.extend([f"{name_norm}%", f"%{name_norm}%"])

            

        socio_match_sql = " AND ".join(socio_filters) if socio_filters else "1=1"

        company_name_match = " AND ".join(empresa_filters) if empresa_filters else "1=0"

        

        # Combine the params strictly in the order they appear in the query string:

        # WHERE ... IN (basics) AND ( socio_match_sql OR company_name_match )

        params = basics[:] + params_socio + params_empresa



        query = f"""

            SELECT e.razao_social, 

                   concat(est.cnpj_basico, est.cnpj_ordem, est.cnpj_dv) AS cnpj_completo,

                   multiIf(est.situacao_cadastral = '02', 'ATIVA', 'BAIXADA/INATIVA') AS situacao,

                   est.cnae_fiscal AS cnae_principal,

                   multiIf(e.natureza_juridica = '2135', 'SIM', 'NÃO') AS cnpj_cpf_socio,

                   est.correio_eletronico AS email_novo,

                   concat(est.logradouro, ', ', est.numero, ' - ', est.bairro, ' - ', coalesce(m.descricao, 'N/A'), '/', est.uf) AS endereco_completo,

                   multiIf(length(est.telefone1) = 8 AND substring(est.telefone1, 1, 1) IN ('6','7','8','9'), concat('9', est.telefone1), est.telefone1) AS telefone_novo,

                   est.ddd1 AS ddd_novo,

                   multiIf(length(est.telefone1) = 8 AND substring(est.telefone1, 1, 1) IN ('6','7','8','9'), 'CELULAR', 

                           length(est.telefone1) = 9, 'CELULAR', 'FIXO') AS tipo_telefone,

                   coalesce(est.nome_fantasia, e.razao_social) AS nome_fantasia,

                   multiIf(est.data_inicio_atividades != '', 

                           concat(substring(est.data_inicio_atividades, 7, 2), '/', 

                                  substring(est.data_inicio_atividades, 5, 2), '/', 

                                  substring(est.data_inicio_atividades, 1, 4)), 

                           'NÃO INFORMADA') AS data_abertura,

                   coalesce(e.capital_social, 0.0) AS capital_social,

                   multiIf(e.natureza_juridica = '2135', 'EMPRESÁRIO INDIVIDUAL (MEI)', 

                           e.natureza_juridica = '2062', 'SOCIEDADE EMPRESÁRIA LIMITADA',

                           e.natureza_juridica = '2305', 'ENTIDADE SINDICAL',

                           e.natureza_juridica = '3999', 'ASSOCIAÇÃO PRIVADA',

                           concat('CÓDIGO ', e.natureza_juridica)) AS natureza_juridica,

                   multiIf(e.porte_empresa = '01', 'MICRO EMPRESA',

                           e.porte_empresa = '03', 'PEQUENO PORTE',

                           e.porte_empresa = '05', 'DEMAIS (MÉDIA/GRANDE)',

                           'NÃO INFORMADO') AS porte,

                   est.cnae_fiscal_secundaria AS cnae_secundario,

                   arrayStringConcat(groupUniqArray(coalesce(s.nome_socio, 'NÃO ENCONTRADO')), ' / ') AS socio_nome

            FROM hemn.empresas_turbo e

            JOIN hemn.estabelecimento est ON e.cnpj_basico = est.cnpj_basico

            LEFT JOIN hemn.socios_turbo s ON e.cnpj_basico = s.cnpj_basico

            LEFT JOIN hemn.municipio m ON est.municipio = m.codigo

            WHERE e.cnpj_basico IN ({','.join(['%s' for _ in basics])})

              AND ({socio_match_sql} OR {company_name_match})

            GROUP BY 

                e.razao_social, 

                cnpj_completo,

                situacao,

                cnae_principal,

                cnpj_cpf_socio,

                email_novo,

                endereco_completo,

                telefone_novo,

                ddd_novo,

                tipo_telefone,

                nome_fantasia,

                data_abertura,

                capital_social,

                natureza_juridica,

                porte,

                cnae_secundario

            ORDER BY multiIf(situacao = 'ATIVA', 1, 2)

            LIMIT 50

        """

        res = ch_local.query(query, params)

        return pd.DataFrame(res.result_rows, columns=res.column_names)





    def _run_enrich(self, tid, input_file, output_dir, name_col, cpf_col, perfil="TODOS"):

        # V8: Semaforo de concorrencia (max 2 enrichs simultaneos no servidor)

        if not self._enrich_semaphore.acquire(blocking=False):

            self._update_task(

                tid, status="FAILED",

                message="Servidor com capacidade maxima de Enriquecimento PJ no momento (2 jobs em andamento). Aguarde alguns minutos e tente novamente."

            )

            return

        try:

            self._update_task(tid, status="PROCESSING", message="[v2.2.0-PREMIUM] Iniciando Escaneamento Titanium-MT (Motor Paralelo)...")

            status = self.get_task_status(tid)

            if status.get("status") == "CANCELLED": return

            start_time = time.time()

            output_file = os_native.path.join(output_dir, f"Enriquecido_{tid[:8]}.xlsx")

            

            # Memory Optimized: Determine row count without loading full file

            if input_file.endswith('.csv'):

                total_rows = sum(1 for _ in open(input_file, 'r', encoding='utf-8', errors='ignore')) - 1

            else:

                # Use openpyxl to get dimensions without loading data

                import openpyxl

                wb = openpyxl.load_workbook(input_file, read_only=True)

                total_rows = wb.active.max_row - 1

                wb.close()

            

            self._update_task(tid, progress=2, message=f"Lendo base ({total_rows:,} registros)...")



            # === PACOTE A: Leitura UNICA do arquivo (Fix #2) ===

            if input_file.endswith('.csv'):

                with open(input_file, 'r', encoding='utf-8', errors='ignore') as f:

                    primeira = f.readline()

                sep = ';' if ';' in primeira else (',' if ',' in primeira else '\t')

                df_in = pd.read_csv(input_file, sep=sep, engine='python', dtype=str)

            else:

                df_in = pd.read_excel(input_file, dtype=str)



            total = len(df_in)



            # --- PHASE 0: NORMALIZACAO (Fix #4: remove_accents vetorizado) ---

            n_col = name_col if (name_col and name_col in df_in.columns) else (df_in.columns[1] if len(df_in.columns) > 1 else df_in.columns[0])

            c_col = cpf_col if (cpf_col and cpf_col in df_in.columns) else df_in.columns[0]



            df_in['titanium_nome'] = (

                df_in[n_col].fillna('').astype(str).str.upper().str.strip()

                .str.normalize('NFD').str.encode('ascii', 'ignore').str.decode('ascii')

            )

            df_in['titanium_cpf'] = df_in[c_col].fillna('').astype(str).str.replace(r'\D', '', regex=True).str.zfill(11)



            # Listas para uso nas Phases 1+

            all_cpfs = df_in['titanium_cpf'].tolist()

            all_names = df_in['titanium_nome'].tolist()



            # --- EXTRACAO DE UF (DETERMINACAO DA COLUNA) ---

            # Prioridade: Coluna nomeada 'UF' ou 'Estado' -> Coluna C (Index 2)

            uf_col = next((c for c in df_in.columns if str(c).upper() in ['UF', 'ESTADO', 'UF_CLIENTE']), None)

            if not uf_col and len(df_in.columns) > 2:

                uf_col = df_in.columns[2]



            if uf_col:

                df_in['titanium_uf'] = df_in[uf_col].fillna('').astype(str).str.upper().str.strip().str[:2]

            else:

                df_in['titanium_uf'] = ''



            # === Mapa CPF/Nome/Chave -> UF (Fix #1: vetorizado, sem iterrows) ===

            ufs_series = df_in['titanium_uf']

            cpfs_series = df_in['titanium_cpf']

            nomes_series = df_in['titanium_nome']

            client_uf_map = {}

            mask_cpf = cpfs_series != ''

            client_uf_map.update(dict(zip(cpfs_series[mask_cpf], ufs_series[mask_cpf])))

            mask_nome = nomes_series != ''

            client_uf_map.update(dict(zip(nomes_series[mask_nome], ufs_series[mask_nome])))

            mask_chave = (cpfs_series.str.len() >= 11) & (nomes_series != '')

            chaves_series = nomes_series + " ***" + cpfs_series.str[3:9] + "**"

            client_uf_map.update(dict(zip(chaves_series[mask_chave], ufs_series[mask_chave])))



            # --- PHASE 1: BATCH PROCESSING (EXTREME SPEED) ---

            # Generar chaves combinadas (socio_chave) para busca ultra precisa

            # Otimização v1.8.11: Vetorização de Chaves (Elimina iterrows/CPU spike)

            mask_chave = (df_in['titanium_nome'] != '') & (df_in['titanium_cpf'].str.len() >= 11)

            if mask_chave.any():

                df_masked = df_in[mask_chave]

                search_chaves = (df_masked['titanium_nome'] + " ***" + df_masked['titanium_cpf'].str[3:9] + "**").unique().tolist()

            else:

                search_chaves = []



            # Mantemos CPF e Nome isolados apenas para fallbacks (se faltar um deles)

            valid_cpfs = [cpf for cpf in all_cpfs if len(cpf) >= 11]

            search_cpfs = list(set(valid_cpfs))

            

            valid_names = [normalize_name(n) for n in all_names if len(str(n)) > 3]

            search_names = list(set(valid_names))

            

            # Prepare Global Results

            all_results = []

            found_count = 0

            

            # --- PHASE 1: TITANIUM-TURBO DATA GATHERING ---

            # 1. Perfil Configuration

            p_val = str(perfil).upper().strip()

            if p_val == "MEI":

                perfil_cond_sql = "e.natureza_juridica = '2135'"

                is_expansion_required = False

            elif p_val == "NAO MEI":

                perfil_cond_sql = "e.natureza_juridica != '2135'"

                is_expansion_required = True

            else:

                perfil_cond_sql = "1=1"

                is_expansion_required = True



            # 2. Buscar dados em SÓCIOS

            if search_chaves or search_cpfs or search_names:

                self._update_task(tid, progress=10, message="Buscando Sócios...")

                

                # Queries para Sócios

                q_socios_base = f"""

                    SELECT cnpj_basico, nome_socio, socio_chave, {{lookup_col}} AS lookup_key 

                    FROM hemn.socios 

                    WHERE {{lookup_col}} IN %(keys)s

                    {"LIMIT 1 BY socio_chave" if not is_expansion_required else ""}

                """

                

                # === V8 (revisado): SEQUENCIAL com batch_size maior + max_query_size 2MB.

                # Paralelizacao removida pq saturava CPU do CH. Manter sequencial garante

                # uso de maquina similar ao Extrair Dados. O index tokenbf_v1 em nome_socio

                # ja torna o pass de nome ~5x mais rapido sem custo de CPU adicional.

                _ENRICH_BATCH = 8000

                _ENRICH_SETTINGS = {

                    'max_query_size': 2097152,   # 2 MB de IN clause (nomes longos)

                    'max_threads': 1,            # 1 thread por query no CH (mesmo comportamento original)

                }

                results_s = []

                cols_s = []



                # Prioridade 1: Chave Combinada (Socio Chave) — Muito mais precisa

                if search_chaves:

                    r, c = self._batch_query(

                        q_socios_base.format(lookup_col="socio_chave"), "keys", search_chaves,

                        batch_size=_ENRICH_BATCH, tid=tid, base_prog=10, max_prog=18,

                        extra_settings=_ENRICH_SETTINGS

                    )

                    results_s += r

                    if c: cols_s = c



                # Prioridade 2: CPF (apenas CPFs reais)

                if search_cpfs:

                    r, c = self._batch_query(

                        q_socios_base.format(lookup_col="cnpj_cpf_socio"), "keys", search_cpfs,

                        batch_size=_ENRICH_BATCH, tid=tid, base_prog=18, max_prog=26,

                        extra_settings=_ENRICH_SETTINGS

                    )

                    results_s += r

                    if c: cols_s = c



                # Prioridade 3: Nome (rede de seguranca — agora rapido com tokenbf index)

                if search_names:

                    r, c = self._batch_query(

                        q_socios_base.format(lookup_col="nome_socio"), "keys", search_names,

                        batch_size=_ENRICH_BATCH, tid=tid, base_prog=26, max_prog=35,

                        extra_settings=_ENRICH_SETTINGS

                    )

                    results_s += r

                    if c: cols_s = c

                

                if results_s:

                    df_socios = pd.DataFrame(results_s, columns=cols_s)

                    

                    # 3. Buscar dados de EMPRESAS e ESTABELECIMENTOS para os CNPJs encontrados

                    unique_cnpjs = df_socios['cnpj_basico'].unique().tolist()

                    self._update_task(tid, progress=45, message=f"Buscando Dados de {len(unique_cnpjs):,} Empresas...")



                    # Otimização v1.9.2: Filtro Geográfico na Query (Somente UFs presentes na planilha)

                    unique_ufs = [u for u in df_in['titanium_uf'].unique().tolist() if u and len(u) == 2]

                    uf_filter_sql = " AND uf IN %(ufs)s" if unique_ufs else ""

                    # PACOTE A FASE 2: migrado para hemn.comercial_pj (gaveta denormalizada particionada por UF).

                    # Elimina 2 hash JOINs (empresas + estabelecimento + municipio) — ganho 30-50% nesta fase.

                    # cnpj na gaveta e o full 14 chars; cnpj_basico/ordem/dv extraidos via substring.

                    perfil_cond_cp = perfil_cond_sql.replace("e.natureza_juridica", "natureza_juridica")

                    q_info = f"""

                        SELECT

                            substring(cnpj, 1, 8) AS cnpj_basico,

                            razao_social,

                            substring(cnpj, 9, 4) AS cnpj_ordem,

                            substring(cnpj, 13, 2) AS cnpj_dv,

                            situacao_cadastral,

                            uf,

                            ddd1, telefone1, ddd2, telefone2,

                            email AS correio_eletronico,

                            '' AS tipo_logradouro,

                            logradouro, numero,

                            '' AS complemento,

                            bairro, cep, cnae_fiscal,

                            '' AS municipio,

                            municipio_nome

                        FROM hemn.comercial_pj

                        WHERE substring(cnpj, 1, 8) IN %(keys)s AND {perfil_cond_cp} AND situacao_cadastral = '02' {uf_filter_sql}

                        ORDER BY (substring(cnpj, 9, 4) = '0001') DESC

                        LIMIT 1 BY substring(cnpj, 1, 8)

                    """



                    params_info = {}

                    if unique_ufs: params_info["ufs"] = unique_ufs

                    results_info, cols_info = self._batch_query(q_info, "keys", unique_cnpjs, tid=tid, base_prog=45, max_prog=75, extra_params=params_info)

                    

                    if results_info:

                        df_info = pd.DataFrame(results_info, columns=cols_info)

                        

                        # 4. Join Final (Socio + Info)

                        df_final_lookup = pd.merge(df_socios, df_info, on='cnpj_basico', how='inner')

                        

                        # Pack into all_results

                        # Pack into all_results (Otimização v1.8.12: to_dict em vez de iterrows)

                        for d in df_final_lookup.to_dict('records'):

                            addr = self._parse_address_columns(d)

                            cont = self._parse_contact_columns(d)

                            mapping = {'01': 'NULA', '02': 'ATIVA', '03': 'SUSPENSA', '04': 'INAPTA', '08': 'BAIXADA'}

                            # Verificação Geográfica de Segurança (Garante que a UF bate com a solicitada para esse registro)

                            l_key = str(d['lookup_key'])

                            target_uf = client_uf_map.get(l_key, '')

                            if target_uf and target_uf != str(d['uf']):

                                continue



                            mapping = {'01': 'NULA', '02': 'ATIVA', '03': 'SUSPENSA', '04': 'INAPTA', '08': 'BAIXADA'}

                            all_results.append({

                                'lookup_key': l_key,

                                'CNPJ': f"{str(int(d['cnpj_basico'])).zfill(8)}{str(int(d['cnpj_ordem'])).zfill(4)}{str(int(d['cnpj_dv'])).zfill(2)}",

                                'RAZAO_SOCIAL': d['razao_social'],

                                'SITUACAO': mapping.get(str(d['situacao_cadastral']).zfill(2), 'ATIVA'),

                                'CNAE': d['cnae_fiscal'], 'LOGRADOURO': str(addr[0]).upper(), 'NÚMERO': addr[1], 'COMPLEMENTO': addr[2],

                                'BAIRRO': addr[3], 'CIDADE': str(addr[4]).upper(), 'UF_END': addr[5], 'CEP': addr[6],

                                'TELEFONE_1': self._format_phone(d.get('ddd1'), d.get('telefone1')),

                                'TELEFONE_2': self._format_phone(d.get('ddd2'), d.get('telefone2')),

                                'EMAIL': cont[2],

                                'CHAVE_SOCIO': d.get('socio_chave', '')

                            })



                    # Motor One-Shot v5.1.4: Busca Híbrida Inteligente (CPF Flash-IN + Nome Aho-Scan)

                    def run_mei_scan(search_cpfs, search_names, tid, base_p, max_p, ufs=None):

                        results = []

                        ch_local = self._get_ch_client()

                        if not ch_local: return []



                        uf_filter = " AND uf IN %(ufs)s" if ufs else ""

                        params = {}

                        if ufs: params['ufs'] = ufs



                        # 1. BUSCA POR CPF (Modo Flash: IN Operator) - Resolve 99% dos casos em 1 seg

                        if search_cpfs:

                            cpfs = list(set([str(c).strip() for c in search_cpfs if len(str(c)) >= 11]))

                            if cpfs:

                                cpf_batch = 10000

                                for idx in range(0, len(cpfs), cpf_batch):

                                    block = cpfs[idx:idx + cpf_batch]

                                    self._update_task(tid, progress=base_p + 1, message=f"MEI CPF Flash: {idx:,}/{len(cpfs):,}...")

                                    # Otimização Crítica: substring(-11) extrai o CPF do final da Razão Social do MEI

                                    # Fix v1.9.2: Join com estabelecimento para filtrar UF (já que UF não existe em empresas)

                                    if ufs:

                                        sql_cpf = f"""

                                            SELECT e.cnpj_basico, e.razao_social 

                                            FROM hemn.empresas AS e

                                            INNER JOIN hemn.estabelecimento AS estab ON e.cnpj_basico = estab.cnpj_basico

                                            PREWHERE e.natureza_juridica = '2135'

                                            WHERE substring(e.razao_social, -11) IN %(cpfs)s AND estab.uf IN %(ufs)s

                                        """

                                    else:

                                        sql_cpf = "SELECT cnpj_basico, razao_social FROM hemn.empresas PREWHERE natureza_juridica = '2135' WHERE substring(razao_social, -11) IN %(cpfs)s"

                                    

                                    p = params.copy()

                                    p['cpfs'] = block

                                    print(f"[DEBUG] [MEI-CPF] Query: {sql_cpf}")

                                    print(f"[DEBUG] [MEI-CPF] Params Keys: {list(p.keys())}")

                                    try:

                                        res_cpf = ch_local.query(sql_cpf, p)

                                        results.extend(res_cpf.result_rows)

                                    except Exception as e:

                                        print(f"[ERROR] [MEI-CPF] ClickHouse Error: {e}")

                                        raise



                        # 2. BUSCA POR NOME (Rede de Segurança Total) - Varre tudo que não tiver CPF ou se falhou

                        if search_names:

                            names = list(set([str(n).upper().strip() for n in search_names if len(str(n)) > 5]))

                            total_names = len(names)

                            # Para evitar travamento, lotes cirúrgicos de 250 nomes.

                            batch_size = 500

                            for i in range(0, total_names, batch_size):

                                block = names[i:min(i + batch_size, total_names)]

                                self._update_task(tid, progress=base_p + 5 + int((i/total_names)*5), message=f"MEI Name Scan: {i:,}/{total_names:,}...")

                                # Fix v1.9.2: Join com estabelecimento para filtrar UF

                                if ufs:

                                    sql_name = f"""

                                        SELECT e.cnpj_basico, e.razao_social 

                                        FROM hemn.empresas AS e

                                        INNER JOIN hemn.estabelecimento AS estab ON e.cnpj_basico = estab.cnpj_basico

                                        PREWHERE e.natureza_juridica = '2135'

                                        WHERE multiSearchAny(e.razao_social, %(n)s) AND estab.uf IN %(ufs)s

                                    """

                                else:

                                    sql_name = f"SELECT cnpj_basico, razao_social FROM hemn.empresas PREWHERE natureza_juridica = '2135' WHERE multiSearchAny(razao_social, %(n)s)"

                                

                                p = params.copy()

                                p['n'] = block

                                print(f"[DEBUG] [MEI-NAME] Query: {sql_name}")

                                print(f"[DEBUG] [MEI-NAME] Params Keys: {list(p.keys())}")

                                try:

                                    res_name = ch_local.query(sql_name, p)

                                    results.extend(res_name.result_rows)

                                except Exception as e:

                                    print(f"[ERROR] [MEI-NAME] ClickHouse Error: {e}")

                                    raise

                        

                        return results



                    # Otimização v1.8.9: Recalcula listas para Phase 2 MEI

                    # Se tem CPF, ignoramos o NOME (Pois o CPF já existe na razão social do MEI e é 100x mais rápido)

                    # Otimização v1.8.11: Vetorização de Listas MEI

                    has_cpf = df_in['titanium_cpf'].str.len() >= 11

                    search_cpfs_mei = set(df_in.loc[has_cpf, 'titanium_cpf'].tolist())

                    

                    # Nomes apenas para quem não tem CPF válido

                    search_names_mei = set(df_in.loc[~has_cpf & (df_in['titanium_nome'].str.len() > 5), 'titanium_nome'].tolist())



                    # Sincronização v1.8.10: Só executa MEI Scan se perfil for TODOS ou MEI

                    if p_val in ["TODOS", "MEI"] and (search_cpfs_mei or search_names_mei):

                        found_meis = run_mei_scan(list(search_cpfs_mei), list(search_names_mei), tid, 40, 55, ufs=unique_ufs)

                    else:

                        found_meis = []



                    # 2. Busca de Detalhes (Apenas para os encontrados)

                    if found_meis:

                        found_cnpjs = list(set([str(r[0]) for r in found_meis]))

                        self._update_task(tid, progress=55, message=f"Buscando detalhes de {len(found_cnpjs)} MEIs...")

                        

                        ch_local = self._get_ch_client()

                        all_mei_results = []

                        cols_mei = []

                        

                        # Otimização v5.1.2: Lotes de 5.000 para evitar 'Max query size exceeded'

                        cnpj_batch = 10000

                        for idx in range(0, len(found_cnpjs), cnpj_batch):

                            block = found_cnpjs[idx:idx + cnpj_batch]

                            sql_details = """

                            SELECT 

                                e.cnpj_basico AS cnpj_basico, e.razao_social AS razao_social,

                                estab.cnpj_ordem AS cnpj_ordem, estab.cnpj_dv AS cnpj_dv, 

                                estab.situacao_cadastral AS situacao_cadastral, estab.uf AS uf,

                                estab.ddd1 AS ddd1, estab.telefone1 AS telefone1, 

                                estab.ddd2 AS ddd2, estab.telefone2 AS telefone2, 

                                estab.correio_eletronico AS correio_eletronico,

                                estab.logradouro AS logradouro, estab.numero AS numero, 

                                estab.complemento AS complementos, estab.bairro AS bairro, 

                                estab.cep AS cep, estab.cnae_fiscal AS cnae_fiscal, 

                                mun.descricao AS municipio_nome

                            FROM hemn.estabelecimento AS estab

                            INNER JOIN hemn.empresas AS e ON estab.cnpj_basico = e.cnpj_basico

                            LEFT JOIN hemn.municipio AS mun ON estab.municipio = mun.codigo

                            WHERE e.cnpj_basico IN %(cnpjs)s AND estab.situacao_cadastral = '02'

                            """

                            # Log detalhado v5.1.5: Captura possível ambiguidade de UF

                            # print(f"[DEBUG] [MEI-DETAILS] Batch {idx}: {len(block)} CNPJs")

                            try:

                                res_details = ch_local.query(sql_details, {'cnpjs': block})

                                all_mei_results.extend(res_details.result_rows)

                                if not cols_mei: cols_mei = res_details.column_names

                            except Exception as e:

                                print(f"[ERROR] [MEI-DETAILS] ClickHouse Error: {e}")

                                raise



                        # 3. Filtragem Geográfica Final (Precisão v1.8.6)

                        # Identifica coluna de UF para isolamento regional

                        uf_col = next((c for c in df_in.columns if c.upper() in ['UF', 'ESTADO', 'UF_CLIENTE']), None)

                        if uf_col:

                            # Criar mapa CPF -> UF do Cliente para filtragem cirúrgica (Fix #1: vetorizado)

                            ufs_v = df_in[uf_col].fillna('').astype(str).str.upper().str.strip()

                            cpfs_v = df_in['titanium_cpf'].astype(str)

                            nomes_v = df_in['titanium_nome'].astype(str).str.upper().str.strip()

                            client_uf_map = {}

                            mc = cpfs_v != ''

                            client_uf_map.update(dict(zip(cpfs_v[mc], ufs_v[mc])))

                            mn = nomes_v != ''

                            client_uf_map.update(dict(zip(nomes_v[mn], ufs_v[mn])))



                            final_filtered = []

                            for d in all_mei_results:

                                r_social = str(d[1])

                                r_uf = str(d[5])

                                # O MEI deve bater na UF do cliente que originou a busca

                                # Checamos se o CPF ou Nome do MEI e a UF batem com o cliente

                                is_match = False

                                for pattern, uf in client_uf_map.items():

                                    if pattern in r_social and uf == r_uf:

                                        is_match = True

                                        break

                                if is_match: final_filtered.append(d)

                            all_mei_results = final_filtered



                    if all_mei_results:

                        df_mei = pd.DataFrame(all_mei_results, columns=cols_mei).drop_duplicates(subset=['cnpj_basico'])

                        

                        # OTIMIZAÇÃO CRÍTICA: Criar mapas para lookup O(1) em vez de loop O(N^2)

                        # Sincronizado com v1.8.10 (search_cpfs_mei / search_names_mei)

                        cpf_lookup = {str(c).strip(): c for c in search_cpfs_mei}

                        mask_lookup = {f"***{str(c)[3:9]}**": c for c in search_cpfs_mei if len(str(c)) >= 9}

                        name_lookup = {str(n).upper().strip(): n for n in search_names_mei}



                        # OTIMIZAÇÃO v1.8.12: RegEx para busca O(N) em vez de O(N*M)

                        import re

                        name_pattern = re.compile('|'.join([re.escape(n) for n in name_lookup.keys() if len(n) > 5])) if name_lookup else None

                        

                        for d in df_mei.to_dict('records'):

                            addr = self._parse_address_columns(d)

                            cont = self._parse_contact_columns(d)

                            mapping = {'01': 'NULA', '02': 'ATIVA', '03': 'SUSPENSA', '04': 'INAPTA', '08': 'BAIXADA'}

                            

                            r_social = str(d['razao_social']).upper()

                            matched_key = None

                            

                            # 1. Tentar extrair o CPF/Mascara do final da Razão Social (Padrão MEI 2135)

                            parts = r_social.split()

                            if parts:

                                last_p = parts[-1].replace('.', '').replace('-', '')

                                if last_p in cpf_lookup:

                                    matched_key = cpf_lookup[last_p]

                                elif last_p in mask_lookup:

                                    matched_key = mask_lookup[last_p]

                            

                            # 2. Fallback: Busca por Nome via RegEx (Otimização Ultra v1.8.12)

                            if not matched_key and name_pattern:

                                match = name_pattern.search(r_social)

                                if match:

                                    matched_key = name_lookup.get(match.group())

                            

                            if matched_key:

                                all_results.append({

                                    'lookup_key': matched_key,

                                    'CNPJ': f"{str(int(d['cnpj_basico'])).zfill(8)}{str(int(d['cnpj_ordem'])).zfill(4)}{str(int(d['cnpj_dv'])).zfill(2)}",

                                    'RAZAO_SOCIAL': d['razao_social'],

                                    'SITUACAO': mapping.get(str(d['situacao_cadastral']).zfill(2), 'ATIVA'),

                                    'CNAE': d['cnae_fiscal'], 'LOGRADOURO': str(addr[0]).upper(), 'NÚMERO': addr[1], 'COMPLEMENTO': addr[2],

                                    'BAIRRO': addr[3], 'CIDADE': str(addr[4]).upper(), 'UF_END': addr[5], 'CEP': addr[6],

                                    'TELEFONE_1': self._format_phone(d.get('ddd1'), d.get('telefone1')),

                                    'TELEFONE_2': self._format_phone(d.get('ddd2'), d.get('telefone2')),

                                    'EMAIL': cont[2],

                                    'CHAVE_SOCIO': 'MEI (BUSCA DIRETA)'

                                })



            # --- PHASE 2: IN-MEMORY MAPPING & FALLBACKS ---

            self._update_task(tid, progress=80, message="Consolidando base de dados...")

            

            if all_results:

                df_cache = pd.DataFrame(all_results)

                

                # Preparar colunas de lookup no DataFrame original

                df_in['socio_chave_lookup'] = df_in.apply(lambda r: f"{str(r['titanium_nome']).strip()} ***{str(r['titanium_cpf'])[3:9]}**" if len(str(r['titanium_cpf'])) >= 11 else "", axis=1)

                

                # Detectar coluna de UF no input para regra de desempate

                uf_input_col = next((c for c in df_in.columns if c.upper() in ['UF', 'ESTADO', 'UF_CLIENTE']), None)

                if uf_input_col:

                    df_in[uf_input_col] = df_in[uf_input_col].astype(str).str.strip().str.upper()



                # Merge 1: Chave Combinada (Socio Chave) - PRIORIDADE MÁXIMA

                df_in['lookup_key'] = df_in['socio_chave_lookup']

                df_merged = pd.merge(df_in, df_cache, on='lookup_key', how='left')

                

                # Merge 2: CPF Exato (Fallback se não houver nome ou se houver erro na chave)

                null_rows = df_merged['CNPJ'].isna()

                if null_rows.any():

                    df_no_match_cpf = df_in[null_rows].copy()

                    df_no_match_cpf['lookup_key'] = df_no_match_cpf['titanium_cpf']

                    df_merged_cpf = pd.merge(df_no_match_cpf.drop(columns=df_cache.columns.drop('lookup_key'), errors='ignore'), df_cache, on='lookup_key', how='left')

                    df_merged = pd.concat([df_merged[~null_rows], df_merged_cpf], ignore_index=True)



                # Merge 3: Nomes (apenas onde ainda não houve match)

                null_rows = df_merged['CNPJ'].isna()

                if null_rows.any() and not df_cache.empty:

                    # Normalizar chaves no cache para comparação de nomes

                    df_cache['titanium_nome'] = df_cache['lookup_key'].apply(normalize_name)

                    df_cache_names = df_cache.drop_duplicates(subset=['titanium_nome', 'CNPJ'])

                    

                    df_no_match = df_merged[null_rows].drop(columns=df_cache.columns.drop('titanium_nome', errors='ignore'), errors='ignore').copy()

                    df_no_match['titanium_nome'] = df_no_match['titanium_nome'].apply(normalize_name)

                    df_merged_names = pd.merge(df_no_match, df_cache_names, on='titanium_nome', how='left')

                    

                    df_merged = pd.concat([df_merged[~null_rows], df_merged_names], ignore_index=True)



                # Merge 4: MEI Fallback (onde lookup_key era o Nome diretamente da busca de empresas)

                null_rows = df_merged['CNPJ'].isna()

                if null_rows.any() and not df_cache.empty:

                    df_no_match_mei = df_merged[null_rows].copy()

                    

                    # Tentar merge por CPF bruto (O(1) no lookup do cache)

                    df_no_match_mei['lookup_key'] = df_no_match_mei['titanium_cpf']

                    df_merged_mei_cpf = pd.merge(df_no_match_mei.drop(columns=df_cache.columns.drop('lookup_key'), errors='ignore'), df_cache, on='lookup_key', how='left')

                    

                    # Onde ainda sobrar nulo, tentar por Nome bruto

                    null_still = df_merged_mei_cpf['CNPJ'].isna()

                    if null_still.any():

                        df_no_match_name = df_merged_mei_cpf[null_still].copy()

                        df_no_match_name['lookup_key'] = df_no_match_name['titanium_nome']

                        df_merged_mei_name = pd.merge(df_no_match_name.drop(columns=df_cache.columns.drop('lookup_key'), errors='ignore'), df_cache, on='lookup_key', how='left')

                        df_merged_mei_cpf = pd.concat([df_merged_mei_cpf[~null_still], df_merged_mei_name], ignore_index=True)

                    

                    df_merged = pd.concat([df_merged[~null_rows], df_merged_mei_cpf], ignore_index=True)



                # --- LÓGICA DE FILTRO INTELIGENTE POR UF (REGRA DO CLIENTE) ---

                if uf_input_col and not df_merged.empty and 'CNPJ' in df_merged.columns:

                    # 1. Contar ocorrências globais por chave no cache (Unicidade Nacional)

                    counts = df_cache['lookup_key'].value_counts().rename('total_brazil')

                    df_merged = df_merged.merge(counts, on='lookup_key', how='left')

                    

                    # 2. Aplicar Regras de Descarte

                    # Regra 1: Se total_brazil == 1 (MANTÉM SEMPRE)

                    # Regra 2: Se total_brazil > 1 E UF_END == UF_INPUT (MANTÉM)

                    # Regra 3: Se total_brazil > 1 E UF_END != UF_INPUT (DESCARTA)

                    mask_discard = (df_merged['total_brazil'] > 1) & (df_merged['UF_END'] != df_merged[uf_input_col]) & df_merged['CNPJ'].notna()

                    

                    # Remover linhas descartadas

                    df_merged = df_merged[~mask_discard].copy()

                    df_merged = df_merged.drop(columns=['total_brazil'], errors='ignore')

                        

                # 3. Remover registros que não foram encontrados (limpar linhas vazias no retorno)

                df_merged = df_merged.dropna(subset=['CNPJ'])

                # V8 DEDUP: remove linhas duplicadas (mesma pessoa+empresa retornada multiplas vezes pelos 3 passes)

                # Os 3 passes (chave/CPF/nome) podem retornar a mesma combinacao CPF+CNPJ — dedupe garante unicidade.

                _dedup_cpf = c_col if c_col in df_merged.columns else 'titanium_cpf'

                if _dedup_cpf in df_merged.columns and 'CNPJ' in df_merged.columns:

                    _before_dedup = len(df_merged)

                    df_merged = df_merged.drop_duplicates(subset=[_dedup_cpf, 'CNPJ'])

                    print(f"[ENRICH DEDUP] {_before_dedup} -> {len(df_merged)} (removidas {_before_dedup - len(df_merged)} duplicatas)")

                found_count = len(df_merged)

                # === V8 Output Layout: A=CPF (do user), B=NOME COMPLETO (do user), C+ = enriquecimento ===

                # Recupera a coluna original de CPF/Nome do usuario (preferencia: c_col/n_col detectados);

                # se nao existirem, usa titanium_cpf/titanium_nome (versoes normalizadas) como fallback.

                df_out = df_merged.copy()

                # Preenche CPF: valor original do user, com fallback pra titanium_cpf

                if c_col and c_col in df_out.columns:

                    df_out['__cpf_user'] = df_out[c_col].astype(str).fillna('').replace({'nan': ''})

                    # Se a coluna original veio vazia, completa com titanium_cpf

                    if 'titanium_cpf' in df_out.columns:

                        empty_mask = df_out['__cpf_user'].str.strip() == ''

                        df_out.loc[empty_mask, '__cpf_user'] = df_out.loc[empty_mask, 'titanium_cpf'].astype(str)

                elif 'titanium_cpf' in df_out.columns:

                    df_out['__cpf_user'] = df_out['titanium_cpf'].astype(str)

                else:

                    df_out['__cpf_user'] = ''

                # Preenche NOME: valor original do user, com fallback pra titanium_nome

                if n_col and n_col in df_out.columns and n_col != c_col:

                    df_out['__nome_user'] = df_out[n_col].astype(str).fillna('').replace({'nan': ''})

                    if 'titanium_nome' in df_out.columns:

                        empty_mask = df_out['__nome_user'].str.strip() == ''

                        df_out.loc[empty_mask, '__nome_user'] = df_out.loc[empty_mask, 'titanium_nome'].astype(str)

                elif 'titanium_nome' in df_out.columns:

                    df_out['__nome_user'] = df_out['titanium_nome'].astype(str)

                else:

                    df_out['__nome_user'] = ''

                # Monta df_final com a ordem desejada

                final_columns = [

                    ('__cpf_user', 'CPF'),

                    ('__nome_user', 'NOME COMPLETO'),

                    ('CNPJ', 'CNPJ'),

                    ('RAZAO_SOCIAL', 'RAZAO_SOCIAL'),

                    ('SITUACAO', 'SITUACAO'),

                    ('CNAE', 'CNAE'),

                    ('LOGRADOURO', 'LOGRADOURO'),

                    ('NÚMERO', 'NÚMERO'),

                    ('COMPLEMENTO', 'COMPLEMENTO'),

                    ('BAIRRO', 'BAIRRO'),

                    ('CIDADE', 'CIDADE'),

                    ('UF_END', 'UF_END'),

                    ('CEP', 'CEP'),

                    ('TELEFONE_1', 'TELEFONE_1'),

                    ('TELEFONE_2', 'TELEFONE_2'),

                    ('EMAIL', 'EMAIL'),

                ]

                build_cols = []

                rename_map = {}

                for src, dst in final_columns:

                    if src in df_out.columns:

                        build_cols.append(src)

                        rename_map[src] = dst

                df_final = df_out[build_cols].rename(columns=rename_map)

                # === Defensiva final: remove qualquer linha sem CNPJ valido (NaN ou string vazia) ===

                print(f"[ENRICH DEBUG] df_out.columns: {list(df_out.columns)}")

                print(f"[ENRICH DEBUG] df_out.shape: {df_out.shape}")

                print(f"[ENRICH DEBUG] df_out CNPJ notna count: {df_out['CNPJ'].notna().sum() if 'CNPJ' in df_out.columns else 'N/A'}")

                print(f"[ENRICH DEBUG] df_final.columns: {list(df_final.columns)}")

                print(f"[ENRICH DEBUG] df_final.shape antes do filter: {df_final.shape}")

                print(f"[ENRICH DEBUG] df_final CNPJ notna count antes: {df_final['CNPJ'].notna().sum() if 'CNPJ' in df_final.columns else 'N/A'}")

                if 'CNPJ' in df_final.columns:

                    df_final = df_final[df_final['CNPJ'].notna()]

                    df_final = df_final[df_final['CNPJ'].astype(str).str.strip() != '']

                    df_final = df_final.reset_index(drop=True)

                print(f"[ENRICH DEBUG] df_final.shape DEPOIS do filter: {df_final.shape}")

                # Atualiza o contador apos a limpeza final

                found_count = len(df_final)

            else:

                found_count = 0

                # Sem matches — retorna planilha vazia com header consistente

                df_final = pd.DataFrame(columns=['CPF', 'NOME COMPLETO', 'CNPJ', 'RAZAO_SOCIAL', 'SITUACAO',

                                                  'CNAE', 'LOGRADOURO', 'NÚMERO', 'COMPLEMENTO', 'BAIRRO',

                                                  'CIDADE', 'UF_END', 'CEP', 'TELEFONE_1', 'TELEFONE_2', 'EMAIL'])



            # Salvar no Excel de forma ultra rapida

            status = self.get_task_status(tid)

            if status.get("status") == "CANCELLED": return

            self._update_task(tid, progress=95, message="Salvando arquivo .xlsx resultante...")

            # V8 fix: xlsxwriter com constant_memory estava corrompendo dados (1233 linhas viram NaN ao escrever).

            # Ressetar index, garantir dtype object nas colunas de string e usar openpyxl simples.

            try:

                df_final = df_final.reset_index(drop=True)

                # Forca dtype object pra evitar inferencia automatica que zera strings vazias

                for col in df_final.columns:

                    df_final[col] = df_final[col].astype(object)

            except Exception as _e:

                print(f"[ENRICH WRITE] sanitize falhou: {_e}")

            print(f"[ENRICH WRITE] df_final final shape: {df_final.shape}, CNPJ notna: {df_final['CNPJ'].notna().sum() if 'CNPJ' in df_final.columns else 'N/A'}")

            df_final.to_excel(output_file, index=False, sheet_name='Enriquecido', engine='openpyxl')

            total_time = time.time() - start_time

            

            # Converter numpy.int64 para int python nativo para evitar erro 500 no FastAPI

            fc_native = int(found_count)

            total_native = int(total)

            

            msg = f"TITANIUM-DONE: V2.0 Processou {total_native} linhas. {fc_native} encontrados em {total_time:.1f}s."

            self._update_task(tid, status="COMPLETED", progress=100, message=msg, result_file=output_file, record_count=fc_native)



        except Exception as e:

            import traceback

            err_msg = f"{str(e)} | {traceback.format_exc()[-200:]}"

            self._update_task(tid, status="FAILED", message=f"TITANIUM-ERROR: {err_msg}")

        finally:

            # V8: libera o semaforo (aceita o proximo enrich na fila)

            try:

                self._enrich_semaphore.release()

            except Exception:

                pass



    # --- EXTRACTION (FULL FILTERS) ---

    # ============================================================

    # DETALHADO — KPIs / Mapa / Operadoras / Extração por período

    # Snapshot pré-calculado 1×/dia (após carrier update das 22h).

    # Endpoints leem do SQLite direto. Live query é fallback raro.

    # ============================================================



    DETALHADO_UFS = ['SP','RJ','MG','RS','PR','BA','SC','PA','GO','PE','CE','DF',

                     'ES','MT','AM','PB','MA','RN','MS','AL','PI','SE','RO','TO',

                     'AC','AP','RR']



    _detalhado_refresh_running = False



    def _maybe_trigger_refresh_async(self):

        """Dispara refresh em background se ainda não estiver rodando.

        Usado como fallback quando endpoint detecta snapshot vazio.

        Pausa: se /tmp/_hemn_pause_auto_refresh existir, NAO dispara

        (usado pra evitar disputa com refresh standalone manual no ClickHouse)."""

        try:

            import os as _os_chk

            if _os_chk.path.exists('/tmp/_hemn_pause_auto_refresh'):

                return

        except Exception:

            pass

        if self._detalhado_refresh_running:

            return

        def _runner():

            try:

                self._detalhado_refresh_running = True

                print(f"[DETALHADO refresh] disparado por fallback")

                self.refresh_detalhado_snapshot()

            finally:

                self._detalhado_refresh_running = False

        threading.Thread(target=_runner, daemon=True).start()



    def _get_detalhado_snapshot(self, uf=None):

        """Lê snapshot do SQLite. uf=None ou 'BR' retorna agregado Brasil. Retorna None se não existir."""

        key = (uf or 'BR').upper()

        try:

            conn = sqlite3.connect(self.db_path, timeout=10)

            cur = conn.execute("SELECT payload, updated_at FROM detalhado_snapshot WHERE uf = ?", (key,))

            row = cur.fetchone()

            conn.close()

            if not row:

                return None

            try:

                data = json.loads(row[0])

            except Exception:

                return None

            data['updated_at'] = row[1]

            return data

        except Exception as e:

            print(f"[DETALHADO snapshot read] {e}")

            return None



    def _save_detalhado_snapshot(self, uf, payload):

        key = (uf or 'BR').upper()

        try:

            conn = sqlite3.connect(self.db_path, timeout=30)

            conn.execute("PRAGMA journal_mode=WAL")

            conn.execute(

                "INSERT OR REPLACE INTO detalhado_snapshot (uf, payload, updated_at) VALUES (?, ?, ?)",

                (key, json.dumps(payload, ensure_ascii=False), datetime.utcnow().isoformat() + 'Z')

            )

            conn.commit()

            conn.close()

            return True

        except Exception as e:

            print(f"[DETALHADO snapshot write {key}] {e}")

            return False



    def _compute_detalhado_payload(self, uf=None):

        """Calcula payload completo (KPIs + MEI + operadoras movel/fixo + DDD) pra uma UF.

        Reusa as funções live (que já têm lock interno).

        Retorna dict ou None se algo crítico falhar."""

        try:

            ov_movel = self._compute_overview_live(uf=uf, tipo='movel')

            if ov_movel is None or 'kpis' not in ov_movel:

                return None

            ov_fixo = self._compute_overview_live(uf=uf, tipo='fixo')

            ddd_rows = self._compute_ddd_live(uf=uf)

            mei_pair = self._compute_mei_breakdown_pair(uf=uf) or {}

            payload = {

                'uf': (uf or 'BR'),

                'kpis': ov_movel.get('kpis', {}),

                'mei': ov_movel.get('mei', {}),

                'operadoras_movel': ov_movel.get('operadoras', {}),

                'operadoras_fixo': (ov_fixo or {}).get('operadoras', {}),

                'ddd_rows': ddd_rows or [],

                'situacao_breakdown': self._compute_situacao_live(uf=uf),

                'mei_breakdown_yes': mei_pair.get('mei_yes'),

                'mei_breakdown_no': mei_pair.get('mei_no'),

            }

            return payload

        except Exception as e:

            print(f"[DETALHADO compute {uf}] {e}")

            return None



    def refresh_detalhado_snapshot(self, only_uf=None):

        """Recalcula snapshot pra todas as UFs (ou só uma se `only_uf` for passada).

        Também grava o snapshot 'BR' (sem filtro) e o mapa por UF.

        Não roda em paralelo: 1 UF de cada vez (pra não estourar ClickHouse).

        Chamado: 1× por dia no fim do carrier_update + manualmente via endpoint admin."""

        t0 = time.time()

        targets = [only_uf.upper()] if only_uf else (['BR'] + list(self.DETALHADO_UFS))

        success = []

        failed = []



        # 1) Map (contagem por UF)

        try:

            map_data = self._compute_map_live()

            if map_data is not None:

                self._save_detalhado_snapshot('__MAP__', {'states': map_data})

                print(f"[DETALHADO refresh] map salvo ({len(map_data)} UFs)")

        except Exception as e:

            print(f"[DETALHADO refresh map] {e}")



        # 2) Por UF (BR = sem filtro = todas)

        for u in targets:

            uf_param = None if u == 'BR' else u

            try:

                payload = self._compute_detalhado_payload(uf=uf_param)

                if payload and self._save_detalhado_snapshot(u, payload):

                    success.append(u)

                    print(f"[DETALHADO refresh] {u} ok ({payload['kpis'].get('total', 0):,} empresas)")

                else:

                    failed.append(u)

                    print(f"[DETALHADO refresh] {u} FALHOU (payload None ou save falhou)")

            except Exception as e:

                failed.append(u)

                print(f"[DETALHADO refresh] {u} EXCEPTION: {e}")



        dur = time.time() - t0

        # Limpa cache em memória

        self._detalhado_cache = {}

        print(f"[DETALHADO refresh] concluído em {dur:.1f}s — sucesso={len(success)} falhas={len(failed)} ({','.join(failed) if failed else '-'})")

        return {'success': success, 'failed': failed, 'duration_seconds': round(dur, 1)}



    # --- Versões "live" (computam direto do ClickHouse) ---



    def _compute_overview_live(self, uf=None, tipo='movel'):

        """Versão live (ClickHouse direto). Retorna dict no mesmo shape do snapshot.uf."""

        if not self.is_linux:

            return None

        ch = self._get_ch_client()

        if not ch:

            return None

        where_uf = " WHERE uf = {uf:String}" if uf else ""

        params = {'uf': uf} if uf else {}

        sql = f"""

        SELECT count(*) AS total,

            countIf(length(telefone1) = 9 OR (length(telefone1) = 8 AND substring(telefone1,1,1) IN ('6','7','8','9'))) AS movel,

            countIf(length(telefone1) = 8 AND substring(telefone1,1,1) IN ('2','3','4','5')) AS fixo,

            countIf(telefone1 = '' OR length(telefone1) < 8) AS invalido,

            countIf(natureza_juridica = '2135') AS mei,

            countIf(natureza_juridica != '2135' AND natureza_juridica != '') AS nao_mei

        FROM hemn.comercial_pj

        {where_uf}

        """

        with self._detalhado_query_lock:

            try:

                row = ch.query(sql, parameters=params).result_rows[0]

            except Exception as e:

                print(f"[DETALHADO live overview {uf}] {e}")

                return None

        operadoras = self._count_operadoras_fixo(uf) if tipo == 'fixo' else self._count_operadoras_buckets(uf)

        return {

            'kpis': {'total': int(row[0]), 'movel': int(row[1]), 'fixo': int(row[2]), 'invalido': int(row[3])},

            'mei': {'mei': int(row[4]), 'nao_mei': int(row[5])},

            'operadoras': operadoras,

        }



    def _compute_ddd_live(self, uf=None):

        if not self.is_linux:

            return []

        ch = self._get_ch_client()

        if not ch:

            return []

        where_uf = " AND uf = {uf:String}" if uf else ""

        params = {'uf': uf} if uf else {}

        sql = f"""

        SELECT ddd1 AS ddd,

            countIf(length(telefone1) = 8 AND substring(telefone1,1,1) IN ('2','3','4','5')) AS fixo,

            countIf(length(telefone1) = 9 OR (length(telefone1) = 8 AND substring(telefone1,1,1) IN ('6','7','8','9'))) AS celular

        FROM hemn.comercial_pj

        WHERE ddd1 != '' AND telefone1 != ''

            {where_uf}

        GROUP BY ddd1

        ORDER BY (fixo + celular) DESC

        """

        with self._detalhado_query_lock:

            try:

                rows = ch.query(sql, parameters=params).result_rows

                return [{'ddd': r[0], 'fixo': int(r[1]), 'celular': int(r[2])} for r in rows if r and r[0]]

            except Exception as e:

                print(f"[DETALHADO live ddd {uf}] {e}")

                return []



    def _compute_map_live(self):

        if not self.is_linux:

            return None

        ch = self._get_ch_client()

        if not ch:

            return None

        with self._detalhado_query_lock:

            try:

                rows = ch.query("SELECT uf, count(*) AS total FROM hemn.comercial_pj WHERE uf != '' GROUP BY uf ORDER BY total DESC").result_rows

                return [{'uf': r[0], 'total': int(r[1])} for r in rows]

            except Exception as e:

                print(f"[DETALHADO live map] {e}")

                return None



    def _compute_situacao_live(self, uf=None):

        """Breakdown por situacao_cadastral x tipo de linha. Retorna [{situacao, fixo, movel, invalido}, ...]."""

        if not self.is_linux:

            return []

        ch = self._get_ch_client()

        if not ch:

            return []

        where_uf = " AND uf = {uf:String}" if uf else ""

        params = {'uf': uf} if uf else {}

        sql = f"""

        SELECT situacao_cadastral AS sit,

            countIf(length(telefone1) = 8 AND substring(telefone1,1,1) IN ('2','3','4','5')) AS fixo,

            countIf(length(telefone1) = 9 OR (length(telefone1) = 8 AND substring(telefone1,1,1) IN ('6','7','8','9'))) AS movel,

            countIf(telefone1 = '' OR length(telefone1) < 8) AS invalido

        FROM hemn.comercial_pj

        WHERE situacao_cadastral != ''

            {where_uf}

        GROUP BY situacao_cadastral

        """

        mapping = {'01': 'NULA', '02': 'ATIVA', '03': 'SUSPENSA', '04': 'INAPTA', '08': 'BAIXADA'}

        order = ['ATIVA', 'BAIXADA', 'SUSPENSA', 'INAPTA', 'NULA']

        with self._detalhado_query_lock:

            try:

                rows = ch.query(sql, parameters=params).result_rows

            except Exception as e:

                print(f"[DETALHADO live situacao {uf}] {e}")

                return []

        by_label = {}

        for r in rows:

            label = mapping.get(str(r[0]).zfill(2), str(r[0]))

            by_label[label] = {

                'situacao': label,

                'fixo': int(r[1]),

                'movel': int(r[2]),

                'invalido': int(r[3]),

            }

        return [by_label[k] for k in order if k in by_label]



    def get_detailed_overview(self, uf=None, tipo='movel'):

        """KPIs do módulo DETALHADO. Tenta snapshot SQLite primeiro (instantâneo).

        Fallback: live query no ClickHouse (raro — só se snapshot vazio)."""

        tipo = (tipo or 'movel').lower()

        if tipo not in ('movel', 'fixo'):

            tipo = 'movel'



        # 1) Snapshot pré-calculado

        snap = self._get_detalhado_snapshot(uf)

        if snap and 'kpis' in snap:

            op_key = 'operadoras_fixo' if tipo == 'fixo' else 'operadoras_movel'

            return {

                'uf': snap.get('uf', uf or 'BR'),

                'tipo': tipo,

                'kpis': snap.get('kpis', {}),

                'mei': snap.get('mei', {}),

                'operadoras': snap.get(op_key, {}),

                'updated_at': snap.get('updated_at'),

                'source': 'snapshot',

            }



        # 2) Fallback: live (e dispara refresh em background)

        print(f"[DETALHADO] snapshot vazio pra {uf or 'BR'} — fallback live + agendando refresh")

        self._maybe_trigger_refresh_async()



        cache_key = f"overview_{uf or 'BR'}_{tipo}"

        cached = self._detalhado_cache.get(cache_key)

        if cached and (time.time() - cached['ts']) < 300:

            return cached['data']



        if not self.is_linux:

            return {"error": "ClickHouse indisponível (ambiente não-linux)"}

        ch = self._get_ch_client()

        if not ch:

            return {"error": "Falha ao conectar no ClickHouse"}



        where_uf = " WHERE uf = {uf:String}" if uf else ""

        params = {'uf': uf} if uf else {}



        # Query única (KPIs + MEI no mesmo SELECT, sem JOIN — comercial_pj já tem tudo consolidado)

        sql_kpi = f"""

        SELECT

            count(*) AS total,

            countIf(

                length(telefone1) = 9

                OR (length(telefone1) = 8 AND substring(telefone1,1,1) IN ('6','7','8','9'))

            ) AS movel,

            countIf(

                length(telefone1) = 8 AND substring(telefone1,1,1) IN ('2','3','4','5')

            ) AS fixo,

            countIf(

                telefone1 = '' OR length(telefone1) < 8

            ) AS invalido,

            countIf(natureza_juridica = '2135') AS mei,

            countIf(natureza_juridica != '2135' AND natureza_juridica != '') AS nao_mei

        FROM hemn.comercial_pj

        {where_uf}

        """

        with self._detalhado_query_lock:

            try:

                row = ch.query(sql_kpi, parameters=params).result_rows[0]

                kpi_total, kpi_movel, kpi_fixo, kpi_invalido = int(row[0]), int(row[1]), int(row[2]), int(row[3])

                mei_count, nao_mei_count = int(row[4]), int(row[5])

            except Exception as e:

                print(f"[DETALHADO] kpi query falhou: {e}")

                kpi_total = kpi_movel = kpi_fixo = kpi_invalido = 0

                mei_count = nao_mei_count = 0



        # Operadoras (movel via portabilidade, fixo via prefix_tree)

        if tipo == 'fixo':

            operadoras = self._count_operadoras_fixo(uf)

        else:

            operadoras = self._count_operadoras_buckets(uf)



        result = {

            "uf": uf or "BR",

            "tipo": tipo,

            "kpis": {

                "total": kpi_total,

                "movel": kpi_movel,

                "fixo": kpi_fixo,

                "invalido": kpi_invalido,

            },

            "mei": {"mei": mei_count, "nao_mei": nao_mei_count},

            "operadoras": operadoras,

            "_cache_ts": int(time.time()),

        }

        self._detalhado_cache[cache_key] = {"ts": time.time(), "data": result}

        return result



    def _count_operadoras_buckets(self, uf=None, situacao=None, sample_limit=100000):

        """Bucketiza celulares da base em VIVO/CLARO/TIM/OI/OUTRAS via portabilidade SQLite.

        Usa amostra de até `sample_limit` numeros (default 100k) por performance.

        Nao-encontrados na portabilidade caem em OUTRAS.

        situacao: '01'/'02'/'03'/'04'/'08' para filtrar; None ou 'TODOS' para nao filtrar."""

        if not self.is_linux:

            return {"VIVO": 0, "CLARO": 0, "TIM": 0, "OI": 0, "OUTRAS": 0, "_sample": 0}

        ch = self._get_ch_client()

        if not ch:

            return {"VIVO": 0, "CLARO": 0, "TIM": 0, "OI": 0, "OUTRAS": 0, "_sample": 0}



        where_uf = " AND uf = {uf:String}" if uf else ""

        sit = (situacao or '').strip().upper()

        where_sit = " AND situacao_cadastral = {situacao:String}" if sit and sit != 'TODOS' else ""

        params = {}

        if uf: params['uf'] = uf

        if where_sit: params['situacao'] = sit



        sql = f"""

        SELECT concat(ddd1,

            multiIf(length(telefone1) = 8 AND substring(telefone1,1,1) IN ('6','7','8','9'),

                    concat('9', telefone1), telefone1)) AS tel_full

        FROM hemn.comercial_pj

        WHERE ddd1 != ''

            AND (length(telefone1) = 9 OR (length(telefone1) = 8 AND substring(telefone1,1,1) IN ('6','7','8','9')))

            {where_uf}

            {where_sit}

        LIMIT {int(sample_limit)}

        """

        with self._detalhado_query_lock:

            try:

                rows = ch.query(sql, parameters=params).result_rows

            except Exception as e:

                print(f"[DETALHADO] sample celulares falhou: {e}")

                return {"VIVO": 0, "CLARO": 0, "TIM": 0, "OI": 0, "OUTRAS": 0, "_sample": 0}



        telefones = [r[0] for r in rows if r and r[0]]

        sample_size = len(telefones)

        if sample_size == 0:

            return {"VIVO": 0, "CLARO": 0, "TIM": 0, "OI": 0, "OUTRAS": 0, "_sample": 0}



        # 2-step lookup: portabilidade SQLite (numeros portados) + prefix_tree ANATEL (operadora original).

        # Portabilidade tem ~57M registros (so portados). Brasil tem 250M+ celulares -> a maioria nunca portou

        # e fica na operadora original do prefixo. Sem o fallback prefix_tree, ~80% caem em OUTRAS errado.

        op_counts = {}

        not_found_phones = []

        try:

            conn = sqlite3.connect(self.db_carrier, timeout=60)

            conn.execute("PRAGMA journal_mode=WAL")

            conn.execute("CREATE TEMP TABLE IF NOT EXISTS _det_query_tels (telefone TEXT PRIMARY KEY)")

            conn.execute("DELETE FROM _det_query_tels")

            conn.executemany("INSERT OR IGNORE INTO _det_query_tels VALUES (?)", [(t,) for t in telefones])

            cur = conn.execute(

                "SELECT p.telefone, p.operadora_id FROM portabilidade p INNER JOIN _det_query_tels q ON p.telefone = q.telefone"

            )

            found = dict(cur.fetchall())

            conn.close()



            for t in telefones:

                opid = found.get(t)

                if opid is None:

                    not_found_phones.append(t)

                else:

                    k = str(opid)

                    op_counts[k] = op_counts.get(k, 0) + 1

        except Exception as e:

            print(f"[DETALHADO] portabilidade lookup falhou: {e}")

            # Sem portabilidade, ainda tentamos prefix_tree no sample inteiro

            not_found_phones = list(telefones)

            op_counts = {}



        # Fallback: prefix_tree ANATEL pra quem nao foi portado.

        # Otimizado: itera so os tamanhos distintos de prefixo (~6-7) com dict lookup O(1),

        # ao inves de varrer 100k+ chaves do prefix_tree por telefone.

        prefix_op_counts = {}

        still_unknown = 0

        prefix_tree = getattr(self, 'prefix_tree', {}) or {}

        if prefix_tree and not_found_phones:

            prefix_lengths = sorted({len(p) for p in prefix_tree.keys()}, reverse=True)

            for tel in not_found_phones:

                company = None

                for L in prefix_lengths:

                    cand = tel[:L]

                    company = prefix_tree.get(cand)

                    if company:

                        break

                if company:

                    prefix_op_counts[company] = prefix_op_counts.get(company, 0) + 1

                else:

                    still_unknown += 1

        else:

            still_unknown = len(not_found_phones)



        buckets = {"VIVO": 0, "CLARO": 0, "TIM": 0, "OI": 0, "OUTRAS": still_unknown}

        def _route(key, cnt):

            norm = (self.get_op_name(key) or str(key) or "").upper()

            if "VIVO" in norm or "TELEFONICA" in norm:

                buckets["VIVO"] += cnt

            elif "CLARO" in norm or "EMBRATEL" in norm:

                buckets["CLARO"] += cnt

            elif "TIM" in norm:

                buckets["TIM"] += cnt

            elif norm == "OI" or norm.startswith("OI ") or "TELEMAR" in norm or "OI / TELEMAR" in norm:

                buckets["OI"] += cnt

            else:

                buckets["OUTRAS"] += cnt

        for opid, cnt in op_counts.items():

            _route(opid, cnt)

        for company, cnt in prefix_op_counts.items():

            _route(company, cnt)



        buckets["_sample"] = sample_size

        return buckets



    def _count_operadoras_fixo(self, uf=None, situacao=None, sample_limit=100000):

        """Bucketiza FIXOS via self.prefix_tree (base ANATEL local).

        Operadora fixa e determinada pelo prefixo do numero (sem portabilidade).

        situacao: '01'/'02'/'03'/'04'/'08' para filtrar; None ou 'TODOS' para nao filtrar."""

        if not self.is_linux:

            return {"VIVO": 0, "CLARO": 0, "TIM": 0, "OI": 0, "OUTRAS": 0, "_sample": 0}

        ch = self._get_ch_client()

        if not ch:

            return {"VIVO": 0, "CLARO": 0, "TIM": 0, "OI": 0, "OUTRAS": 0, "_sample": 0}



        where_uf = " AND uf = {uf:String}" if uf else ""

        sit = (situacao or '').strip().upper()

        where_sit = " AND situacao_cadastral = {situacao:String}" if sit and sit != 'TODOS' else ""

        params = {}

        if uf: params['uf'] = uf

        if where_sit: params['situacao'] = sit



        # Pega telefones fixos (8 digitos comecando 2-5)

        sql = f"""

        SELECT concat(ddd1, telefone1) AS tel_full

        FROM hemn.comercial_pj

        WHERE ddd1 != ''

            AND length(telefone1) = 8

            AND substring(telefone1, 1, 1) IN ('2','3','4','5')

            {where_uf}

            {where_sit}

        LIMIT {int(sample_limit)}

        """

        with self._detalhado_query_lock:

            try:

                rows = ch.query(sql, parameters=params).result_rows

            except Exception as e:

                print(f"[DETALHADO] sample fixos falhou: {e}")

                return {"VIVO": 0, "CLARO": 0, "TIM": 0, "OI": 0, "OUTRAS": 0, "_sample": 0}



        telefones = [r[0] for r in rows if r and r[0]]

        sample_size = len(telefones)

        if sample_size == 0:

            return {"VIVO": 0, "CLARO": 0, "TIM": 0, "OI": 0, "OUTRAS": 0, "_sample": 0}



        buckets = {"VIVO": 0, "CLARO": 0, "TIM": 0, "OI": 0, "OUTRAS": 0}

        prefix_tree = getattr(self, 'prefix_tree', {}) or {}

        # Otimizado: itera so os tamanhos distintos de prefixo (~7) com lookup O(1) no dict,

        # ao inves de varrer 103k+ chaves do prefix_tree por telefone (era O(n*m) = bilhoes).

        prefix_lengths = sorted({len(p) for p in prefix_tree.keys()}, reverse=True) if prefix_tree else []



        for tel in telefones:

            company = None

            for L in prefix_lengths:

                cand = tel[:L]

                company = prefix_tree.get(cand)

                if company:

                    break

            if not company:

                buckets["OUTRAS"] += 1

                continue

            norm = (self.get_op_name(company) or company or "").upper()

            if "VIVO" in norm or "TELEFONICA" in norm:

                buckets["VIVO"] += 1

            elif "CLARO" in norm or "EMBRATEL" in norm:

                buckets["CLARO"] += 1

            elif "TIM" in norm:

                buckets["TIM"] += 1

            elif norm == "OI" or norm.startswith("OI ") or "TELEMAR" in norm or "OI / TELEMAR" in norm:

                buckets["OI"] += 1

            else:

                buckets["OUTRAS"] += 1



        buckets["_sample"] = sample_size

        return buckets



    def get_detailed_ddd(self, uf=None):

        """Lista [{ddd, fixo, celular}] por DDD. Lê snapshot SQLite (instantâneo)."""

        snap = self._get_detalhado_snapshot(uf)

        if snap and 'ddd_rows' in snap:

            return snap['ddd_rows']

        # Fallback live

        print(f"[DETALHADO] ddd snapshot vazio pra {uf or 'BR'} — fallback live")

        self._maybe_trigger_refresh_async()

        cache_key = f"ddd_{uf or 'BR'}"

        cached = self._detalhado_cache.get(cache_key)

        if cached and (time.time() - cached['ts']) < 300:

            return cached['data']



        if not self.is_linux:

            return []

        ch = self._get_ch_client()

        if not ch:

            return []



        where_uf = " AND uf = {uf:String}" if uf else ""

        params = {'uf': uf} if uf else {}



        sql = f"""

        SELECT

            ddd1 AS ddd,

            countIf(

                length(telefone1) = 8 AND substring(telefone1,1,1) IN ('2','3','4','5')

            ) AS fixo,

            countIf(

                length(telefone1) = 9 OR (length(telefone1) = 8 AND substring(telefone1,1,1) IN ('6','7','8','9'))

            ) AS celular

        FROM hemn.comercial_pj

        WHERE ddd1 != ''

            AND telefone1 != ''

            {where_uf}

        GROUP BY ddd1

        ORDER BY (fixo + celular) DESC

        """

        with self._detalhado_query_lock:

            try:

                rows = ch.query(sql, parameters=params).result_rows

                data = [{"ddd": r[0], "fixo": int(r[1]), "celular": int(r[2])} for r in rows if r and r[0]]

            except Exception as e:

                print(f"[DETALHADO] ddd query falhou: {e}")

                data = []

        self._detalhado_cache[cache_key] = {"ts": time.time(), "data": data}

        return data



    def get_detailed_map(self):

        """Contagem por UF pra renderizar choropleth. Lê snapshot SQLite."""

        snap = self._get_detalhado_snapshot('__MAP__')

        if snap and 'states' in snap:

            return snap['states']

        # Fallback live

        print(f"[DETALHADO] map snapshot vazio — fallback live")

        self._maybe_trigger_refresh_async()

        cache_key = "map_br"

        cached = self._detalhado_cache.get(cache_key)

        if cached and (time.time() - cached['ts']) < 600:

            return cached['data']



        if not self.is_linux:

            return []

        ch = self._get_ch_client()

        if not ch:

            return []



        sql = """

        SELECT uf, count(*) AS total

        FROM hemn.comercial_pj

        WHERE uf != ''

        GROUP BY uf

        ORDER BY total DESC

        """

        with self._detalhado_query_lock:

            try:

                rows = ch.query(sql).result_rows

                data = [{"uf": r[0], "total": int(r[1])} for r in rows]

            except Exception as e:

                print(f"[DETALHADO] map query falhou: {e}")

                data = []

        self._detalhado_cache[cache_key] = {"ts": time.time(), "data": data}

        return data



    def _compute_mei_breakdown_pair(self, uf=None):

        """Computa MEI e Nao-MEI em UMA SO query agrupada por (mei_flag x situacao).

        Retorna {'mei_yes': {...}, 'mei_no': {...}}. Cada metade tem o mesmo shape do antigo."""

        empty = {'mei_yes': None, 'mei_no': None}

        if not self.is_linux:

            return empty

        ch = self._get_ch_client()

        if not ch:

            return empty

        where_uf = " AND uf = {uf:String}" if uf else ""

        params = {'uf': uf} if uf else {}

        sql = f"""

        SELECT

            if(natureza_juridica = '2135', 'M', 'N') AS mei_flag,

            situacao_cadastral AS sit,

            count(*) AS total,

            countIf(length(telefone1) = 8 AND substring(telefone1,1,1) IN ('2','3','4','5')) AS fixo,

            countIf(length(telefone1) = 9 OR (length(telefone1) = 8 AND substring(telefone1,1,1) IN ('6','7','8','9'))) AS movel,

            countIf(telefone1 = '' OR length(telefone1) < 8) AS invalido

        FROM hemn.comercial_pj

        WHERE situacao_cadastral != '' AND natureza_juridica != ''

            {where_uf}

        GROUP BY mei_flag, situacao_cadastral

        """

        mapping = {'01': 'NULA', '02': 'ATIVA', '03': 'SUSPENSA', '04': 'INAPTA', '08': 'BAIXADA'}

        order = ['ATIVA', 'BAIXADA', 'SUSPENSA', 'INAPTA', 'NULA']

        with self._detalhado_query_lock:

            try:

                rows = ch.query(sql, parameters=params).result_rows

            except Exception as e:

                print(f"[DETALHADO] mei_breakdown_pair {uf} falhou: {e}")

                return empty

        def _build(filt_rows, is_mei):

            by_label = {}

            tot_fixo = tot_movel = tot_inval = tot_geral = 0

            for r in filt_rows:

                label = mapping.get(str(r[1]).zfill(2), str(r[1]))

                cnt, fx, mv, iv = int(r[2]), int(r[3]), int(r[4]), int(r[5])

                by_label[label] = {'situacao': label, 'total': cnt, 'fixo': fx, 'movel': mv, 'invalido': iv}

                tot_geral += cnt; tot_fixo += fx; tot_movel += mv; tot_inval += iv

            return {

                'uf': uf or 'BR',

                'is_mei': bool(is_mei),

                'total': tot_geral,

                'situacoes': [by_label[k] for k in order if k in by_label],

                'tipos_telefone': {'fixo': tot_fixo, 'movel': tot_movel, 'invalido': tot_inval},

            }

        return {

            'mei_yes': _build([r for r in rows if r[0] == 'M'], True),

            'mei_no':  _build([r for r in rows if r[0] == 'N'], False),

        }



    def get_detailed_mei_breakdown(self, uf=None, is_mei=True):

        """Detalhamento de MEI ou Nao-MEI: situacao cadastral + tipos de telefone.

        Le snapshot primeiro, fallback live + agenda refresh."""

        snap = self._get_detalhado_snapshot(uf)

        if snap:

            key = 'mei_breakdown_yes' if is_mei else 'mei_breakdown_no'

            cached = snap.get(key)

            if cached:

                return cached

        # snapshot vazio ou sem o campo (gerado antes desta feature) -> fallback + agendar

        print(f"[DETALHADO] mei_breakdown snapshot vazio pra {uf or 'BR'} -- fallback live + agendando refresh")

        self._maybe_trigger_refresh_async()

        pair = self._compute_mei_breakdown_pair(uf=uf)

        return (pair or {}).get('mei_yes' if is_mei else 'mei_no')



    def get_detailed_operadoras(self, uf=None, tipo='movel', situacao=None):

        """Distribuicao de operadoras (VIVO/CLARO/TIM/OI/OUTRAS).

        Sem situacao (ou TODOS): le do snapshot. Com situacao especifica: query live filtrada."""

        tipo = (tipo or 'movel').lower()

        if tipo not in ('movel', 'fixo'):

            tipo = 'movel'

        sit = (situacao or '').strip().upper()

        no_filter = (not sit) or sit == 'TODOS'

        if no_filter:

            snap = self._get_detalhado_snapshot(uf)

            if snap:

                op_key = 'operadoras_fixo' if tipo == 'fixo' else 'operadoras_movel'

                return {'operadoras': snap.get(op_key, {}), 'source': 'snapshot', 'situacao': 'TODOS'}

            self._maybe_trigger_refresh_async()

            op = self._count_operadoras_fixo(uf=uf) if tipo == 'fixo' else self._count_operadoras_buckets(uf=uf)

            return {'operadoras': op, 'source': 'live', 'situacao': 'TODOS'}

        op = self._count_operadoras_fixo(uf=uf, situacao=sit) if tipo == 'fixo' else self._count_operadoras_buckets(uf=uf, situacao=sit)

        return {'operadoras': op, 'source': 'live', 'situacao': sit}



    def get_detailed_situacao(self, uf=None):

        """Breakdown por situacao_cadastral x tipo de linha. Le snapshot ou fallback live."""

        snap = self._get_detalhado_snapshot(uf)

        if snap and isinstance(snap.get('situacao_breakdown'), list) and snap['situacao_breakdown']:

            return snap['situacao_breakdown']

        # snapshot vazio ou foi gerado antes desta feature -> fallback + agendar refresh

        print(f"[DETALHADO] situacao snapshot vazio pra {uf or 'BR'} -- fallback live + agendando refresh")

        self._maybe_trigger_refresh_async()

        cache_key = f"sit_{(uf or 'BR').upper()}"

        cached = self._detalhado_cache.get(cache_key)

        if cached and (time.time() - cached['ts']) < 600:

            return cached['data']

        data = self._compute_situacao_live(uf=uf)

        self._detalhado_cache[cache_key] = {"ts": time.time(), "data": data}

        return data



    def start_extraction_period(self, filters, output_dir, username=None):

        """Variante do start_extraction com filtro adicional de data_inicio_atividades.

        filters extra: data_de (YYYY-MM-DD), data_ate (YYYY-MM-DD).

        Validação: SP <= 5d, demais UFs <= 10d."""

        from datetime import datetime as _dt

        # UF e obrigatoria pra extracao por periodo (escopo nacional eh muito caro

        # e gera resultados grandes demais — periodo deve ser sempre escopado a 1 estado).

        uf = (filters.get('uf') or '').upper().strip()

        if not uf:

            raise ValueError("UF e obrigatoria para extracao por periodo. Informe o estado (ex: SP, RJ).")

        if len(uf) != 2 or not uf.isalpha():

            raise ValueError(f"UF invalida: '{uf}'. Use 2 letras (ex: SP, RJ).")

        try:

            de = _dt.strptime(str(filters.get('data_de', '')).strip(), '%Y-%m-%d')

            ate = _dt.strptime(str(filters.get('data_ate', '')).strip(), '%Y-%m-%d')

        except Exception:

            raise ValueError("Datas inválidas. Formato esperado: YYYY-MM-DD")

        if ate < de:

            raise ValueError("Data final não pode ser anterior à inicial.")



        max_days = 5 if uf == 'SP' else 10

        delta_days = (ate - de).days

        if delta_days > max_days:

            raise ValueError(f"Intervalo máximo permitido para {uf or 'esta UF'} é de {max_days} dias (recebido: {delta_days}).")



        # Injeta filtros internos consumidos por _run_extraction

        f = dict(filters)

        f['_data_de'] = de.strftime('%Y%m%d')

        f['_data_ate'] = ate.strftime('%Y%m%d')

        f.setdefault('situacao', '02')   # ativas por padrão

        f.setdefault('tipo_tel', 'TODOS')

        f.setdefault('perfil', 'TODOS')



        summary_parts = []

        if uf: summary_parts.append(f"UF: {uf}")

        summary_parts.append(f"Periodo: {de.strftime('%d/%m/%Y')} a {ate.strftime('%d/%m/%Y')}")

        f_summary = " | ".join(summary_parts)



        tid = self._create_task(module="EXTRACTION_PERIOD", username=username, filters=f_summary)

        threading.Thread(target=self._run_extraction, args=(tid, f, output_dir), daemon=True).start()

        return tid



    # ============================================================

    # /DETALHADO

    # ============================================================



    def start_extraction(self, filters, output_dir, username=None):

        summary_parts = []

        if filters.get('uf'): summary_parts.append(f"UF: {filters['uf']}")

        if filters.get('cidade'): summary_parts.append(f"Cidade: {filters['cidade']}")

        if filters.get('cnae'): summary_parts.append(f"CNAE: {filters['cnae']}")

        if filters.get('situacao'): 

            sit_map = {"02": "Ativas", "04": "Baixadas", "08": "Suspensas"}

            summary_parts.append(f"Status: {sit_map.get(filters['situacao'], filters['situacao'])}")

        

        f_summary = " | ".join(summary_parts) if summary_parts else "Extração Completa"

        

        tid = self._create_task(module="EXTRACTION", username=username, filters=f_summary)

        threading.Thread(target=self._run_extraction, args=(tid, filters, output_dir), daemon=True).start()

        return tid



    def _run_extraction(self, tid, filters, output_dir):

        print(f"[DEBUG] Starting _run_extraction for tid: {tid}")

        try:

            self._update_task(tid, status="PROCESSING", progress=1, message="Iniciando motores ClickHouse...")

            print(f"[DEBUG] [_run_extraction] Task {tid} set to PROCESSING")

            status = self.get_task_status(tid)

            if status.get("status") == "CANCELLED":

                print(f"[DEBUG] [_run_extraction] Task {tid} was CANCELLED before start")

                return

            output_file = os_native.path.join(output_dir, f"Extracao_{tid[:8]}.xlsx")

            

            estab_conds = []

            empresas_conds = []

            params = {}

            

            sit = filters.get("situacao", "02")

            if sit != "TODOS":

                estab_conds.append("estab_inner.situacao_cadastral = %(sit)s")

                params['sit'] = sit



            if filters.get("uf"): 

                estab_conds.append("estab_inner.uf = %(uf)s")

                params['uf'] = filters["uf"].strip().upper()

            

            if filters.get("cidade"):

                raw_cidade = str(filters["cidade"]).replace(';', ',')

                cidade_list = [c.strip().upper() for c in raw_cidade.split(',') if c.strip()]

                if cidade_list:

                    if len(cidade_list) == 1:

                        estab_conds.append("m.descricao LIKE %(cid)s")

                        params['cid'] = f"%{cidade_list[0]}%"

                    else:

                        estab_conds.append("m.descricao IN %(cidades)s")

                        params['cidades'] = cidade_list

            

            if filters.get("cnae"): 

                # Suporte a múltiplos CNAEs/Prefixos separados por , ou ;

                raw_cnae = filters["cnae"].replace(';', ',')

                cnae_list = [c.strip() for c in raw_cnae.split(',') if c.strip()]

                

                if cnae_list:

                    cnae_clauses = []

                    for i, c_prefix in enumerate(cnae_list):

                        p_name = f"cnae_pref_{i}"

                        cnae_clauses.append(f"startsWith(estab_inner.cnae_fiscal, %({p_name})s)")

                        params[p_name] = c_prefix

                    

                    estab_conds.append(f"({' OR '.join(cnae_clauses)})")



            # Filtro de Perfil (MEI / NAO MEI) - APLICADO NA TABELA EMPRESAS

            perfil = str(filters.get("perfil", "TODOS")).upper().strip()

            print(f"[DEBUG] PERFIL FILTRO: {perfil}")

            if perfil == "MEI":

                empresas_conds.append("natureza_juridica = '2135'")

            elif perfil == "NAO MEI":

                empresas_conds.append("natureza_juridica != '2135'")



            # Filtro de Órgãos Governamentais

            if filters.get("sem_governo"):

                gov_keywords = [

                    "FEDERAL", "GOVERNO", "PUBLICO", "PUBLICA", "ESTADUAL", "ESTADO", 

                    "MUNICIPIO", "MUNICIPAL", "POLICIA", "BOMBEIRO", "BANCO DO BRASIL", "CORREIOS", 

                    "MINISTERIO", "ADVOCACIA-GERAL", "BANCO CENTRAL", "CASA CIVIL", "CONTROLADORIA-GERAL",

                    "GABINETE DE SEGURANCA", "SECRETARIA"

                ]

                # Filter in companies table (natureza_juridica check is also good, but name is what user asked)

                # Grouping keywords to avoid huge query, using multiSearchAny for performance

                empresas_conds.append("NOT multiSearchAnyCaseInsensitive(razao_social, %(gov_keys)s)")

                params['gov_keys'] = gov_keywords



            # Filtro de periodo (DETALHADO) — data_inicio_atividades em formato YYYYMMDD.

            # Coluna NAO existe em hemn.comercial_pj (gaveta desnormalizada), so em hemn.estabelecimento.

            # Estrategia: monta INNER JOIN com derivada filtrando (data + uf), o que serve simultaneamente

            # pra filtrar registros E trazer a coluna DATA_ABERTURA pro SELECT da planilha.

            period_select = ""

            period_join = ""

            if filters.get("_data_de") and filters.get("_data_ate"):

                sub_uf = ""

                period_uf = (filters.get("uf") or "").strip().upper()

                if period_uf:

                    sub_uf = " AND uf = %(_period_uf)s"

                    params['_period_uf'] = period_uf

                period_join = (

                    "INNER JOIN (SELECT concat(cnpj_basico, cnpj_ordem, cnpj_dv) AS cnpj_full, "

                    "data_inicio_atividades AS dt_abertura "

                    "FROM hemn.estabelecimento "

                    "WHERE data_inicio_atividades >= %(data_de)s "

                    "AND data_inicio_atividades <= %(data_ate)s" + sub_uf + ") "

                    "AS _period ON _period.cnpj_full = estab_inner.cnpj"

                )

                # Formata YYYYMMDD -> DD/MM/YYYY direto no SQL (igual padrao usado em outros pontos do codigo)

                period_select = (",\n                    "

                    "if(length(_period.dt_abertura) = 8, "

                    "concat(substring(_period.dt_abertura, 7, 2), '/', "

                    "substring(_period.dt_abertura, 5, 2), '/', "

                    "substring(_period.dt_abertura, 1, 4)), '') AS DATA_ABERTURA")

                params['data_de'] = filters["_data_de"]

                params['data_ate'] = filters["_data_ate"]



            print(f"[DEBUG] [_run_extraction] filters built: estab_conds={len(estab_conds)}, empresas_conds={len(empresas_conds)}")



            tipo_req = filters.get("tipo_tel", "TODOS")

            if tipo_req == "CELULAR":

                estab_conds.append("((length(estab_inner.telefone1) >= 8 AND substring(estab_inner.telefone1, 1, 1) IN ('6','7','8','9')) OR (length(estab_inner.telefone2) >= 8 AND substring(estab_inner.telefone2, 1, 1) IN ('6','7','8','9')))")

            elif tipo_req == "FIXO":

                estab_conds.append("((length(estab_inner.telefone1) >= 8 AND substring(estab_inner.telefone1, 1, 1) IN ('2','3','4','5')) OR (length(estab_inner.telefone2) >= 8 AND substring(estab_inner.telefone2, 1, 1) IN ('2','3','4','5')))")

            elif tipo_req == "AMBOS":

                estab_conds.append("((length(estab_inner.telefone1) >= 8 AND substring(estab_inner.telefone1, 1, 1) IN ('6','7','8','9') AND length(estab_inner.telefone2) >= 8 AND substring(estab_inner.telefone2, 1, 1) IN ('2','3','4','5')) OR (length(estab_inner.telefone1) >= 8 AND substring(estab_inner.telefone1, 1, 1) IN ('2','3','4','5') AND length(estab_inner.telefone2) >= 8 AND substring(estab_inner.telefone2, 1, 1) IN ('6','7','8','9')))")



            # Filtro de DDD da Região (Independente do Tipo de Telefone)

            if filters.get("filtrar_ddd_regiao") and filters.get("uf"):

                target_uf = filters["uf"].strip().upper()

                ddds = self.UF_DDD_MAP.get(target_uf, [])

                if ddds:

                    estab_conds.append("(estab_inner.ddd1 IN %(ddds_reg)s OR estab_inner.ddd2 IN %(ddds_reg)s)")

                    params['ddds_reg'] = ddds

            elif filters.get("somente_com_telefone"):

                # Fallback legado

                estab_conds.append("(estab_inner.telefone1 != '' OR estab_inner.telefone2 != '')")



            cep_file = filters.get("cep_file")

            cep_df = None

            cep_col = None

            had_num_col = False  # rastreia para emitir aviso se planilha tinha NUMERO

            if cep_file and os_native.path.exists(cep_file):

                self._update_task(tid, progress=2, message="Lendo planilha de filtro de CEPs...")

                try:

                    if cep_file.lower().endswith('.csv'):

                        try:

                            # Tenta com utf-8-sig para ignorar BOM

                            cep_df = pd.read_csv(cep_file, sep=';', dtype=str, on_bad_lines='skip', encoding='utf-8-sig')

                            if len(cep_df.columns) <= 1:

                                cep_df = pd.read_csv(cep_file, sep=',', dtype=str, on_bad_lines='skip', encoding='utf-8-sig')

                        except Exception:

                            cep_df = pd.read_csv(cep_file, sep=None, engine='python', dtype=str, on_bad_lines='skip', encoding='utf-8-sig')

                    else:

                        cep_df = pd.read_excel(cep_file, dtype=str)



                    print(f"[DEBUG] [_run_extraction] CEP file loaded. Columns: {list(cep_df.columns)}")

                    cep_col = next((c for c in cep_df.columns if "CEP" in str(c).upper()), None)

                    had_num_col = any("NUMERO" in str(c).upper().replace('Ú', 'U') for c in cep_df.columns)

                    print(f"[DEBUG] [_run_extraction] Detected: cep_col='{cep_col}', had_num_col={had_num_col}")

                    # FAIL-FAST: usuario subiu planilha mas nenhuma coluna CEP foi encontrada

                    if not cep_col:

                        self._update_task(tid, status="FAILED", message=f"Planilha de CEPs enviada, mas nenhuma coluna contendo 'CEP' foi detectada no cabecalho. Colunas encontradas: {list(cep_df.columns)[:10]}")

                        return

                    # Memoria: manter apenas a coluna CEP — NUMERO ignorado nesta versao

                    cep_df = cep_df[[cep_col]].copy()

                    local_df = cep_df.dropna(subset=[cep_col]).copy()

                    series_cep = local_df[cep_col].astype(str).str.replace(r'\D', '', regex=True)

                    series_cep = series_cep[series_cep != '']

                    series_cep = series_cep.str.zfill(8)

                    valid_ceps = sorted({c for c in series_cep if c and c != '00000000' and len(c) == 8})

                    # CAP: 200k para usuario comum, 2M para staff (rede de seguranca contra erro)

                    user_role = (filters.get('_user_role') or '').upper()

                    is_staff = user_role in ('ADMIN', 'MAYK')

                    cap = 2_000_000 if is_staff else 200_000

                    if len(valid_ceps) > cap:

                        cap_label = f"{cap:,}".replace(',', '.')

                        current_label = f"{len(valid_ceps):,}".replace(',', '.')

                        self._update_task(tid, status="FAILED", message=(

                            f"Sua planilha tem {current_label} CEPs unicos — limite atual e {cap_label}.\n"

                            f"Sugestoes:\n"

                            f"1. Divida a planilha em 2 partes e rode 2 extracoes.\n"

                            f"2. Combine os filtros UF + CNAE para refinar sem precisar de tantos CEPs.\n"

                            f"3. Para extracoes maiores, fale com o suporte."

                        ))

                        return

                    if valid_ceps:

                        estab_conds.append("estab_inner.cep IN %(ceps)s")

                        params['ceps'] = valid_ceps

                        print(f"[INFO] [_run_extraction] CEP filter ativo: {len(valid_ceps)} CEPs unicos.")

                    if had_num_col:

                        print("[INFO] [_run_extraction] Coluna NUMERO detectada na planilha mas IGNORADA — filtro atual e por CEP. Resultado incluira todas as empresas dos CEPs informados.")

                    cep_df = local_df

                except Exception as e:

                    self._update_task(tid, status="FAILED", message=f"Erro ao analisar planilha CEP: {str(e)}")

                    return



            # REESTRUTURAÇÃO DA QUERY (FASE 4 - ULTRA ROBUSTO)

            estab_where = " AND ".join(estab_conds) if estab_conds else "1=1"

            empresas_where = " AND ".join(empresas_conds) if empresas_conds else "1=1"



            # MIGRAÇÃO PARA GAVETA: hemn.comercial_pj é particionada por UF e denormalizada

            # (não precisa de JOIN com empresas/municipio). Resultado: ClickHouse abre só

            # a partição da UF do filtro = 20-30x menos leitura, sem custo de JOIN.

            # Mantemos o alias "estab_inner" no FROM para que estab_conds (que já usa esse

            # prefixo) continue válido sem reescrever o builder de condições.

            # Mapeia "m.descricao" (que existia via JOIN com municipio) para a coluna direta.

            combined_where = estab_where

            if empresas_where != "1=1":

                combined_where = f"({estab_where}) AND ({empresas_where})"

            combined_where = combined_where.replace("m.descricao", "estab_inner.municipio_nome")



            # Toggle "Incluir Nomes dos Sócios" (commit nova função 2026-04)

            incluir_socios = bool(filters.get('incluir_socios'))

            socios_select = ""

            socios_join = ""

            if incluir_socios:

                socios_select = ",\n                    COALESCE(s.socios_lista, '') AS SOCIOS"

                socios_join = """LEFT JOIN (

                    SELECT cnpj_basico, arrayStringConcat(arraySlice(groupArray(nome_socio), 1, 3), ' | ') AS socios_lista

                    FROM hemn.socios

                    GROUP BY cnpj_basico

                ) AS s ON substring(estab_inner.cnpj, 1, 8) = s.cnpj_basico"""



            q = f"""

                SELECT

                    estab_inner.razao_social AS NOME,

                    estab_inner.cnpj AS CNPJ,

                    estab_inner.situacao_cadastral AS SITUACAO,

                    estab_inner.cnae_fiscal AS CNAE,

                    estab_inner.logradouro AS RUA,

                    estab_inner.numero AS NUMERO,

                    '' AS COMPLEMENTO,

                    estab_inner.bairro AS BAIRRO,

                    estab_inner.municipio_nome AS CIDADE,

                    estab_inner.uf AS UF,

                    estab_inner.cep AS CEP,

                    estab_inner.ddd1 AS DDD1,

                    estab_inner.telefone1 AS TEL1,

                    estab_inner.ddd2 AS DDD2,

                    estab_inner.telefone2 AS TEL2{period_select}{socios_select}

                FROM hemn.comercial_pj AS estab_inner

                {period_join}

                {socios_join}

                WHERE {combined_where}

                LIMIT 20000000

            """

            

            # FASE 9: MOTOR DE LOTES HEMN (SINGLE OU MULTI QUERY)

            total_records_final = 0

            header_written = False

            

            import xlsxwriter

            workbook = xlsxwriter.Workbook(output_file, {'constant_memory': True, 'tmpdir': '/tmp'})

            sheet = workbook.add_worksheet("Extracao Hemn")

            header_fmt = workbook.add_format({'bold': True, 'bg_color': '#3a7bd5', 'font_color': 'white'})



            ch_local = self._get_ch_client()



            # Caminho unico: extracao com filtros (incluindo CEP IN %(ceps)s se houver upload).

            # NUMERO foi removido em 2026-04 — vide nota acima sobre 'had_num_col'.

            if True:

                self._update_task(tid, progress=15, message="Executando consulta no ClickHouse...")

                result = ch_local.query(q, params, settings={'query_id': tid})

                rows = result.result_rows

                cols = result.column_names

                total_rows_found = len(rows)

                

                self._update_task(tid, progress=20, message=f"Iniciando processamento de {total_rows_found:,} registros...")



                batch_size = 250000

                operator_task_cache = {} # Cache persistente entre lotes p/ evitar consultas duplicadas de operadora

                for i in range(0, total_rows_found, batch_size):

                    status = self.get_task_status(tid)

                    if status.get("status") == "CANCELLED":

                        workbook.close()

                        return



                    chunk = rows[i:i + batch_size]

                    df = pd.DataFrame(chunk, columns=cols)

                    

                    self._update_task(tid, progress=min(95, round(20 + (i/max(1, total_rows_found) * 75), 1)), 

                                      message=f"Processando: {i:,} / {total_rows_found:,} registros...")



                    if cep_df is not None and '_match_cep' in cep_df.columns:

                        df['_match_cep'] = df['CEP'].astype(str).str.replace(r'\D', '', regex=True).str.zfill(8)

                        valid_sheet_ceps = cep_df[['_match_cep']].drop_duplicates()

                        df = df.merge(valid_sheet_ceps, on='_match_cep', how='inner')

                        df = df.drop(columns=[c for c in ['_match_cep'] if c in df.columns])



                    if df.empty: continue

                    

                    df_processed = self._process_extraction_dataframe(tid, df, filters, workbook, sheet, header_fmt, header_written, total_records_final, operator_cache=operator_task_cache)

                    total_records_final += len(df_processed)

                    header_written = True



            workbook.close()

            # Fim da FASE 9



            self._update_task(tid, status="COMPLETED", progress=100, message=f"Extração Pronta! {total_records_final:,} encontrados.", result_file=output_file, record_count=total_records_final)

        except Exception as e:

            import traceback

            err_msg = traceback.format_exc()

            print(f"[CRITICAL] EXTRACTION THREAD FAILED for {tid}: {e}\n{err_msg}")

            try:

                self._update_task(tid, status="FAILED", message=f"TITANIUM-MT ERROR: {str(e)}")

            except:

                pass



    def _process_extraction_dataframe(self, tid, df, filters, workbook, sheet, header_fmt, header_written, start_row_count, operator_cache=None):

        """

        Sub-motor de processamento de dataframe para extração (VETORIZADO).

        Normaliza telefones, filtra operadoras e escreve no Excel com performance titanium.

        """

        if df.empty: return df



        # 1. Normalização de Telefones (Vetorizada)

        for col in ['DDD1', 'TEL1', 'DDD2', 'TEL2']:

            if col not in df.columns: df[col] = ""

            df[col] = df[col].astype(str).str.replace(r'\.0$', '', regex=True).replace(['nan', 'NaN', 'None', '<NA>'], '')



        # Limpeza e concatenação (T1)

        df['full_t1'] = (df['DDD1'] + df['TEL1']).str.replace(r'\D', '', regex=True)

        # Regra do 9º dígito (Vetorizada)

        mask10_t1 = (df['full_t1'].str.len() == 10) & (df['full_t1'].str[2].isin(['6','7','8','9']))

        df.loc[mask10_t1, 'full_t1'] = df['full_t1'].str[:2] + '9' + df['full_t1'].str[2:]

        

        # Limpeza e concatenação (T2)

        df['full_t2'] = (df['DDD2'] + df['TEL2']).str.replace(r'\D', '', regex=True)

        mask10_t2 = (df['full_t2'].str.len() == 10) & (df['full_t2'].str[2].isin(['6','7','8','9']))

        df.loc[mask10_t2, 'full_t2'] = df['full_t2'].str[:2] + '9' + df['full_t2'].str[2:]



        # Filtro de Junk (Zeros ou Vazios)

        def clean_junk(ser):

            # Transforma em vazio se for só zeros ou se após o DDD (2 primeiros dígitos) for só zeros

            is_all_zero = ser.str.replace('0', '') == ''

            is_local_zero = (ser.str.len() >= 2) & (ser.str.slice(2).str.replace('0', '') == '')

            return np.where(is_all_zero | is_local_zero, "", ser)

        

        df['full_t1'] = clean_junk(df['full_t1'])

        df['full_t2'] = clean_junk(df['full_t2'])



        # Identificação de Região (Pernambuco Fix) - RIGOROSO v2.0

        if filters.get("filtrar_ddd_regiao") and filters.get("uf"):

            target_uf = str(filters.get("uf")).upper().strip()

            # Mapa real de DDDs por UF (Caso self.UF_DDD_MAP falhe)

            ddd_map = self.UF_DDD_MAP if hasattr(self, 'UF_DDD_MAP') else {

                'AC':['68'], 'AL':['82'], 'AP':['96'], 'AM':['92','97'], 'BA':['71','73','74','75','77'],

                'CE':['85','88'], 'DF':['61'], 'ES':['27','28'], 'GO':['62','64'], 'MA':['98','99'],

                'MT':['65','66'], 'MS':['67'], 'MG':['31','32','33','34','35','37','38'],

                'PA':['91','93','94'], 'PB':['83'], 'PR':['41','42','43','44','45', '46'],

                'PE':['81','87'], 'PI':['86','89'], 'RJ':['21','22','24'], 'RN':['84'],

                'RS':['51','53','54','55'], 'RO':['69'], 'RR':['95'], 'SC':['47','48','49'],

                'SP':['11','12','13','14','15','16','17','18','19'], 'SE':['79'], 'TO':['63']

            }

            valid_ddds = [str(d) for d in ddd_map.get(target_uf, [])]

            if valid_ddds:

                df['is_reg1'] = df['DDD1'].astype(str).isin(valid_ddds)

                df['is_reg2'] = df['DDD2'].astype(str).isin(valid_ddds)

            else:

                df['is_reg1'] = True

                df['is_reg2'] = True

        else:

            df['is_reg1'] = True

            df['is_reg2'] = True



        # Identificação de Tipos (Vetorizada)

        def get_tipo_vec(col):

            # O TEL1/TEL2 original (já limpo de .0 e nan)

            clean = df[col].str.replace(r'\D', '', regex=True)

            is_cel = (clean.str.len() == 9) | ((clean.str.len() == 8) & (clean.str.slice(0,1).isin(['6','7','8','9'])))

            return np.where(clean.str.len() > 0, np.where(is_cel, "CELULAR", "FIXO"), None)



        df['t1_tipo'] = get_tipo_vec('TEL1')

        df['t2_tipo'] = get_tipo_vec('TEL2')



        tipo_req = filters.get("tipo_tel", "TODOS")

        if tipo_req == "AMBOS":

            mask = ((df['t1_tipo'] == "CELULAR") & (df['t2_tipo'] == "FIXO")) | ((df['t1_tipo'] == "FIXO") & (df['t2_tipo'] == "CELULAR"))

            df = df[mask].copy()

        elif tipo_req != "TODOS":

            df = df[(df['t1_tipo'] == tipo_req) | (df['t2_tipo'] == tipo_req)].copy()



        if df.empty: return df



        # Seleção do Telefone (Vetorizada com Filtro Regional)

        if tipo_req in ["CELULAR", "FIXO"]:

            # Só seleciona se for do tipo correto E (se filtro on) for da região correta

            mask1 = (df['t1_tipo'] == tipo_req) & (df['is_reg1']) & (df['full_t1'] != "")

            mask2 = (df['t2_tipo'] == tipo_req) & (df['is_reg2']) & (df['full_t2'] != "")

            df['TELEFONE SOLICITADO'] = np.where(mask1, df['full_t1'], 

                                        np.where(mask2, df['full_t2'], ""))

        elif tipo_req == "AMBOS":

            # Para AMBOS, mostramos os dois, mas filtrados por região se necessário

            t1_valid = (df['full_t1'] != "") & (df['is_reg1'])

            t2_valid = (df['full_t2'] != "") & (df['is_reg2'])

            t1_part = np.where(t1_valid, df['t1_tipo'].astype(str) + ": " + df['full_t1'], "")

            t2_part = np.where(t2_valid, df['t2_tipo'].astype(str) + ": " + df['full_t2'], "")

            df['TELEFONE SOLICITADO'] = t1_part + np.where((t1_part != "") & (t2_part != ""), " | ", "") + t2_part

        else: # TODOS

            # Prioridade: 1. Celular Regional, 2. Fixo Regional, 3. Qualquer Celular (se filtro off), 4. Qualquer Fixo (se filtro off)

            c1_reg = (df['t1_tipo'] == "CELULAR") & (df['is_reg1']) & (df['full_t1'] != "")

            c2_reg = (df['t2_tipo'] == "CELULAR") & (df['is_reg2']) & (df['full_t2'] != "")

            f1_reg = (df['t1_tipo'] == "FIXO") & (df['is_reg1']) & (df['full_t1'] != "")

            f2_reg = (df['t2_tipo'] == "FIXO") & (df['is_reg2']) & (df['full_t2'] != "")

            

            df['TELEFONE SOLICITADO'] = np.where(c1_reg, df['full_t1'],

                                        np.where(c2_reg, df['full_t2'],

                                        np.where(f1_reg, df['full_t1'],

                                        np.where(f2_reg, df['full_t2'], ""))))

        

        # Se após a seleção o telefone estiver vazio e o usuário pediu filtragem ou somente com telefone, removemos a linha

        if filters.get("filtrar_ddd_regiao") or filters.get("somente_com_telefone"):

            df = df[df['TELEFONE SOLICITADO'] != ""].copy()

            if df.empty: return df

                                        

        # Drop de colunas auxiliares

        drop_cols = [c for c in ['DDD1', 'TEL1', 'DDD2', 'TEL2', 't1_tipo', 't2_tipo', 'full_t1', 'full_t2'] if c in df.columns]

        df = df.drop(columns=drop_cols)

        

        # Enriquecimento de Operadora (com cache multi-lote)

        df = self._append_operator_column(tid, df, task_cache=operator_cache)



        # Filtros de Operadora (Vetorizados)

        # Trata "", None e "NENHUMA"/"TODAS" como "sem filtro" — frontend pode mandar string vazia

        op_inc = (str(filters.get("operadora_inc") or "TODAS")).upper().strip()

        op_exc = (str(filters.get("operadora_exc") or "NENHUMA")).upper().strip()

        if not op_inc: op_inc = "TODAS"

        if not op_exc: op_exc = "NENHUMA"



        if 'OPERADORA DO TELEFONE' in df.columns:

            if op_exc and op_exc != "NENHUMA":

                # Suporta múltiplas operadoras separadas por ; ou ,

                exc_list = [x.strip() for x in op_exc.replace(';',',').split(',') if x.strip() and x.strip() != "NENHUMA"]

                if exc_list:

                    parts = ["VIVO|TELEFONICA" if x == "VIVO" else re.escape(x) for x in exc_list]

                    pattern = "|".join(parts)

                    mask_exc = df['OPERADORA DO TELEFONE'].str.upper().str.contains(pattern, na=False, regex=True)

                    df = df[~mask_exc]



            if op_inc and op_inc != "TODAS":

                inc_list = [x.strip() for x in op_inc.replace(';',',').split(',') if x.strip() and x.strip() != "TODAS"]

                if inc_list:

                    parts = ["VIVO|TELEFONICA" if x == "VIVO" else re.escape(x) for x in inc_list]

                    pattern = "|".join(parts)

                    mask_inc = df['OPERADORA DO TELEFONE'].str.upper().str.contains(pattern, na=False, regex=True)

                    df = df[mask_inc]



        if df.empty: return df



        # Formatação Final (Paridade Total com o Print)

        # 1. Normalização de nomes técnicos para amigáveis

        final_mapping = {

            'NOME': 'NOME DA EMPRESA', 

            'SITUACAO': 'SITUACAO CADASTRAL', 

            'RUA': 'LOGRADOURO', 

            'NUMERO': 'NUMERO DA FAIXADA',

            'CEP': 'CEP',

            'BAIRRO': 'BAIRRO',

            'CIDADE': 'CIDADE',

            'UF': 'UF',

            'CNAE': 'CNAE',

            'CNPJ': 'CNPJ',

            'TELEFONE SOLICITADO': 'TELEFONE SOLICITADO',

            'OPERADORA DO TELEFONE': 'OPERADORA DO TELEFONE',

            'DATA_ABERTURA': 'DATA DE ABERTURA'

        }

        

        # Filtro de colunas extras e renomeação

        df = df.rename(columns=final_mapping)

        

        sit_map = {'01':'NULA','02':'ATIVA','03':'SUSPENSA','04':'INAPTA','08':'BAIXADA'}

        if 'SITUACAO CADASTRAL' in df.columns:

            df['SITUACAO CADASTRAL'] = df['SITUACAO CADASTRAL'].astype(str).str.zfill(2).map(sit_map).fillna(df['SITUACAO CADASTRAL'])



        # ORDEM EXATA DO PRINT GOLDEN:

        # CNPJ | NOME DA EMPRESA | SITUACAO CADASTRAL | CNAE | LOGRADOURO | NUMERO DA FAIXADA | [COMPLEMENTO se Regional] | BAIRRO | CIDADE | UF | CEP | TELEFONE | OPERADORA

        final_columns = ['CNPJ', 'NOME DA EMPRESA', 'SITUACAO CADASTRAL', 'CNAE', 'LOGRADOURO', 'NUMERO DA FAIXADA']

        

        uf_req = str(filters.get("uf", "")).strip().upper()

        if uf_req in ["DF", "GO", "MT", "MS"]:

            if 'COMPLEMENTO' not in df.columns: df['COMPLEMENTO'] = ""

            final_columns.append('COMPLEMENTO')

            

        final_columns.extend(['BAIRRO', 'CIDADE', 'UF', 'CEP', 'TELEFONE SOLICITADO', 'OPERADORA DO TELEFONE'])

        # Modulo DETALHADO (extracao por periodo): inclui DATA DE ABERTURA no output

        if filters.get('_data_de') and filters.get('_data_ate'):

            if 'DATA DE ABERTURA' not in df.columns: df['DATA DE ABERTURA'] = ""

            final_columns.append('DATA DE ABERTURA')

        # Toggle "Incluir Nomes dos Sócios": se ON, adiciona coluna SÓCIOS no final.

        # Para empresas sem sócios cadastrados (MEI/individual), usa o nome da empresa como fallback.

        if filters.get('incluir_socios') and 'SOCIOS' in df.columns:

            df['SOCIOS'] = df['SOCIOS'].astype(str).replace(['nan','NaN','None','<NA>'], '')

            empty_mask = (df['SOCIOS'].str.strip() == '')

            if empty_mask.any() and 'NOME DA EMPRESA' in df.columns:

                df.loc[empty_mask, 'SOCIOS'] = df.loc[empty_mask, 'NOME DA EMPRESA']

            final_columns.append('SOCIOS')



        # Assegurar que todas existem

        for c in final_columns:

            if c not in df.columns: df[c] = ""

            else: df[c] = df[c].astype(str).replace(['nan', 'NaN', 'None', '<NA>'], "")

        

        df_final = df[final_columns].fillna("")



        # Escrita no Excel (Vetorizada com write_row e Split de Abas)

        EXCEL_LIMIT = 1000000 

        global_row = start_row_count

        

        # Acessar a última worksheet ativa ou criar a primeira

        if not header_written or not workbook.worksheets():

            if not workbook.worksheets():

                active_sheet = workbook.add_worksheet("Extracao_1")

            else:

                active_sheet = workbook.worksheets()[-1]

            

            # ESCREVER CABEÇALHO SEMPRE NA LINHA 0

            active_sheet.write_row(0, 0, df_final.columns, header_fmt)

        else:

            active_sheet = workbook.worksheets()[-1]



        # Escrita em massa

        for r_idx, row_data in enumerate(df_final.values):

            # Verificar se atingiu o limite para trocar de aba

            if global_row > 0 and (global_row % EXCEL_LIMIT) == 0:

                sheet_idx = (global_row // EXCEL_LIMIT) + 1

                active_sheet = workbook.add_worksheet(f"Extracao_{sheet_idx}")

                active_sheet.write_row(0, 0, df_final.columns, header_fmt)

            

            # LINHA 0 É SEMPRE O CABEÇALHO. DADOS COMEÇAM NA LINHA 1.

            row_in_sheet = (global_row % EXCEL_LIMIT) + 1

            active_sheet.write_row(row_in_sheet, 0, row_data)

            global_row += 1



            

        return df_final



    # --- UNIFY ---

    def start_unify(self, file_paths, output_dir, username=None):

        f_summary = f"Unificar {len(file_paths)} arquivos"

        tid = self._create_task(module="UNIFY", username=username, filters=f_summary)

        threading.Thread(target=self._run_unify, args=(tid, file_paths, output_dir), daemon=True).start()

        return tid



    def _run_unify(self, tid, file_paths, output_dir):

        try:

            self._update_task(tid, status="PROCESSING", message="Unificando arquivos...")

            status = self.get_task_status(tid)

            if status.get("status") == "CANCELLED": return

            output_file = os_native.path.join(output_dir, f"Unificado_{tid[:8]}.xlsx")

            dfs = []

            for i, p in enumerate(file_paths):

                if p.endswith('.csv'):

                    dfs.append(pd.read_csv(p, sep=None, engine='python', dtype=str))

                else:

                    dfs.append(pd.read_excel(p, dtype=str))

                self._update_task(tid, progress=int(((i+1)/len(file_paths))*80))

            

            df_final = pd.concat(dfs, ignore_index=True)

            df_final.to_excel(output_file, index=False)

            self._update_task(tid, status="COMPLETED", progress=100, message="Arquivos unificados!", result_file=output_file, record_count=len(df_final))

        except Exception as e:

            self._update_task(tid, status="FAILED", message=f"Erro: {str(e)}")



    # --- CARRIER ---

    def batch_carrier(self, input_file, output_dir, phone_col, username=None):

        tid = self._create_task(module="CARRIER", username=username)

        threading.Thread(target=self._run_carrier, args=(tid, input_file, output_dir, phone_col), daemon=True).start()

        return tid



    def _run_carrier(self, tid, input_file, output_dir, phone_col):

        self._update_task(tid, status="PROCESSING", message="Iniciando Escaneamento Portabilidade-MT...")

        try:

            status = self.get_task_status(tid)

            if status.get("status") == "CANCELLED": return

            output_file = os_native.path.join(output_dir, f"Portabilidade_{tid[:8]}.xlsx")

            

            # Load Data

            if input_file.endswith('.csv'):

                with open(input_file, 'r', encoding='utf-8', errors='ignore') as f:

                    primeira = f.readline()

                sep = ';' if ';' in primeira else (',' if ',' in primeira else '\t')

                df = pd.read_csv(input_file, sep=sep, engine='python', dtype=str)

            else:

                df = pd.read_excel(input_file, dtype=str)

            

            total = len(df)

            self._update_task(tid, progress=5, message=f"Processando {total:,} registros...")



            # Determine Phone Column

            t_col = phone_col if (phone_col and phone_col in df.columns) else df.columns[0]

            

            # Normalize Phones

            df['titanium_tel'] = df[t_col].fillna('').astype(str).str.replace(r'\D', '', regex=True)

            df['titanium_tel'] = df['titanium_tel'].apply(lambda x: x[2:] if x.startswith('55') else x)

            

            phones_to_query = df['titanium_tel'].unique().tolist()

            # Double lookup: try 10 digits if 11 digit fails (some DBs have legacy format)

            phones_to_query_alt = [p[:len(p)-9] + p[len(p)-8:] for p in phones_to_query if len(p) == 11]

            all_queries = list(set(phones_to_query + phones_to_query_alt))



            op_results = {}

            op_map = self._get_carrier_map()

            

            # Batch Query SQLite (Titanium-MT Style)

            conn = sqlite3.connect(self.db_carrier)

            batch_size = 800

            for i in range(0, len(all_queries), batch_size):

                # Check cancellation

                if i % 4000 == 0:

                    status = self.get_task_status(tid)

                    if status.get("status") == "CANCELLED": 

                        conn.close()

                        return

                    self._update_task(tid, progress=int(5 + (i/len(all_queries))*85), message=f"Consultando Operadoras: {i:,}/{len(all_queries):,}...")



                batch = all_queries[i : i + batch_size]

                placeholders = ','.join(['?'] * len(batch))

                rows = conn.execute(f"SELECT telefone, operadora_id FROM portabilidade WHERE telefone IN ({placeholders})", batch).fetchall()

                for tel_db, op_id in rows:

                    op_results[str(tel_db)] = op_map.get(str(op_id), "OUTRA")

            conn.close()



            def _smart_map(t):

                # 1. Try SQLite portabilidade (Real Ported Data)

                res = None

                if t in op_results: res = op_results[t]

                if not res and len(t) == 11:

                    t10 = t[:2] + t[3:]

                    if t10 in op_results: res = op_results[t10]

                if not res:

                    t13 = "55" + t

                    if t13 in op_results: res = op_results[t13]

                

                if res and res != "OUTRA": return res

                

                # 2. Fallback: Prefix Lookup (Non-ported Original Carrier)

                num = re.sub(r'\D', '', t)

                if num.startswith("55"): num = num[2:]

                for length in range(7, 3, -1):

                    pref = num[:length]

                    if pref in self.prefix_tree:

                        return self.get_op_name(self.prefix_tree[pref])

                

                return res if res else "NÃO CONSTA"



            df['OPERADORA'] = df['titanium_tel'].map(_smart_map)

            

            # Cleanup and Save

            df.drop(columns=['titanium_tel'], inplace=True)

            df.to_excel(output_file, index=False)

            

            self._update_task(tid, status="COMPLETED", progress=100, message="Portabilidade Concluída!", result_file=output_file, record_count=total)

        except Exception as e:

            self._update_task(tid, status="FAILED", message=f"Erro: {str(e)}")



    def get_single_carrier(self, phone):

        try:

            conn = sqlite3.connect(self.db_carrier)

            tel = ''.join(filter(str.isdigit, str(phone)))

            if tel.startswith('55'): tel = tel[2:]

            

            # Smart Lookup (11 -> 10 -> 13)

            variations = [tel]

            if len(tel) == 11: variations.append(tel[:2] + tel[3:])

            variations.append("55" + tel)

            

            c = None

            for v in variations:

                c = conn.execute("SELECT operadora_id FROM portabilidade WHERE telefone = ?", (v,)).fetchone()

                if c: break



            # Operadora original do prefixo (ANATEL) - sempre calculada para detectar portabilidade
            num = re.sub(r'\D', '', tel)
            original_op = None
            for length in range(7, 3, -1):
                pref = num[:length]
                if pref in self.prefix_tree:
                    original_op = self.get_op_name(self.prefix_tree[pref])
                    break

            if c:
                op_name = self.get_op_name(c[0])
                # Esta na tabela portabilidade E operadora atual difere da do prefixo => portado
                portado = bool(original_op) and (op_name != original_op)
                res = {"operadora": op_name, "tipo": "Móvel", "portado": portado, "operadora_original": original_op or op_name}
            else:
                op_name = original_op or "NÃO CONSTA"
                res = {"operadora": op_name, "tipo": "Original/Prefix", "portado": False, "operadora_original": op_name}

            

            conn.close()

            return res

        except Exception as e:

            return {"operadora": f"ERRO: {str(e)}", "tipo": "N/D"}



    # --- SPLIT ---

    def start_split(self, input_file, output_dir, username=None):

        fname = os_native.path.basename(input_file)

        f_summary = f"Fatiar: {fname}"

        tid = self._create_task(module="SPLIT", username=username, filters=f_summary)

        threading.Thread(target=self._run_split, args=(tid, input_file, output_dir), daemon=True).start()

        return tid



    def _run_split(self, tid, input_file, output_dir):

        print(f"[SPLIT] Iniciando tarefa {tid} para arquivo: {input_file}")

        self._update_task(tid, status="PROCESSING", message="Iniciando fatiamento...")

        try:

            output_file = os_native.path.join(output_dir, f"Dividido_{tid[:8]}.xlsx")

            import xlsxwriter

            

            # Se for Excel, usamos o modo padrão (Excel gigante é raro, o limite é 1M de linhas)

            # Se for CSV, usamos chunks para evitar estourar a RAM

            is_csv = not input_file.lower().endswith(('.xlsx', '.xls', '.xlsm'))

            

            writer = pd.ExcelWriter(output_file, engine='xlsxwriter', engine_kwargs={'options': {'constant_memory': True}})

            total_rows = 0

            chunk_size = 1000000



            if is_csv:

                self._update_task(tid, message="Lendo CSV em blocos (otimizado)...")

                # Detectar separador e encoding de forma robusta

                sep = ','

                enc = 'utf-8'

                try:

                    with open(input_file, 'rb') as f:

                        raw = f.read(10240) # Aumentado para 10KB para amostragem melhor

                        # Tenta utf-8

                        try:

                            sample = raw.decode('utf-8')

                            enc = 'utf-8'

                        except:

                            sample = raw.decode('latin-1')

                            enc = 'latin-1'

                        

                        delims = [';', ',', '\t', '|']

                        max_count = 0

                        for d in delims:

                            count = sample.count(d)

                            if count > max_count:

                                max_count = count

                                sep = d

                    print(f"[SPLIT] Detecção: sep='{sep}', enc='{enc}'")

                except Exception as e:

                    print(f"[SPLIT] Erro na detecção: {e}")



                total_rows = sum(1 for _ in open(input_file, 'rb'))

                processed_rows = 0

                reader = pd.read_csv(input_file, sep=sep, chunksize=chunk_size, dtype=str, encoding=enc, on_bad_lines='skip', header=None, engine='c')

                

                for i, chunk in enumerate(reader):

                    sheet_name = f"Lote_{i+1}"

                    chunk.to_excel(writer, sheet_name=sheet_name, index=False)

                    processed_rows += len(chunk)

                    prog = min(99, int((processed_rows / (total_rows or 1)) * 100))

                    self._update_task(tid, progress=prog, message=f"Fatiando: {processed_rows:,}/{total_rows:,} linhas...")

                    print(f"[SPLIT] {tid} -> Processadas {processed_rows:,} linhas")

                total_rows = processed_rows

            else:

                self._update_task(tid, message="Lendo Excel gigante...")

                df = pd.read_excel(input_file, dtype=str)

                total_rows = len(df)

                for i in range(0, total_rows, chunk_size):

                    chunk = df.iloc[i : i + chunk_size]

                    sheet_name = f"Lote_{(i//chunk_size)+1}"

                    chunk.to_excel(writer, sheet_name=sheet_name, index=False)

                    self._update_task(tid, progress=int((i/total_rows)*100), message=f"Escrevendo {sheet_name}...")

            

            writer.close()

            print(f"[SPLIT] Sucesso: {output_file} ({total_rows} linhas)")

            self._update_task(tid, status="COMPLETED", progress=100, message=f"Dividido com sucesso! {total_rows:,} linhas.", result_file=output_file, record_count=total_rows)

        except Exception as e:

            print(f"[SPLIT ERROR] {tid}: {str(e)}")

            self._update_task(tid, status="FAILED", message=f"Erro: {str(e)}")



    def _get_carrier_map(self):

        op_map = {

            "55320":"VIVO / TELEFONICA", "55323":"VIVO / TELEFONICA", "55215":"VIVO / TELEFONICA", "55324":"VIVO / TELEFONICA", "55377":"VIVO / TELEFONICA",

            "55321":"CLARO", "55351":"CLARO", "55322":"CLARO", "55371":"CLARO", "55306":"ALGAR",

            "55341":"TIM", "55343":"SERCOMTEL", "55331":"OI", "55312":"ALGAR", "55345":"ALGAR",

            "55112":"ALGAR", "55121":"CLARO", "55123":"TIM", "55131":"OI", "55141":"TIM",

            "55143":"SERCOMTEL", "55204":"SERCOMTEL", "55301":"ARQIA", "55303":"SISTEER",

            "55305":"VIRGIN", "55315":"TELECALL", "55327":"ALGAR", "55348":"TRANSATEL",

            "55367":"CUBIC", "55393":"GLOBALSTAR", "55106":"AMERICADOMÉSTICA", "55307":"AMERICADOMÉSTICA",

            "55211":"BRISANET", "55322":"BRISANET", "55101":"TELECALL", "55191":"IPCORP"

        }

        full_map = {

            "55100":"Interjato", "55102":"Life Telecom", "55103":"DD Telecom", "55104":"DB3", "55105":"Unifique",

            "55107":"Vipnet", "55108":"Osi Telecom", "55109":"Tellig", "55111":"SuperTV", "55113":"Fonar",

            "55116":"Fidelity", "55117":"Transit", "55118":"Spin", "55119":"ZaaZ", "55120":"GTI", "55124":"Megatelecom",

            "55126":"IDT Brasil", "55127":"VIRTUAL", "55128":"Voxbras", "55129":"T-Leste", "55130":"WKVE",

            "55132":"CONVERGIA", "55133":"NWI", "55136":"DSLI", "55137":"Golden Line", "55138":"Tesa", "55139":"Netglobalis",

            "55140":"Hoje Telecom", "55142":"GT Group", "55144":"Sul Internet", "55145":"Cirion", "55147":"British Telecom",

            "55148":"Netserv", "55149":"BBS Options", "55150":"Cambridge", "55151":"Vero S.A.", "55152":"Alares",

            "55153":"GIGA MAIS", "55154":"Viafibra", "55155":"Big Telco", "55156":"Engevox", "55157":"One Telecom",

            "55158":"Voitel", "55161":"Vonex", "55162":"Mundivox", "55163":"Hello Brazil", "55164":"Brasil Tecpar",

            "55166":"Quex", "55167":"DIRECTCALL", "55168":"Tremnet", "55170":"Locaweb", "55171":"YIP", "55172":"BTT",

            "55173":"Brastel", "55174":"TELEFREE", "55176":"Adylnet", "55177":"Opção", "55178":"Nortelpa", "55179":"Option",

            "55180":"Plumium", "55181":"Datora", "55184":"BRFibra", "55185":"UltraNet", "55186":"Teletel", "55187":"EAI",

            "55188":"TVN VOZ", "55189":"CONECTA", "55192":"LIG16", "55193":"Vipway", "55194":"76 Telecom", "55196":"iVATi",

            "55197":"EAD", "55198":"Ligue Telecom", "55199":"Ensite", "55200":"Porto Velho", "55201":"Hit Telecom",

            "55202":"Smartspace", "55203":"Valenet", "55205":"UFINET", "55206":"Desktop", "55208":"Ponto Telecom",

            "55209":"WCS", "55210":"Morango", "55213":"Tri Telecom", "55214":"Orange", "55216":"Ampernet",

            "55217":"TERAPAR", "55218":"Grandi Telecom", "55219":"Redcontrol", "55220":"Net Angra", "55221":"GGNET",

            "55222":"ADP3", "55223":"Citta", "55224":"Algar VGL", "55225":"Conectcor", "55226":"Unitelco", "55227":"Tely",

            "55228":"GO1", "55229":"Tubaron", "55230":"Titania", "55231":"Brasilfone", "55232":"Bignet", "55233":"Superonda",

            "55234":"Córdia", "55235":"BRDigital", "55236":"MG Conecta", "55237":"Goiás Telecom", "55238":"Mhnet",

            "55239":"Vonare", "55240":"Predialnet", "55241":"Nedel", "55242":"Tcheturbo", "55245":"Toque Telecom",

            "55246":"Neo Telecom", "55247":"RED Telecom", "55248":"Itelco", "55249":"Você Telecom", "55250":"LINKTEL",

            "55251":"DBUG", "55252":"BLUE TELECOM", "55253":"ALAWEB", "55254":"Brasrede", "55255":"OXMAN",

            "55256":"Nova Tecnologia", "55258":"Conecta Provedor", "55259":"Sothis", "55260":"Conectv", "55261":"Empire",

            "55262":"Ves Telecom", "55263":"MD Brasil", "55264":"NGT Telecom", "55265":"QNET Telecom", "55266":"8BR",

            "55267":"ORION", "55268":"Nip Telecom", "55269":"Connectronic", "55270":"Mundo Telecom", "55271":"SpeedTravel",

            "55272":"PAK Telecom", "55273":"Process", "55274":"BRAZILIAN", "55275":"High Connect", "55277":"Gente",

            "55278":"Bitcom", "55279":"C3 Telecom", "55280":"NT2 Telecom", "55281":"CPNET", "55282":"Viva Vox",

            "55283":"Rocketnet", "55285":"Alfa Telecom", "55286":"Next Telecom", "55287":"Sulnet", "55288":"Fiber One",

            "55289":"ITANET", "55290":"Superip", "55291":"Link 10", "55292":"Gigalink", "55293":"Vip Telecom",

            "55294":"Iveloz", "55295":"Baldussi", "55296":"Global Lines", "55297":"Softdados", "55298":"Advance",

            "55299":"Ôtima Telecom", "55300":"Bras Nuvem", "55301":"Arqia", "55304":"Terapar", "55306":"Surf Telecom",

            "55308":"Ligue Telecom", "55309":"Unifique", "55310":"Vecto Mobile", "55311":"NLT", "55313":"Emnify",

            "55322":"Brisanet", "55343":"Sercomtel Celular", "55348":"Transatel", "55356":"Airnity", "55360":"Neko Serviços",

            "55364":"1nce", "55370":"Iez Telecom", "55375":"Superchip", "55388":"Connect Iot", "55391":"IPCorp",

            "55393":"Globalstar", "55401":"ACB Fibra", "55404":"Client CO", "55410":"Supranet", "55411":"Tel & Globe",

            "55420":"MNET", "55440":"IBI Telecom", "55460":"AMULTIPHONE", "55470":"Customer First", "55474":"BR DID",

            "55477":"ACESSE COMUNICACAO", "55479":"MAISVOIP", "55480":"Softpollus", "55481":"Redfox Fiber",

            "55482":"TRI D TELECOM", "55483":"Vono Tecnologia", "55484":"ABRATEL", "55485":"RCE IT",

            "55486":"PANDA NETWORK", "55487":"FIVENET", "55488":"GL Fibra", "55489":"Bitwave", "55490":"Nicnet",

            "55491":"TOTAL PHONE", "55492":"NETFLEX", "55493":"INOVA FIBRA", "55494":"Allfiber", "55495":"Rolim Net",

            "55496":"ALOSEC", "55497":"NET WAY", "55498":"Digital Net", "55499":"Seagate", "55512":"Plis Inteligência",

            "55534":"Vox One", "55538":"Lestetronics", "55553":"FB Net", "55557":"Soft System", "55570":"Atelex",

            "55576":"RMNetwork", "55580":"Pronto Fibra", "55584":"ITELECOM", "55585":"Global Fibra", "55589":"Osirnet",

            "55590":"LH TELECOM", "55600":"MEGALINK", "55603":"Cyber Internet", "55605":"Ame Telecom",

            "55614":"L T Soluções", "55618":"T. T. A. Soluções", "55623":"Sigatel", "55629":"Internet 10",

            "55630":"R2 Dados", "55633":"The Fiber", "55634":"Wn Telecom", "55639":"Niqturbo", "55640":"Nova NET",

            "55645":"JR Telecom", "55651":"Erictel", "55653":"Cn Telecom", "55654":"Lastnet", "55655":"Mega Net",

            "55660":"Viatec", "55662":"TR Telecom", "55665":"Indanet", "55668":"Brlognet", "55673":"UBA Conect",

            "55677":"BrasilNet", "55678":"Contato Internet", "55679":"Servstar", "55682":"Turbonet", "55684":"Gr@mnet",

            "55688":"Signet", "55694":"Handix", "55698":"Sati Telecom", "55701":"Drcall", "55703":"Innova",

            "55704":"Vale Vox", "55707":"G2G Internet", "55712":"Rio Grande", "55722":"GlobalNet", "55724":"New Voice",

            "55733":"Internet Play", "55746":"Onlink", "55751":"Upix Networks", "55761":"Conecte Telecom",

            "55764":"He Net", "55777":"METODO TELECOM", "55782":"Sonik", "55790":"WGO Multimidia", "55791":"4NET",

            "55820":"NPX", "55834":"Desktop Internet", "55853":"Assim", "55880":"Coelho Tecnologia",

            "55889":"Conectlan", "55890":"Btelway", "55916":"Multiware", "55947":"Brphonia", "55974":"Serra Geral",

            "55975":"Vulcanet", "55984":"FTTH Telecom"

        }

        op_map.update(full_map)

        

        # MERGE WITH DYNAMICALLY LOADED ANATEL DICT (From CSV)

        if hasattr(self, 'anatel_dict'):

            op_map.update(self.anatel_dict)

            

        # NORMALIZAÇÃO HEMN PARA FILTROS ROBUSTOS

        normalized = {}

        for k, v in op_map.items():

            vu = str(v).upper().strip()

            # Identificação Específica (evita que "VOICE/VOIP" seja detectado como "OI")

            if "TELEFONICA" in vu or "VIVO" in vu: normalized[k] = "VIVO / TELEFONICA"

            elif "CLARO" in vu: normalized[k] = "CLARO"

            elif "TIM" in vu: normalized[k] = "TIM"

            elif vu == "OI" or vu.startswith("OI ") or "OI S.A" in vu or "OI FIXO" in vu: normalized[k] = "OI / TELEMAR"

            else: normalized[k] = vu

        return normalized



    def _append_operator_column(self, tid, df, task_cache=None):

        """

        Enriquecimento de operadora de ALTA PERFORMANCE.

        Usa cache em lote (task_cache: dict opcional persistido entre lotes da extração).

        """

        if 'TELEFONE SOLICITADO' not in df.columns or df.empty:

            df['OPERADORA DO TELEFONE'] = "NÃO CONSTA"

            return df



        self._update_task(tid, progress=90, message="Identificando operadoras...")

        try:

            # 1. Limpeza Vetorizada de Telefones

            df['_clean_tel_enrich'] = df['TELEFONE SOLICITADO'].astype(str).str.replace(r'\D', '', regex=True).str.replace(r'^55', '', regex=True)

            

            unique_phones = df['_clean_tel_enrich'].unique()

            phones_to_query = [p for p in unique_phones if p and p != 'nan']



            # Filtro Preventivo de Cache: skipa o que já foi resolvido em lote anterior

            if task_cache:

                phones_to_query = [p for p in phones_to_query if p not in task_cache]



            # Gerar variações (11 -> 10 para portabilidade legada)

            all_queries = set(phones_to_query)

            for p in phones_to_query:

                if len(p) == 11: all_queries.add(p[:2] + p[3:]) # DDD + 8 dígitos



            # 2. Consulta SQLite em Lotes (Cache de Resultados)

            op_results = task_cache if task_cache is not None else {}

            op_map = self._get_carrier_map()

            

            all_queries_list = list(all_queries)

            batch_size = 1000  # Aumentado para melhor performance

            total_queries = len(all_queries_list)

            

            # OTIMIZAÇÃO CRÍTICA: Abrir conexão UMA VEZ

            conn = sqlite3.connect(self.db_carrier)

            cursor = conn.cursor()

            

            for i in range(0, total_queries, batch_size):

                # Feedback de progresso para evitar a percepção de travamento em 90%

                if i % 5000 == 0 or i == 0:

                    p_val = 90 + int((i / max(1, total_queries)) * 9)

                    self._update_task(tid, progress=p_val, message=f"Consultando Operadoras: {i:,}/{total_queries:,}...")



                batch = all_queries_list[i : i + batch_size]

                placeholders = ','.join(['?'] * len(batch))

                rows = cursor.execute(f"SELECT telefone, operadora_id FROM portabilidade WHERE telefone IN ({placeholders})", batch).fetchall()

                for tel_db, op_id in rows:

                    op_val = op_map.get(str(op_id), "OUTRA")

                    op_results[str(tel_db)] = op_val

                    if task_cache is not None:

                        task_cache[str(tel_db)] = op_val

            

            conn.close()



            # 3. Mapeamento Inteligente e Fallback de Prefixo (Híbrido Vetorizado)

            # Primeiro nível: Mapeamento direto da Portabilidade

            df['OPERADORA DO TELEFONE'] = df['_clean_tel_enrich'].map(op_results).astype('object')

            

            # Segundo nível: Portabilidade com 10 dígitos (se o de 11 falhou)

            mask_retry_10 = (df['OPERADORA DO TELEFONE'].isna()) & (df['_clean_tel_enrich'].str.len() == 11)

            if mask_retry_10.any():

                df.loc[mask_retry_10, 'OPERADORA DO TELEFONE'] = (df.loc[mask_retry_10, '_clean_tel_enrich'].str[:2] + df.loc[mask_retry_10, '_clean_tel_enrich'].str[3:]).map(op_results)



            # Terceiro nível: Prefixo Anatel (Fallback)

            mask_prefix = df['OPERADORA DO TELEFONE'].isna() | (df['OPERADORA DO TELEFONE'] == "OUTRA")

            if mask_prefix.any():

                # Loop limitado por comprimentos de prefixo (mais rápido que loop por linha)

                prefix_results = pd.Series(index=df.index, dtype=str)

                nums = df.loc[mask_prefix, '_clean_tel_enrich']

                

                # De 7 a 4 dígitos

                for length in [7, 6, 5, 4]:

                    if nums.empty: break

                    prefs = nums.str[:length]

                    # Mapear prefixos usando o prefix_tree em memória

                    mapped = prefs.map(self.prefix_tree).dropna()

                    if not mapped.empty:

                        # Converter códigos Anatel para nomes reais

                        unique_codes = mapped.unique()

                        code_to_name = {c: self.get_op_name(c).upper() for c in unique_codes}

                        resolved = mapped.map(code_to_name)

                        prefix_results.update(resolved)

                        # Remover já encontrados para não sobrepor com prefixos menores

                        nums = nums.drop(index=mapped.index)

                

                df.loc[mask_prefix, 'OPERADORA DO TELEFONE'] = prefix_results.fillna(df.loc[mask_prefix, 'OPERADORA DO TELEFONE'])



            df['OPERADORA DO TELEFONE'] = df['OPERADORA DO TELEFONE'].fillna("NÃO CONSTA")

            df = df.drop(columns=['_clean_tel_enrich'])

            return df

        except Exception as e:

            import traceback

            print(f"Erro no filtro de operadora: {e}")

            traceback.print_exc()

            df['OPERADORA DO TELEFONE'] = "ERRO"

            return df

